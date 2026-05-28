# === IMPORTS ===
import dearpygui.dearpygui as dpg
import pandapower as pp
import pandas as pd
import numpy as np
import threading
import traceback
import socketserver
import webbrowser
import base64
import os
import sys
from collections import defaultdict
import html as _hl
import folium
import math

# === PATH AND DIRECTORY MANAGEMENT ===
def resource_path(relative_path):
    if getattr(sys, 'frozen', False):
        base = sys._MEIPASS
    else:
        base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base, relative_path)

def output_path(relative_path):
    if getattr(sys, 'frozen', False):
        base = os.path.dirname(sys.executable)
    else:
        base = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base, relative_path)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# === GRID CONFIGURATION AND CONSTANTS ===
SUBGRIDS         = ["LEYTE-SAMAR", "CEBU", "NEGROS", "BOHOL", "PANAY"]
SLACK_SUBGRID    = "LEYTE-SAMAR"
USE_DEP_CAPACITY = True
MAP_PORT         = 8766
MAP_FILE         = "combined_visayas_map.html"
SLD_FILE         = "combined_visayas_sld.html"

V_KV     = 230.0
V_KV_138 = 138.0

# === HVDC AND TIE LINE DEFINITIONS ===
def mva_to_i_ka(mva, v_kv=V_KV):
    return mva / (math.sqrt(3) * v_kv)

def mva_to_i_ka_138(mva):
    return mva / (math.sqrt(3) * V_KV_138)

HVDC_DEFAULTS = [
    {"name":"LVHVDC","bus":"ORMOC 1","subgrid":"LEYTE-SAMAR",
     "min_p_mw":-250.0,"max_p_mw":250.0,"vm_pu":1.0,"cost":2775.6396,"is_slack":True},
    {"name":"VMHVDC","bus":"DUMANJUG","subgrid":"CEBU",
     "min_p_mw":-140.63,"max_p_mw":140.63,"vm_pu":1.0,"cost":3011.1459,"is_slack":False},
]

def load_hvdc_connections():
    hvdc_path = os.path.join(period_base_dir(), "hvdc.csv")
    if os.path.exists(hvdc_path):
        try:
            df = pd.read_csv(hvdc_path)
            df.columns = df.columns.str.strip()
            conns = []
            for _, row in df.iterrows():
                conns.append({
                    "name":     str(row["name"]),
                    "bus":      str(row["bus"]),
                    "subgrid":  str(row["subgrid"]),
                    "min_p_mw": float(row["min_p_mw"]),
                    "max_p_mw": float(row["max_p_mw"]),
                    "vm_pu":    float(row["vm_pu"]),
                    "cost":     float(row["cost"]),
                    "is_slack": str(row["is_slack"]).strip().lower() in ("true","1","yes"),
                })
            return conns
        except Exception:
            pass
    return [dict(h) for h in HVDC_DEFAULTS]

HVDC_CONNECTIONS = HVDC_DEFAULTS

TIE_BUS_SUBGRID = {
    "BAROTAC VIEJO 1": "PANAY",
    "GAHIT CTS":       "NEGROS",
    "CALATRAVA CTS":   "NEGROS",
    "TALAVERA CTS":    "CEBU",
    "DAAN BANTAYAN":   "CEBU",
    "TABANGO 1":       "LEYTE-SAMAR",
    "ARGAO CTS":       "CEBU",
    "MARIBOJOC CTS":   "BOHOL",
    "MAASIN 1":        "LEYTE-SAMAR",
    "UBAY":            "BOHOL",
    "SAMBOAN":         "CEBU",
    "AMLAN 1":         "NEGROS",
}

TIE_LINES = [
    ("BAROTAC VIEJO 1", "GAHIT CTS",     18.5,  "tie_PANAY_NEGROS_1",  mva_to_i_ka(444)),
    ("BAROTAC VIEJO 1", "GAHIT CTS",     18.5,  "tie_PANAY_NEGROS_2",  mva_to_i_ka(444)),
    ("DAAN BANTAYAN",   "TABANGO 1",     31.22, "tie_CEBU_LEYTE_1",    mva_to_i_ka(200)),
    ("DAAN BANTAYAN",   "TABANGO 1",     31.22, "tie_CEBU_LEYTE_2",    mva_to_i_ka(240)),
    ("MAASIN 1",        "UBAY",          65.0,  "tie_LEYTE_BOHOL_1",   mva_to_i_ka(108)),
    ("ARGAO CTS",       "MARIBOJOC CTS", 22.0,  "tie_CEBU_BOHOL_1",    mva_to_i_ka_138(666)),
    ("ARGAO CTS",       "MARIBOJOC CTS", 22.0,  "tie_CEBU_BOHOL_2",    mva_to_i_ka_138(666)),
    ("SAMBOAN",         "AMLAN 1",       15.0,  "tie_CEBU_NEGROS_1",   mva_to_i_ka_138(108)),
    ("SAMBOAN",         "AMLAN 1",       15.0,  "tie_CEBU_NEGROS_2",   mva_to_i_ka_138(108)),
    ("CALATRAVA CTS",   "TALAVERA CTS",  26.17, "tie_NEGROS_CEBU_1",   mva_to_i_ka(444)),
    ("CALATRAVA CTS",   "TALAVERA CTS",  26.17, "tie_NEGROS_CEBU_2",   mva_to_i_ka(444)),
]

TIE_R_OHM_PER_KM = 0.0277
TIE_X_OHM_PER_KM = 0.1200
TIE_C_NF_PER_KM  = 110.0

SUBGRID_BOUNDS = {
    "LEYTE-SAMAR": {"lat":(9.90,12.70),  "lng":(124.20,125.80)},
    "CEBU":        {"lat":(9.40,11.30),  "lng":(123.25,124.10)},
    "NEGROS":      {"lat":(9.05,11.05),  "lng":(122.30,123.50)},
    "BOHOL":       {"lat":(9.55, 9.80),  "lng":(123.80,124.20)},
    "PANAY":       {"lat":(10.35,11.95), "lng":(121.90,123.15)},
}
BUS_COLORS = {"slack":"red","pv":"green","pq":"blue"}
BUS_RADIUS  = {"slack":13,"pv":10,"pq":8}

# === USER CONSTRAINTS MANAGEMENT ===
CONSTRAINT_DEFAULTS = {
    "vm_min_pu":            0.95,
    "vm_max_pu":            1.05,
    "max_line_loading":     100.0,
    "max_trafo_loading":    100.0,
    "nr_max_iteration":     50,
    "nr_tolerance_mva":     1e-6,
    "opf_max_iteration":    100,
    "opf_delta":            1e-10,
    "tie_r_ohm_per_km":     TIE_R_OHM_PER_KM,
    "tie_x_ohm_per_km":     TIE_X_OHM_PER_KM,
    "tie_c_nf_per_km":      TIE_C_NF_PER_KM,
    "offset_step_m":        60.0,
}

user_constraints: dict = dict(CONSTRAINT_DEFAULTS)

def get_constraint(key):
    return user_constraints.get(key, CONSTRAINT_DEFAULTS[key])

# === OPENSTREETMAP (OSM) DATA PROCESSING ===
def _load_osm_vertices():
    _vp = output_path('visayas_line_vertices.csv')
    if not os.path.exists(_vp):
        _vp = resource_path(os.path.join('visualization', 'visayas_line_vertices.csv'))
    if not os.path.exists(_vp):
        return {}, {}, []
    try:
        df = pd.read_csv(_vp)
        df.columns = df.columns.str.strip()
        if "network_name" not in df.columns:
            df["network_name"] = ""
        df["network_name"] = df["network_name"].fillna("")

        lines = {}
        for osm_id, grp in df.groupby("osm_id"):
            grp = grp.sort_values("sequence")
            coords = [[float(r["lat"]), float(r["lon"])] for _, r in grp.iterrows()]
            if len(coords) < 2:
                continue
            lines[str(osm_id)] = {
                "name":         str(grp["name"].iloc[0]),
                "network_name": str(grp["network_name"].iloc[0]).strip(),
                "kv":           int(grp["kv"].iloc[0]),
                "c":            coords,
            }

        seg_groups = defaultdict(list)
        for oid, info in lines.items():
            nn = info["network_name"].lower()
            if nn:
                seg_groups[nn].append(info["c"])

        name_index = {}
        for nn, segs in seg_groups.items():
            if len(segs) == 1:
                name_index[nn] = segs[0]
            else:
                segs_sorted = sorted(segs, key=lambda s: s[0][0], reverse=True)
                merged = []
                for s in segs_sorted:
                    if merged and merged[-1] == s[0]:
                        merged.extend(s[1:])
                    else:
                        merged.extend(s)
                name_index[nn] = merged

        rtree = [
            (oid, info["c"][0][0], info["c"][0][1],
                  info["c"][-1][0], info["c"][-1][1], info["kv"])
            for oid, info in lines.items()
        ]
        return lines, name_index, rtree
    except Exception:
        return {}, {}, []

OSM_LINES, OSM_NAME_INDEX, OSM_RTREE = _load_osm_vertices()

def reload_osm_vertices():
    global OSM_LINES, OSM_NAME_INDEX, OSM_RTREE
    OSM_LINES, OSM_NAME_INDEX, OSM_RTREE = _load_osm_vertices()

def find_osm_coords(line_name, la0, ln0, la1, ln1):
    if not line_name:
        return None
    coords = OSM_NAME_INDEX.get(line_name.strip().lower())
    if coords is None:
        return None
    path = list(coords)
    def d2(a,b,c,e): return (a-c)**2+(b-e)**2
    if d2(la0,ln0,*path[-1]) < d2(la0,ln0,*path[0]):
        path.reverse()
    path[0]  = [la0, ln0]
    path[-1] = [la1, ln1]
    return path

# === STATE AND APPLICATION DATA MANAGEMENT ===
state = {
    "running":        False,
    "done":           False,
    "error":          None,
    "net":            None,
    "bus_map":        None,
    "subgrid_nets":   None,
    "status":         "Idle -- click Run to start",
    "map_ready":      False,
    "map_server":     None,
    "base_dir":       output_path(''),
    "last_run_label": "",
}

filter_state = {"subgrid": "All"}

period_state = {"period": "Base"}

BASE_PERIOD_FOLDER = "202512221000"

def _folder_to_label(name):
    try:
        from datetime import datetime
        dt = datetime.strptime(name, "%Y%m%d%H%M")
        return dt.strftime("%b %d, %Y %H:%M")
    except Exception:
        return name

_period_label_to_folder = {}

def discover_periods():
    global _period_label_to_folder
    script_dir = os.path.dirname(os.path.abspath(__file__))
    parent_dir = os.path.dirname(script_dir)
    found = set()

    exclude_names = {'dist', 'build', 'VisayasGrid', '__pycache__', 'visualization'} | set(SUBGRIDS)

    for search_dir in [script_dir, parent_dir]:
        try:
            for entry in os.scandir(search_dir):
                if not entry.is_dir(): continue
                if entry.path == script_dir: continue
                if entry.name in exclude_names: continue
                if any(os.path.isdir(os.path.join(entry.path, sg)) for sg in SUBGRIDS):
                    found.add(entry.name)
        except Exception:
            pass

    _period_label_to_folder = {}
    base_label = _folder_to_label(BASE_PERIOD_FOLDER)
    _period_label_to_folder[base_label] = "Base"
    for f in sorted(found):
        lbl = _folder_to_label(f)
        _period_label_to_folder[lbl] = f
    return [base_label] + [_folder_to_label(f) for f in sorted(found)]

def period_base_dir():
    label  = period_state["period"]
    folder = _period_label_to_folder.get(label, "Base")
    if folder == "Base":
        return resource_path('visualization')
    return resource_path(folder)
    script_dir = BASE_DIR
    inside  = os.path.join(script_dir, folder)
    if os.path.isdir(inside): return inside
    sibling = os.path.join(os.path.dirname(script_dir), folder)
    if os.path.isdir(sibling): return sibling
    return script_dir

def refresh_period_list():
    periods = discover_periods()
    if dpg.does_item_exist("period_combo"):
        dpg.configure_item("period_combo", items=periods)
        default_label = periods[0] if periods else "Base"
        if period_state["period"] not in periods:
            period_state["period"] = default_label
            dpg.set_value("period_combo", default_label)
        elif not dpg.get_value("period_combo"):
            dpg.set_value("period_combo", period_state["period"])

UPLOADABLE = [
    "gens.csv", "loads.csv", "shunts.csv",
    "gen_costs.csv",
    "lines.csv", "transformers.csv", "std_types.csv",
    "buses.csv", "geo_coords.csv",
]
UPLOADABLE_OPTIONAL = []
ALL_UPLOADABLE = UPLOADABLE + UPLOADABLE_OPTIONAL

GLOBAL_UPLOADABLE = ["hvdc.csv", "visayas_line_vertices.csv"]

SUBGRIDS_WITH_GLOBAL = SUBGRIDS + ["VISAYAS"]
upload_state = {
    "subgrid":    SUBGRIDS[0],
    "component":  "gens.csv",
    "status":     "",
    "preview_df": None,
}

mem_overrides: dict = {}
outage_state: set = set()

comparison = {
    "baseline_net":   None,
    "modified_net":   None,
    "has_baseline":   False,
    "has_modified":   False,
    "baseline_label": "",
    "modified_label": "",
}

# === PANDAPOWER NETWORK INITIALIZATION AND MERGING ===
def load_subgrid(folder):
    base = os.path.join(resource_path('visualization'), folder)
    def csv(name):
        key = (folder, name)
        if key in mem_overrides:
            return mem_overrides[key].copy()

        label = period_state["period"]
        raw_folder = _period_label_to_folder.get(label, "Base")

        if raw_folder != "Base":
            period_path = os.path.join(resource_path(raw_folder), folder, name)
            if os.path.exists(period_path):
                return pd.read_csv(period_path)

        base_path = os.path.join(base, name)
        return pd.read_csv(base_path)
    net = pp.create_empty_network()
    std_types_df=csv("std_types.csv")
    if "vkr_percent.1" in std_types_df.columns:
        std_types_df = std_types_df.drop(columns=["vkr_percent.1"])
    buses_df=csv("buses.csv")
    transformers_df=csv("transformers.csv"); loads_df=csv("loads.csv")
    gens_df=csv("gens.csv"); shunts_df=csv("shunts.csv")
    lines_df=csv("lines.csv"); gen_costs_df=csv("gen_costs.csv")
    _period = period_state["period"]
    _period_base = os.path.join(state["base_dir"], _period, folder) if _period != "Base" else base
    HAS_3W=(
        (os.path.exists(os.path.join(_period_base,"std_types_3w.csv")) or
         os.path.exists(os.path.join(base,"std_types_3w.csv"))) and
        (os.path.exists(os.path.join(_period_base,"transformers_3w.csv")) or
         os.path.exists(os.path.join(base,"transformers_3w.csv")))
    )
    for _,row in std_types_df.iterrows():
        pp.create_std_type(net,{"sn_mva":row["sn_mva"],"vn_hv_kv":row["vn_hv_kv"],
            "vn_lv_kv":row["vn_lv_kv"],"vk_percent":row["vk_percent"],
            "vkr_percent":row["vkr_percent"],"pfe_kw":row["pfe_kw"],
            "i0_percent":row["i0_percent"],"shift_degree":row["shift_degree"]},
            name=row["name"],element="trafo")
    if HAS_3W:
        _std3w = csv("std_types_3w.csv")
        if "vkr_percent.1" in _std3w.columns:
            _std3w = _std3w.drop(columns=["vkr_percent.1"])
        for _,row in _std3w.iterrows():
            pp.create_std_type(net,{k:row[k] for k in [
                "sn_hv_mva","sn_mv_mva","sn_lv_mva","vn_hv_kv","vn_mv_kv","vn_lv_kv",
                "vk_hv_percent","vk_mv_percent","vk_lv_percent","vkr_hv_percent",
                "vkr_mv_percent","vkr_lv_percent","pfe_kw","i0_percent",
                "shift_mv_degree","shift_lv_degree"]},name=row["name"],element="trafo3w")
    for _,row in buses_df.iterrows():
        pp.create_bus(net,vn_kv=row["vn_kv"],name=row["name"])
    bus_index={name:idx for idx,name in net.bus["name"].items()}
    net.bus["min_vm_pu"] = get_constraint("vm_min_pu")
    net.bus["max_vm_pu"] = get_constraint("vm_max_pu")
    for _,row in transformers_df.iterrows():
        pp.create_transformer(net,hv_bus=bus_index[row["hv_bus"]],
            lv_bus=bus_index[row["lv_bus"]],std_type=row["std_type"],name=row["name"])
    net.trafo["max_loading_percent"] = get_constraint("max_trafo_loading")
    if HAS_3W:
        for _,row in csv("transformers_3w.csv").iterrows():
            pp.create_transformer3w(net,hv_bus=bus_index[row["hv_bus"]],
                mv_bus=bus_index[row["mv_bus"]],lv_bus=bus_index[row["lv_bus"]],
                std_type=row["std_type"],name=row["name"])
    for _,row in loads_df.iterrows():
        pp.create_load(net,bus=bus_index[row["bus"]],p_mw=row["p_mw"],name=row["name"])
    for _,row in gens_df.iterrows():
        pp.create_gen(net,bus=bus_index[row["bus"]],p_mw=row["p_mw"],name=row["name"])
    for _,row in shunts_df.iterrows():
        pp.create_shunt(net,bus=bus_index[row["bus"]],q_mvar=row["q_mvar"],name=row["name"])
    for _,row in lines_df.iterrows():
        pp.create_line_from_parameters(net,from_bus=bus_index[row["from_bus"]],
            to_bus=bus_index[row["to_bus"]],length_km=row["length_km"],
            r_ohm_per_km=row["r_ohm_per_km"],x_ohm_per_km=row["x_ohm_per_km"],
            c_nf_per_km=row["c_nf_per_km"],max_i_ka=row["max_i_ka"],name=row["name"])
    net.line["max_loading_percent"]  = get_constraint("max_line_loading")
    return net, gen_costs_df


def merge_networks(subgrid_nets):
    combined=pp.create_empty_network(); bus_map={}
    for folder,(net,_) in subgrid_nets.items():
        for elem in ["trafo","trafo3w"]:
            for tn,td in net.std_types[elem].items():
                if tn not in combined.std_types[elem]:
                    pp.create_std_type(combined,td,name=tn,element=elem)
    for folder,(net,_) in subgrid_nets.items():
        for _,row in net.bus.iterrows():
            ni=pp.create_bus(combined,vn_kv=row["vn_kv"],name=row["name"])
            combined.bus.loc[ni, "min_vm_pu"] = row.get("min_vm_pu", get_constraint("vm_min_pu"))
            combined.bus.loc[ni, "max_vm_pu"] = row.get("max_vm_pu", get_constraint("vm_max_pu"))
            bus_map[(folder,row["name"])]=ni
    def bmap(f,n):
        if (f,n) not in bus_map:
            raise KeyError(f"Bus '{n}' not found in subgrid '{f}'.")
        return bus_map[(f,n)]
    for folder,(net,_) in subgrid_nets.items():
        ob={idx:n for idx,n in net.bus["name"].items()}
        for _,row in net.trafo.iterrows():
            pp.create_transformer(combined,hv_bus=bmap(folder,ob[row["hv_bus"]]),
                lv_bus=bmap(folder,ob[row["lv_bus"]]),std_type=row["std_type"],name=row["name"])
        if not net.trafo3w.empty:
            for _,row in net.trafo3w.iterrows():
                pp.create_transformer3w(combined,hv_bus=bmap(folder,ob[row["hv_bus"]]),
                    mv_bus=bmap(folder,ob[row["mv_bus"]]),lv_bus=bmap(folder,ob[row["lv_bus"]]),
                    std_type=row["std_type"],name=row["name"])
        for _,row in net.load.iterrows():
            pp.create_load(combined,bus=bmap(folder,ob[row["bus"]]),p_mw=row["p_mw"],name=row["name"])
        for _,row in net.gen.iterrows():
            pp.create_gen(combined,bus=bmap(folder,ob[row["bus"]]),p_mw=row["p_mw"],name=row["name"])
        for _,row in net.shunt.iterrows():
            pp.create_shunt(combined,bus=bmap(folder,ob[row["bus"]]),q_mvar=row["q_mvar"],name=row["name"])
        for _,row in net.line.iterrows():
            pp.create_line_from_parameters(combined,
                from_bus=bmap(folder,ob[row["from_bus"]]),
                to_bus=bmap(folder,ob[row["to_bus"]]),
                length_km=row["length_km"],r_ohm_per_km=row["r_ohm_per_km"],
                x_ohm_per_km=row["x_ohm_per_km"],c_nf_per_km=row["c_nf_per_km"],
                max_i_ka=row["max_i_ka"],name=row["name"])
    if not combined.trafo.empty:
        combined.trafo["max_loading_percent"] = get_constraint("max_trafo_loading")
    combined.line["max_loading_percent"] = get_constraint("max_line_loading")
    return combined,bus_map

# === SINGLE LINE DIAGRAM (SLD) UI CONSTANTS AND HELPERS ===
import html as _sld_hl
from collections import deque

_SLD_COL = {
    "PANAY":        (   20,  940, "#fb923c", "#fb923c"),
    "NEGROS":       (  960, 1920, "#f87171", "#f87171"),
    "CEBU":         ( 1940, 2880, "#60a5fa", "#60a5fa"),
    "LEYTE-SAMAR":  ( 2900, 3880, "#f472b6", "#f472b6"),
    "BOHOL":        ( 3900, 4620, "#a78bfa", "#a78bfa"),
}
_SLD_LEGEND_X = 4640

_SLD_VN_COLOR2 = {
    230:"#f87171", 138:"#fb923c", 115:"#fb923c",
     69:"#38bdf8",  35:"#a78bfa",  34:"#a78bfa",  33:"#a78bfa",
     18:"#4ade80",  14:"#e879f9",  13:"#e879f9",
      7:"#fbbf24",   6:"#fbbf24",   4:"#f9a8d4",
}
_SLD_VN_SW2 = {
    230:5, 138:4, 115:4, 69:3,
     35:2.5, 34:2.5, 33:2.5, 18:2.5, 14:2.5, 13:2.5,
      7:2, 6:2, 4:2,
}

_CHAR_W         = 6.2
_MIN_HALF       = 50
_SUB_PAD_H      = 14
_SUB_PAD_TOP    = 26
_SUB_PAD_BOT    = 10
_BUS_LABEL_H    = 14
_BUS_KV_H       = 12
_BUS_ROW_GAP    = 20
_GEN_R          = 11
_GEN_STEM       = 18
_GEN_LABEL_H    = 18
_LOAD_HW        = 10
_LOAD_H         = 16
_LOAD_STEM      = 12
_LOAD_LABEL_H   = 10
_TRAFO_R        = 9
_TRAFO_GAP      = 3
_TRAFO_STEM     = 8
_SUB_GAP_V      = 28
_COL_INNER_PAD  = 12
_COL_GAP_H      = 12

def _e(s):
    return _sld_hl.escape(str(s))

def _nearest(vn, table):
    vn_r = round(float(vn))
    return table.get(vn_r, table[min(table, key=lambda k: abs(k - vn_r))])

def _load_col(pct):
    if pct > 90: return "#ef4444"
    if pct > 70: return "#f97316"
    return "#22d3ee"

def _substation_name(bus_name):
    import re
    m = re.match(r'^(.+?)\s+\d+[A-Za-z]?$', bus_name.strip())
    return m.group(1) if m else bus_name.strip()

def _pitch(names, sym_w):
    if not names: return 0
    max_lbl = max(len(n) for n in names) * _CHAR_W
    return max(sym_w + 6, max_lbl + 6)

def _fan_width(names, sym_w):
    if not names: return 0
    return len(names) * _pitch(names, sym_w)

def _bus_half(bus_name, gen_names, load_names):
    name_half = len(bus_name) * _CHAR_W / 2 + _SUB_PAD_H
    gen_half  = _fan_width(gen_names,  2 * _GEN_R)  / 2 + _SUB_PAD_H if gen_names  else 0
    load_half = _fan_width(load_names, 2 * _LOAD_HW) / 2 + _SUB_PAD_H if load_names else 0
    return max(_MIN_HALF, name_half, gen_half, load_half)

def _row_h(gen_names, load_names, sw):
    h = _BUS_LABEL_H + sw + _BUS_KV_H
    if gen_names: h += _GEN_STEM + 2 * _GEN_R + _GEN_LABEL_H
    if load_names: h += _LOAD_STEM + _LOAD_H + _LOAD_LABEL_H
    return h

def _trafo_block_h():
    return _TRAFO_STEM * 2 + 2 * _TRAFO_R * 2 - _TRAFO_GAP + 6

def _sub_height(rows):
    h = _SUB_PAD_TOP + _SUB_PAD_BOT
    for i, row in enumerate(rows):
        h += _BUS_ROW_GAP + _row_h(row["gen_names"], row["load_names"], row["sw"])
        if i < len(rows) - 1:
            h += _trafo_block_h()
    return h

def _sub_width(rows):
    max_half = max((_bus_half(r["name"], r["gen_names"], r["load_names"])
                    for r in rows), default=_MIN_HALF)
    return 2 * max_half + 2 * _SUB_PAD_H

def _order_by_connectivity(sub_keys, sub_to_buses, net):
    if not sub_keys: return []
    bus_to_sub = {}
    for key, bidxs in sub_to_buses.items():
        for b in bidxs:
            bus_to_sub[b] = key
    key_set = set(sub_keys)
    adj = defaultdict(set)
    for _, row in net.line.iterrows():
        ka = bus_to_sub.get(int(row["from_bus"]))
        kb = bus_to_sub.get(int(row["to_bus"]))
        if ka and kb and ka != kb and ka in key_set and kb in key_set:
            adj[ka].add(kb); adj[kb].add(ka)
    for _, row in net.trafo.iterrows():
        ka = bus_to_sub.get(int(row["hv_bus"]))
        kb = bus_to_sub.get(int(row["lv_bus"]))
        if ka and kb and ka != kb and ka in key_set and kb in key_set:
            adj[ka].add(kb); adj[kb].add(ka)
    def dom_vn(key):
        return max((float(net.bus.loc[b, "vn_kv"])
                    for b in sub_to_buses.get(key, [])
                    if b in net.bus.index), default=0)
    remaining = set(sub_keys)
    ordered = []; visited = set(); queue = deque()
    while remaining:
        if not queue:
            start = max(remaining, key=dom_vn)
            queue.append(start); visited.add(start)
        cur = queue.popleft()
        ordered.append(cur); remaining.discard(cur)
        for nb in sorted(adj[cur] - visited, key=dom_vn, reverse=True):
            visited.add(nb); queue.append(nb)
    return ordered

# === SINGLE LINE DIAGRAM (SLD) HTML GENERATION CORE ===
def build_sld_html(net, bus_map, subgrid_nets):
    has_res = not net.res_bus.empty

    eb = set(net.ext_grid["bus"].values)
    gb = set(net.gen["bus"].values)
    def btype(i):
        if i in eb: return "slack"
        if i in gb: return "pv"
        return "pq"

    b2sg = {ci: sg for (sg, _), ci in bus_map.items()}

    bus_gens  = defaultdict(list)
    bus_loads = defaultdict(list)
    for i in net.gen.index:
        bus_gens[int(net.gen.loc[i, "bus"])].append(i)
    for i in net.load.index:
        bus_loads[int(net.load.loc[i, "bus"])].append(i)

    sub_groups = defaultdict(list)
    for idx, brow in net.bus.iterrows():
        sg  = b2sg.get(idx, "LEYTE-SAMAR")
        sub = _substation_name(brow["name"])
        sub_groups[(sg, sub)].append((idx, brow["name"], float(brow["vn_kv"])))

    sub_plan = {}
    for key, buses in sub_groups.items():
        rows = []
        for (bidx, bname, vn) in sorted(buses, key=lambda t: -t[2]):
            sw     = _nearest(vn, _SLD_VN_SW2)
            gidxs  = bus_gens[bidx]
            lidxs  = bus_loads[bidx]
            gnames = [str(net.gen.loc[g,  "name"]) for g in gidxs]
            lnames = [str(net.load.loc[l, "name"]) for l in lidxs]
            rows.append({
                "idx": bidx, "name": bname, "vn": vn, "sw": sw,
                "gidxs": gidxs, "lidxs": lidxs,
                "gen_names": gnames, "load_names": lnames,
            })
        sub_plan[key] = rows

    sub_pos  = {}
    sub_dims = {}
    sg_heights = {}

    for sg, (sg_x0, sg_x1, _, _) in _SLD_COL.items():
        keys = [k for k in sub_plan if k[0] == sg]
        if not keys: continue
        sub_to_buses = {k: [r["idx"] for r in sub_plan[k]] for k in keys}
        ordered  = _order_by_connectivity(keys, sub_to_buses, net)
        inner_w  = sg_x1 - sg_x0 - 2 * _COL_INNER_PAD
        col_w    = (inner_w - _COL_GAP_H) / 2
        col_x    = [sg_x0 + _COL_INNER_PAD,
                    sg_x0 + _COL_INNER_PAD + col_w + _COL_GAP_H]
        col_y    = [56.0, 56.0]
        for key in ordered:
            rows = sub_plan[key]
            w    = min(_sub_width(rows), col_w)
            h    = _sub_height(rows)
            sub_dims[key] = (w, h)
            ci   = 0 if col_y[0] <= col_y[1] else 1
            sub_pos[key]  = (col_x[ci], col_y[ci])
            col_y[ci]    += h + _SUB_GAP_V
        sg_heights[sg] = max(col_y) + 20

    canvas_h = max(sg_heights.values(), default=2000) + 60

    bar_pos = {}
    for (sg, sub), rows in sub_plan.items():
        if (sg, sub) not in sub_pos: continue
        bx, by = sub_pos[(sg, sub)]
        w, _   = sub_dims[(sg, sub)]
        cx_sub = bx + w / 2
        y_cur  = by + _SUB_PAD_TOP + _BUS_ROW_GAP
        for i, row in enumerate(rows):
            bar_y = y_cur + _BUS_LABEL_H
            half  = _bus_half(row["name"], row["gen_names"], row["load_names"])
            bar_pos[row["name"]] = (cx_sub, bar_y, half, row["sw"])
            y_cur = bar_y + row["sw"] + _BUS_KV_H
            if row["gen_names"]:  y_cur += _GEN_STEM + 2 * _GEN_R + _GEN_LABEL_H
            if row["load_names"]: y_cur += _LOAD_STEM + _LOAD_H + _LOAD_LABEL_H
            if i < len(rows) - 1: y_cur += _trafo_block_h()
            y_cur += _BUS_ROW_GAP

    parts = []

    for sg, (sg_x0, sg_x1, sg_bc, sg_tc) in _SLD_COL.items():
        h = sg_heights.get(sg, canvas_h)
        parts.append(
            f'<rect class="sbox" x="{sg_x0}" y="14" '
            f'width="{sg_x1-sg_x0}" height="{h}" stroke="{sg_bc}" rx="6" '
            f'pointer-events="none"/>\n'
            f'<text class="stitle" x="{sg_x0+14}" y="38" fill="{sg_tc}">{sg}</text>\n'
        )

    tie_names = {t[3] for t in TIE_LINES}
    corridor_count = defaultdict(int)
    for _, row in net.line.iterrows():
        if row["name"] in tie_names: continue
        fn = net.bus.loc[row["from_bus"], "name"] if row["from_bus"] in net.bus.index else None
        tn = net.bus.loc[row["to_bus"],   "name"] if row["to_bus"]   in net.bus.index else None
        if fn and tn and fn in bar_pos and tn in bar_pos:
            corridor_count[frozenset([fn, tn])] += 1

    corridor_drawn = defaultdict(int)
    for lidx, row in net.line.iterrows():
        if row["name"] in tie_names: continue
        fn = net.bus.loc[row["from_bus"], "name"] if row["from_bus"] in net.bus.index else None
        tn = net.bus.loc[row["to_bus"],   "name"] if row["to_bus"]   in net.bus.index else None
        if not fn or not tn or fn not in bar_pos or tn not in bar_pos: continue

        ci, yi, hi, swi = bar_pos[fn]
        cj, yj, hj, swj = bar_pos[tn]
        corr  = frozenset([fn, tn])
        n_par = corridor_count[corr]
        drawn = corridor_drawn[corr]
        corridor_drawn[corr] += 1

        vn_ = float(net.bus.loc[row["from_bus"], "vn_kv"])
        sw_ = max(1.2, _nearest(vn_, _SLD_VN_SW2) * 0.30)

        is_outage_ = row["name"] in outage_state or not row.get("in_service", True)
        if is_outage_:
            loading = None
            col_    = "#444444"
            tip     = f"{_e(row['name'])}\n{_e(fn)}  \u2192  {_e(tn)}\n⚡ OUT OF SERVICE"
        elif has_res and lidx in net.res_line.index:
            loading = float(net.res_line.loc[lidx, "loading_percent"])
            col_    = _load_col(loading)
            tip     = (f"{_e(row['name'])}\n{_e(fn)}  \u2192  {_e(tn)}"
                       f"\nLoading: {loading:.1f}%")
        else:
            loading = None
            col_    = _nearest(vn_, _SLD_VN_COLOR2)
            tip     = f"{_e(row['name'])}\n{_e(fn)}  \u2192  {_e(tn)}"
        STEP   = 9
        offset = (drawn - (n_par - 1) / 2.0) * STEP
        same_sub = (_substation_name(fn) == _substation_name(tn))

        if same_sub:
            xi = ci + offset; xj = cj + offset
        else:
            go_right = cj > ci
            xi = (ci + hi + 5 + drawn * 7) if go_right else (ci - hi - 5 - drawn * 7)
            xj = (cj - hj - 5 - drawn * 7) if go_right else (cj + hj + 5 + drawn * 7)

        if abs(yi - yj) < 4:
            hw_y = yi + swi / 2 + 32 + drawn * 14
            d    = (f"M {xi:.1f},{yi} L {xi:.1f},{hw_y:.1f} "
                    f"L {xj:.1f},{hw_y:.1f} L {xj:.1f},{yj}")
            lmx, lmy = (xi + xj) / 2, hw_y - 6
        else:
            my = (yi + yj) / 2
            d  = (f"M {xi:.1f},{yi} L {xi:.1f},{my:.1f} "
                  f"L {xj:.1f},{my:.1f} L {xj:.1f},{yj}")
            lmx, lmy = (xi + xj) / 2, my - 6

        outage_dash    = ' stroke-dasharray="10,6"' if is_outage_ else ""
        outage_opacity = ' opacity="0.5"'            if is_outage_ else ""
        node_opacity   = ' opacity="0.5"'            if is_outage_ else ""

        parts.append(
            f'<path d="{d}" stroke="transparent" stroke-width="14" '
            f'fill="none" style="pointer-events:stroke">'
            f'<title>{tip}</title></path>\n'
        )
        parts.append(
            f'<path d="{d}" stroke="{col_}" stroke-width="{sw_:.1f}" '
            f'fill="none"{outage_dash}{outage_opacity} style="pointer-events:none">'
            f'<title>{tip}</title></path>\n'
        )

        for nx_, ny_ in [(xi, yi), (xj, yj)]:
            parts.append(
                f'<circle cx="{nx_:.1f}" cy="{ny_:.1f}" r="9" '
                f'fill="transparent" stroke="none" '
                f'style="pointer-events:all">'
                f'<title>{tip}</title></circle>\n'
                f'<circle cx="{nx_:.1f}" cy="{ny_:.1f}" r="2.8" '
                f'fill="{col_}" stroke="#0b0f14" stroke-width="1.0" '
                f'style="pointer-events:none"'
                f'{node_opacity}/>\n'
            )

        if loading is not None and loading > 70:
            parts.append(
                f'<text style="font-size:6px;fill:{col_};text-anchor:middle;'
                f'font-family:\'Courier New\',monospace" '
                f'x="{lmx:.1f}" y="{lmy:.1f}">{loading:.0f}%</text>\n'
            )

    for (sg, sub), rows in sub_plan.items():
        if (sg, sub) not in sub_pos: continue
        bx, by = sub_pos[(sg, sub)]
        w, h   = sub_dims[(sg, sub)]
        sg_bc  = _SLD_COL[sg][2]
        cx_sub = bx + w / 2
        y_cur  = by + _SUB_PAD_TOP + _BUS_ROW_GAP

        parts.append(
            f'<rect x="{bx:.1f}" y="{by:.1f}" width="{w:.1f}" height="{h:.1f}" '
            f'rx="4" fill="rgba(255,255,255,0.025)" stroke="{sg_bc}" '
            f'stroke-width="0.7" stroke-dasharray="5,3" pointer-events="none"/>\n'
            f'<text style="font-size:7.5px;fill:{sg_bc};font-weight:700;'
            f'font-family:\'Courier New\',monospace" '
            f'x="{bx+6:.1f}" y="{by+16:.1f}">{_e(sub)}</text>\n'
        )

        for i, row in enumerate(rows):
            bar_y  = y_cur + _BUS_LABEL_H
            sw_    = row["sw"]
            col_   = _nearest(row["vn"], _SLD_VN_COLOR2)
            half   = _bus_half(row["name"], row["gen_names"], row["load_names"])
            x0b    = cx_sub - half
            x1b    = cx_sub + half
            gidxs  = row["gidxs"]
            lidxs  = row["lidxs"]
            gnames = row["gen_names"]
            lnames = row["load_names"]
            bidx   = row["idx"]

            if has_res and bidx in net.res_bus.index:
                res_  = net.res_bus.loc[bidx]
                vm_   = float(res_["vm_pu"])
                warn  = vm_ < 0.95 or vm_ > 1.05
                bcol  = "#fbbf24" if warn else col_
                lmp_s = (f"\nLMP: \u20b1{float(res_['lam_p']):.2f}/MWh"
                         if "lam_p" in res_.index else "")
                tip_b = (f"{_e(row['name'])}\n{int(row['vn'])} kV"
                         f"\nVm={vm_:.4f}  Va={float(res_['va_degree']):.2f}\u00b0{lmp_s}")
            else:
                bcol  = col_
                tip_b = f"{_e(row['name'])}\n{int(row['vn'])} kV"

            bt  = btype(bidx)
            dot = {"slack":"#ef4444","pv":"#22c55e","pq":"#3b82f6"}[bt]

            parts.append(
                f'<text class="blbl" x="{cx_sub:.1f}" y="{bar_y - sw_ - 4:.1f}" '
                f'text-anchor="middle">{_e(row["name"])}</text>\n'
            )
            parts.append(
                f'<line x1="{x0b:.1f}" y1="{bar_y}" x2="{x1b:.1f}" y2="{bar_y}" '
                f'stroke="transparent" stroke-width="16" stroke-linecap="round" '
                f'style="pointer-events:stroke">'
                f'<title>{tip_b}</title></line>\n'
            )
            parts.append(
                f'<line x1="{x0b:.1f}" y1="{bar_y}" x2="{x1b:.1f}" y2="{bar_y}" '
                f'stroke="{bcol}" stroke-width="{sw_}" stroke-linecap="round" '
                f'style="pointer-events:none">'
                f'<title>{tip_b}</title></line>\n'
            )
            parts.append(
                f'<text class="bkv" x="{cx_sub:.1f}" y="{bar_y + sw_ + 9:.1f}" '
                f'text-anchor="middle">{int(row["vn"])} kV</text>\n'
            )

            if bidx in eb:
                Y0  = bar_y - sw_ - 10
                egn = ""
                for _, eg in net.ext_grid.iterrows():
                    if eg["bus"] == bidx:
                        egn = eg["name"]; break
                c = "#f87171"
                parts.append(
                    f'<line x1="{cx_sub:.1f}" y1="{Y0}" x2="{cx_sub:.1f}" y2="{Y0-10}" stroke="{c}" stroke-width="1.6"/>\n'
                    f'<line x1="{cx_sub-14:.1f}" y1="{Y0-10}" x2="{cx_sub+14:.1f}" y2="{Y0-10}" stroke="{c}" stroke-width="2.2"/>\n'
                    f'<line x1="{cx_sub-11:.1f}" y1="{Y0-15}" x2="{cx_sub+11:.1f}" y2="{Y0-15}" stroke="{c}" stroke-width="1.8"/>\n'
                    f'<line x1="{cx_sub-7:.1f}"  y1="{Y0-20}" x2="{cx_sub+7:.1f}"  y2="{Y0-20}" stroke="{c}" stroke-width="1.4"/>\n'
                    + (f'<text style="font-size:6px;fill:{c};text-anchor:middle;font-family:\'Courier New\',monospace" '
                       f'x="{cx_sub:.1f}" y="{Y0-26}">{_e(egn)}</text>\n' if egn else "")
                )

            y_after_bar = bar_y + sw_ + _BUS_KV_H

            if gnames:
                gp      = _pitch(gnames, 2 * _GEN_R)
                n_g     = len(gidxs)
                rail_y  = y_after_bar + _GEN_STEM
                gen_xs  = [cx_sub + (k - (n_g - 1) / 2.0) * gp for k in range(n_g)]
                parts.append(
                    f'<line x1="{cx_sub:.1f}" y1="{y_after_bar:.1f}" '
                    f'x2="{cx_sub:.1f}" y2="{rail_y:.1f}" '
                    f'stroke="#3d5470" stroke-width="1.1"/>\n'
                )
                if n_g > 1:
                    parts.append(
                        f'<line x1="{gen_xs[0]:.1f}" y1="{rail_y:.1f}" '
                        f'x2="{gen_xs[-1]:.1f}" y2="{rail_y:.1f}" '
                        f'stroke="#3d5470" stroke-width="1.1"/>\n'
                    )
                for k, gi in enumerate(gidxs):
                    gx    = gen_xs[k]
                    cy_c  = rail_y + _GEN_R
                    gname = gnames[k]
                    tip_g = gname
                    p_lbl = ""
                    if has_res and gi in net.res_gen.index:
                        p_mw  = float(net.res_gen.loc[gi, "p_mw"])
                        q_mv  = float(net.res_gen.loc[gi, "q_mvar"])
                        tip_g = f"{gname}\n{p_mw:.1f} MW  {q_mv:.1f} Mvar"
                        p_lbl = (
                            f'<text style="font-size:5.5px;fill:#64748b;text-anchor:middle;'
                            f'font-family:\'Courier New\',monospace" '
                            f'x="{gx:.1f}" y="{cy_c + _GEN_R + 9:.1f}">'
                            f'{p_mw:.1f} MW</text>\n'
                        )
                    parts.append(
                        f'<line x1="{gx:.1f}" y1="{rail_y:.1f}" '
                        f'x2="{gx:.1f}" y2="{cy_c - _GEN_R:.1f}" '
                        f'stroke="#3d5470" stroke-width="1.1"/>\n'
                        f'<circle cx="{gx:.1f}" cy="{cy_c:.1f}" r="{_GEN_R + 8}" '
                        f'fill="transparent" stroke="none" '
                        f'style="pointer-events:all">'
                        f'<title>{_e(tip_g)}</title></circle>\n'
                        f'<circle cx="{gx:.1f}" cy="{cy_c:.1f}" r="{_GEN_R}" '
                        f'class="gen" style="pointer-events:none">'
                        f'<title>{_e(tip_g)}</title></circle>\n'
                        f'<text class="genlbl" text-anchor="middle" '
                        f'x="{gx:.1f}" y="{cy_c + 2.5:.1f}">G</text>\n'
                        f'<text style="font-size:5.5px;fill:#38bdf8;text-anchor:middle;'
                        f'font-family:\'Courier New\',monospace" '
                        f'x="{gx:.1f}" y="{cy_c + _GEN_R + 9:.1f}">'
                        f'{_e(gname)}</text>\n'
                        + p_lbl
                    )
                y_after_bar = rail_y + _GEN_R * 2 + _GEN_LABEL_H

            if lnames:
                lp      = _pitch(lnames, 2 * _LOAD_HW)
                n_l     = len(lidxs)
                rail_y  = y_after_bar + _LOAD_STEM
                load_xs = [cx_sub + (k - (n_l - 1) / 2.0) * lp for k in range(n_l)]
                parts.append(
                    f'<line x1="{cx_sub:.1f}" y1="{y_after_bar:.1f}" '
                    f'x2="{cx_sub:.1f}" y2="{rail_y:.1f}" '
                    f'stroke="#3d5470" stroke-width="1.1"/>\n'
                )
                if n_l > 1:
                    parts.append(
                        f'<line x1="{load_xs[0]:.1f}" y1="{rail_y:.1f}" '
                        f'x2="{load_xs[-1]:.1f}" y2="{rail_y:.1f}" '
                        f'stroke="#3d5470" stroke-width="1.1"/>\n'
                    )
                for k, li in enumerate(lidxs):
                    lx_    = load_xs[k]
                    ty_top = rail_y
                    ty_bot = ty_top + _LOAD_H
                    lname  = lnames[k]
                    p_mw   = None
                    if has_res and li in net.res_load.index:
                        p_mw = float(net.res_load.loc[li, "p_mw"])
                    active = p_mw is not None and abs(p_mw) > 0.01
                    fill_  = "#fb923c" if active else "none"
                    stk_   = "#d97706" if active else "#374151"
                    tip_l  = lname + (f"\n{p_mw:.1f} MW" if p_mw is not None else "")
                    parts.append(
                        f'<line x1="{lx_:.1f}" y1="{rail_y:.1f}" '
                        f'x2="{lx_:.1f}" y2="{ty_top:.1f}" '
                        f'stroke="#3d5470" stroke-width="1.1"/>\n'
                        f'<polygon points="{lx_:.1f},{ty_bot:.1f} '
                        f'{lx_-_LOAD_HW:.1f},{ty_top:.1f} '
                        f'{lx_+_LOAD_HW:.1f},{ty_top:.1f}" '
                        f'fill="{fill_}" stroke="{stk_}" stroke-width="1">'
                        f'<title>{_e(tip_l)}</title></polygon>\n'
                        f'<text style="font-size:5.5px;fill:#94a3b8;text-anchor:middle;'
                        f'font-family:\'Courier New\',monospace" '
                        f'x="{lx_:.1f}" y="{ty_bot + 9:.1f}">{_e(lname)}</text>\n'
                    )

            if i < len(rows) - 1:
                next_row = rows[i + 1]
                t_col    = "#fb923c"
                tip_t    = f"{_e(row['name'])} \u2192 {_e(next_row['name'])}"
                for tidx, trow in net.trafo.iterrows():
                    hb = net.bus.loc[trow["hv_bus"], "name"] if trow["hv_bus"] in net.bus.index else None
                    lb = net.bus.loc[trow["lv_bus"], "name"] if trow["lv_bus"] in net.bus.index else None
                    if {hb, lb} == {row["name"], next_row["name"]}:
                        if has_res and tidx in net.res_trafo.index:
                            lt    = float(net.res_trafo.loc[tidx, "loading_percent"])
                            t_col = "#ef4444" if lt > 90 else "#22c55e"
                            tip_t = f"{_e(trow['name'])}\nLoading: {lt:.1f}%"
                        else:
                            t_col = _nearest(float(net.bus.loc[trow["hv_bus"], "vn_kv"]),
                                             _SLD_VN_COLOR2)
                            tip_t = _e(trow["name"])
                        break

                y_bottom = bar_y + sw_ + _BUS_KV_H
                if gnames: y_bottom += _GEN_STEM + 2 * _GEN_R + _GEN_LABEL_H
                if lnames: y_bottom += _LOAD_STEM + _LOAD_H + _LOAD_LABEL_H

                TR  = _TRAFO_R
                tc1 = y_bottom + _TRAFO_STEM
                tc2 = tc1 + 2 * TR - _TRAFO_GAP
                parts.append(
                    f'<line x1="{cx_sub:.1f}" y1="{y_bottom:.1f}" '
                    f'x2="{cx_sub:.1f}" y2="{tc1 - TR:.1f}" '
                    f'stroke="#3d5470" stroke-width="1.2"/>\n'
                    f'<circle cx="{cx_sub:.1f}" cy="{tc1:.1f}" r="{TR + 8}" '
                    f'fill="transparent" stroke="none" '
                    f'style="pointer-events:all">'
                    f'<title>{tip_t}</title></circle>\n'
                    f'<circle cx="{cx_sub:.1f}" cy="{tc1:.1f}" r="{TR}" '
                    f'fill="none" stroke="{t_col}" stroke-width="1.5" '
                    f'style="pointer-events:none">'
                    f'<title>{tip_t}</title></circle>\n'
                    f'<circle cx="{cx_sub:.1f}" cy="{tc2:.1f}" r="{TR}" '
                    f'fill="none" stroke="{t_col}" stroke-width="1.5"/>\n'
                    f'<line x1="{cx_sub:.1f}" y1="{tc2 + TR:.1f}" '
                    f'x2="{cx_sub:.1f}" y2="{tc2 + TR + _TRAFO_STEM:.1f}" '
                    f'stroke="#3d5470" stroke-width="1.2"/>\n'
                )

            y_cur = bar_y + sw_ + _BUS_KV_H
            if gnames: y_cur += _GEN_STEM + 2 * _GEN_R + _GEN_LABEL_H
            if lnames: y_cur += _LOAD_STEM + _LOAD_H + _LOAD_LABEL_H
            if i < len(rows) - 1: y_cur += _trafo_block_h()
            y_cur += _BUS_ROW_GAP

    drawn_tie = set()
    for (fb_n, tb_n, length_km, tie_name, _) in TIE_LINES:
        corr = frozenset([fb_n, tb_n])
        if corr in drawn_tie: continue
        drawn_tie.add(corr)
        if fb_n not in bar_pos or tb_n not in bar_pos: continue
        ci, yi, hi, _ = bar_pos[fb_n]
        cj, yj, hj, _ = bar_pos[tb_n]
        hw_y = min(yi, yj) - 35
        lbl  = f"{_e(fb_n)} \u2194 {_e(tb_n)}  {length_km:.0f} km"
        mx_  = (ci + cj) / 2
        d    = (f"M {ci:.1f},{yi} L {ci:.1f},{hw_y:.1f} "
                f"L {cj:.1f},{hw_y:.1f} L {cj:.1f},{yj}")
        parts.append(
            f'<path d="{d}" class="ltie" '
            f'marker-end="url(#arr)"><title>{lbl}</title></path>\n'
            f'<text style="font-size:7px;fill:#a78bfa;text-anchor:middle;'
            f'font-family:\'Courier New\',monospace" '
            f'x="{mx_:.1f}" y="{hw_y - 4:.1f}">{lbl}</text>\n'
        )

    lx_, ly_ = _SLD_LEGEND_X + 10, 60
    leg_rows = [
        ("#f87171", 5,   "230 kV Bus",          ""),
        ("#fb923c", 4,   "138 / 115 kV Bus",     ""),
        ("#38bdf8", 3,   "69 kV Bus",            ""),
        ("#a78bfa", 2.5, "34.5 / 33 kV Bus",     ""),
        ("#4ade80", 2.5, "18 kV Bus",            ""),
        ("#e879f9", 2.5, "13.8 kV Bus",          ""),
        ("#fbbf24", 2,   "6.9 kV Bus",           ""),
        ("#22d3ee", 1.6, "Line loading < 70 %",  ""),
        ("#f97316", 1.6, "Line loading 70-90 %", ""),
        ("#ef4444", 1.6, "Line loading > 90 %",  ""),
        ("#a78bfa", 2,   "Interisland tie-line",  'stroke-dasharray="7,3"'),
        ("#ef4444", 2,   "Slack / Ext Grid",      ""),
        ("#22c55e", 2,   "Generator (PV) bus",    ""),
        ("#3b82f6", 2,   "Load-only (PQ) bus",    ""),
        ("#fbbf24", 2,   "Voltage violation",     ""),
    ]
    ROW    = 20
    BOX_H2 = len(leg_rows) * ROW + 170
    parts.append(
        f'<rect x="{lx_-10}" y="{ly_-18}" width="290" height="{BOX_H2}" '
        f'rx="5" fill="#080c11" stroke="#1e3a5f" stroke-width="1" '
        f'pointer-events="none"/>\n'
        f'<text style="font-size:10px;fill:#38bdf8;font-weight:700;'
        f'font-family:\'Courier New\',monospace" x="{lx_}" y="{ly_}">LEGEND</text>\n'
    )
    for i, (lc, lw, ll, ld) in enumerate(leg_rows):
        iy = ly_ + 18 + i * ROW
        parts.append(
            f'<line x1="{lx_}" y1="{iy}" x2="{lx_+50}" y2="{iy}" '
            f'stroke="{lc}" stroke-width="{lw}" {ld}/>\n'
            f'<text style="font-size:7px;fill:#cbd5e1;'
            f'font-family:\'Courier New\',monospace" x="{lx_+56}" y="{iy+3}">'
            f'{ll}</text>\n'
        )
    sy = ly_ + 18 + len(leg_rows) * ROW + 10
    parts.append(
        f'<circle cx="{lx_+18}" cy="{sy+10}" r="{_GEN_R-1}" class="gen"/>'
        f'<text class="genlbl" text-anchor="middle" x="{lx_+18}" y="{sy+13}">G</text>'
        f'<text style="font-size:7px;fill:#cbd5e1;font-family:\'Courier New\',monospace" '
        f'x="{lx_+34}" y="{sy+14}">Generator</text>\n'
        f'<polygon points="{lx_+18},{sy+44} {lx_+7},{sy+30} {lx_+29},{sy+30}" '
        f'fill="#fb923c" stroke="#d97706" stroke-width="1"/>'
        f'<text style="font-size:7px;fill:#cbd5e1;font-family:\'Courier New\',monospace" '
        f'x="{lx_+34}" y="{sy+42}">Active Load</text>\n'
    )

    if has_res:
        tot_gen  = float(net.res_gen["p_mw"].sum())  if not net.res_gen.empty  else 0
        tot_load = float(net.res_load["p_mw"].sum()) if not net.res_load.empty else 0
        sumtxt   = (f"Buses: {len(net.bus)}  |  Lines: {len(net.line)}  "
                    f"|  Gen: {tot_gen:.1f} MW  |  Load: {tot_load:.1f} MW")
    else:
        sumtxt = "Simulation results not yet available"
    parts.append(
        f'<text style="font-size:9px;fill:#1e3a5f;'
        f'font-family:\'Courier New\',monospace" x="30" y="{canvas_h - 20}">'
        f'{_e(sumtxt)}</text>\n'
    )

    svg_body = "".join(parts)
    CANVAS_W = _SLD_LEGEND_X + 310

    html_out = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>Visayas Grid \u2014 Single Line Diagram</title>
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{background:#0b0f14;font-family:'Courier New',monospace;color:#e2e8f0;
      overflow:hidden;user-select:none}}
#hdr{{position:fixed;top:0;left:0;right:0;height:42px;background:#0f1923;
      border-bottom:1px solid #1e3a5f;display:flex;align-items:center;
      gap:10px;padding:0 14px;z-index:100}}
#hdr h1{{font-size:12px;font-weight:700;letter-spacing:3px;color:#38bdf8;
         text-transform:uppercase}}
.sep{{width:1px;height:18px;background:#1e3a5f}}
.leg{{display:flex;align-items:center;gap:4px;font-size:10px;color:#94a3b8}}
.ldot{{width:12px;height:4px;border-radius:2px}}
#zi{{margin-left:auto;font-size:10px;color:#475569}}
#rb{{padding:3px 8px;background:#0f1923;border:1px solid #1e3a5f;
     color:#94a3b8;font-size:10px;cursor:pointer;border-radius:3px}}
#rb:hover{{background:#1e3a5f;color:#e2e8f0}}
#cvs{{position:fixed;top:42px;left:0;right:0;bottom:0}}
svg{{width:100%;height:100%;cursor:grab}}
svg.dragging{{cursor:grabbing}}
.gen{{fill:#0c2340;stroke:#38bdf8;stroke-width:1.4}}
.genlbl{{font-size:6.5px;fill:#38bdf8;font-family:'Courier New',monospace}}
.blbl{{font-size:7.5px;fill:#e2e8f0;font-family:'Courier New',monospace;font-weight:700}}
.bkv{{font-size:6px;fill:#64748b;font-family:'Courier New',monospace}}
.sbox{{fill:none;stroke-width:0.8;stroke-dasharray:5,3;pointer-events:none}}
.stitle{{font-size:12px;font-weight:700;font-family:'Courier New',monospace;
          letter-spacing:2px}}
.ltie{{stroke:#a78bfa;stroke-width:2;fill:none;stroke-dasharray:9,4}}
.tip{{position:fixed;background:#0f1923;border:1px solid #1e3a5f;padding:6px 10px;
      font-size:10px;pointer-events:none;border-radius:4px;z-index:200;
      display:none;max-width:300px;line-height:1.7;white-space:pre-wrap;color:#e2e8f0}}
</style>
</head>
<body>
<div id="hdr">
  <h1>Visayas Grid \u2014 SLD</h1>
  <div class="sep"></div>
  <div class="leg"><div class="ldot" style="background:#f87171"></div>230 kV</div>
  <div class="leg"><div class="ldot" style="background:#fb923c"></div>138 kV</div>
  <div class="leg"><div class="ldot" style="background:#38bdf8"></div>69 kV</div>
  <div class="leg"><div class="ldot" style="background:#e879f9"></div>13.8 kV</div>
  <div class="leg"><div class="ldot" style="background:#4ade80"></div>18 / 34.5 kV</div>
  <div class="leg"><div class="ldot" style="background:#a78bfa"></div>Tie-line</div>
  <div class="sep"></div>
  <div class="leg">&#9711;&nbsp;Gen&nbsp; &#9661;&nbsp;Load&nbsp; &#8284;&nbsp;Slack</div>
  <div id="zi">Scroll&nbsp;=&nbsp;zoom &nbsp;\u00b7&nbsp; Drag&nbsp;=&nbsp;pan</div>
  <button id="rb" onclick="resetView()">Reset</button>
</div>
<div id="cvs">
<svg id="sld" xmlns="http://www.w3.org/2000/svg">
<defs>
  <marker id="arr" markerWidth="6" markerHeight="6" refX="5" refY="3" orient="auto">
    <path d="M0,0 L6,3 L0,6Z" fill="#a78bfa"/>
  </marker>
</defs>
<g id="root">
{svg_body}
</g>
</svg>
</div>
<div class="tip" id="tip"></div>
<script>
const svg = document.getElementById('sld');
const zi  = document.getElementById('zi');
const tip = document.getElementById('tip');
const VW = {CANVAS_W}, VH = {canvas_h};
let vx=0, vy=0, vw=VW, vh=VH, dragging=false, lx=0, ly=0;
function applyView() {{
  svg.setAttribute('viewBox', `${{vx.toFixed(0)}} ${{vy.toFixed(0)}} ${{vw.toFixed(0)}} ${{vh.toFixed(0)}}`);
  zi.textContent = `Zoom ${{(VW / vw * 100).toFixed(0)}}%`;
}}
applyView();
svg.addEventListener('wheel', e => {{
  e.preventDefault();
  const r  = svg.getBoundingClientRect();
  const mx = (e.clientX - r.left) / r.width  * vw + vx;
  const my = (e.clientY - r.top)  / r.height * vh + vy;
  const f  = e.deltaY > 0 ? 1.10 : 0.91;
  vw *= f; vh *= f;
  vx = mx - (e.clientX - r.left) / r.width  * vw;
  vy = my - (e.clientY - r.top)  / r.height * vh;
  applyView();
}}, {{ passive: false }});
svg.addEventListener('mousedown', e => {{ dragging=true; lx=e.clientX; ly=e.clientY; svg.classList.add('dragging'); }});
window.addEventListener('mousemove', e => {{
  if (!dragging) return;
  vx -= (e.clientX - lx) / svg.clientWidth  * vw;
  vy -= (e.clientY - ly) / svg.clientHeight * vh;
  lx = e.clientX; ly = e.clientY; applyView();
}});
window.addEventListener('mouseup', () => {{ dragging=false; svg.classList.remove('dragging'); }});
function resetView() {{ vx=0; vy=0; vw=VW; vh=VH; applyView(); }}
document.querySelectorAll('line,circle,path,polygon,rect').forEach(el => {{
  const t = el.querySelector('title');
  if (!t) return;
  el.addEventListener('mouseenter', () => {{ tip.style.display='block'; tip.textContent=t.textContent; }});
  el.addEventListener('mousemove',  e => {{ tip.style.left=(e.clientX+14)+'px'; tip.style.top=(e.clientY-10)+'px'; }});
  el.addEventListener('mouseleave', () => {{ tip.style.display='none'; }});
}});
</script>
</body>
</html>
"""
    out_dir   = state["base_dir"]

    auto_save = (not dpg.does_item_exist("auto_generate_sld")
                 or dpg.get_value("auto_generate_sld"))
    if auto_save:
        out_path     = os.path.join(out_dir, SLD_FILE)
        out_path_tmp = out_path + ".tmp"
        try:
            with open(out_path_tmp, "w", encoding="utf-8") as fh:
                fh.write(html_out)
            if os.path.exists(out_path):
                os.replace(out_path_tmp, out_path)
            else:
                os.rename(out_path_tmp, out_path)
        except Exception:
            pass

    state["sld_html"] = html_out

# === SIMULATION EXECUTION ===
def run_simulation():
    try:
        os.chdir(state["base_dir"])

        state["status"]="Loading subgrids..."
        subgrid_nets={}
        for folder in SUBGRIDS:
            state["status"]=f"Loading {folder}..."
            subgrid_nets[folder]=load_subgrid(folder)

        state["status"]="Merging networks..."
        net,bus_map=merge_networks(subgrid_nets)

        state["status"]="Adding HVDC connections..."
        global HVDC_CONNECTIONS
        HVDC_CONNECTIONS = load_hvdc_connections()
        hvdc_src = "hvdc.csv" if os.path.exists(
            os.path.join(period_base_dir(),"hvdc.csv")) else "defaults"
        state["status"] = f"Adding HVDC ({hvdc_src})..."
        for hvdc in HVDC_CONNECTIONS:
            cb=bus_map[(hvdc["subgrid"],hvdc["bus"])]
            if hvdc["is_slack"]:
                pp.create_ext_grid(net,bus=cb,vm_pu=hvdc["vm_pu"],name=hvdc["name"])
            else:
                pp.create_gen(net,bus=cb,p_mw=0.0,vm_pu=hvdc["vm_pu"],
                              name=hvdc["name"],controllable=True)
                net.gen.loc[net.gen["name"]==hvdc["name"],"min_p_mw"]=hvdc["min_p_mw"]
                net.gen.loc[net.gen["name"]==hvdc["name"],"max_p_mw"]=hvdc["max_p_mw"]
        assert len(net.ext_grid)==1, f"Expected 1 ext_grid, got {len(net.ext_grid)}"

        state["status"]="Adding tie-lines..."
        for (fb_n,tb_n,length_km,tie_name,i_ka) in TIE_LINES:
            fb=bus_map[(TIE_BUS_SUBGRID[fb_n],fb_n)]
            tb=bus_map[(TIE_BUS_SUBGRID[tb_n],tb_n)]
            pp.create_line_from_parameters(net, from_bus=fb, to_bus=tb,
                                            length_km=length_km,
                                            r_ohm_per_km=get_constraint("tie_r_ohm_per_km"),
                                            x_ohm_per_km=get_constraint("tie_x_ohm_per_km"),
                                            c_nf_per_km=get_constraint("tie_c_nf_per_km"),
                                            max_i_ka=i_ka, name=tie_name)
        net.line["max_loading_percent"]=100

        state["status"]="Applying line outages..."
        if outage_state:
            for lidx, lrow in net.line.iterrows():
                if lrow["name"] in outage_state:
                    net.line.loc[lidx, "in_service"] = False
                    print(f"[OUTAGE] Tripped line: {lrow['name']}")
            print(f"[OUTAGE] {len(outage_state)} line(s) out of service: {outage_state}")

        state["status"]="Running Newton-Raphson..."
        pp.runpp(net, algorithm="nr", init="auto",
                calculate_voltage_angles=True,
                tolerance_mva=1e-6,
                max_iteration=int(get_constraint("nr_max_iteration")),
                numba=True, enforce_q_lims=True)

        state["status"]="Setting up OPF costs..."
        gen_folder={}
        for folder,(sg,_) in subgrid_nets.items():
            for _,row in sg.gen.iterrows(): gen_folder[row["name"]]=folder

        hvdc_gen_names = [h["name"] for h in HVDC_CONNECTIONS if not h["is_slack"]]

        for i in net.gen.index:
            gn=net.gen.loc[i,"name"]
            if gn in hvdc_gen_names: continue
            folder=gen_folder[gn]
            _,gc=subgrid_nets[folder]; cr=gc[gc["name"]==gn].iloc[0].to_dict()
            pmax=cr["dep_capacity"] if USE_DEP_CAPACITY else cr["max_p_mw"]
            net.gen.loc[i,"min_p_mw"]=cr["min_p_mw"]
            net.gen.loc[i,"max_p_mw"]=pmax
            net.gen.loc[i,"min_q_mvar"]=cr["min_q_mvar"]
            net.gen.loc[i,"max_q_mvar"]=cr["max_q_mvar"]
            net.gen.loc[i,"controllable"]= pmax > 0
            if pmax==0:
                net.gen.loc[i,"min_q_mvar"]=0.0
                net.gen.loc[i,"max_q_mvar"]=0.0
            segs=[]
            for k in range(1,4):
                if f"p{2*k-1}" in cr and pd.notna(cr[f"p{2*k-1}"]):
                    segs.append([float(cr[f"p{2*k-1}"]),float(cr[f"p{2*k}"]),float(cr[f"c{k}"])])
            if segs: pp.create_pwl_cost(net,i,"gen",segs)

        slack_idx=net.ext_grid.index[0]
        h0 = HVDC_CONNECTIONS[0]
        net.ext_grid.loc[slack_idx,"min_p_mw"]=h0["min_p_mw"]
        net.ext_grid.loc[slack_idx,"max_p_mw"]=h0["max_p_mw"]
        net.ext_grid.loc[slack_idx,"controllable"]=True
        pp.create_pwl_cost(net,slack_idx,"ext_grid",
            [[h0["min_p_mw"],h0["max_p_mw"],h0["cost"]]])

        h1 = HVDC_CONNECTIONS[1]
        vmhvdc_idx=net.gen[net.gen["name"]==h1["name"]].index[0]
        pp.create_pwl_cost(net,vmhvdc_idx,"gen",
            [[h1["min_p_mw"],h1["max_p_mw"],h1["cost"]]])

        state["status"]="Running OPF..."
        pp.runopp(net, verbose=False, delta=1e-10, init="pf",
          max_iteration=int(get_constraint("opf_max_iteration")))
        state["net"]=net; state["bus_map"]=bus_map; state["subgrid_nets"]=subgrid_nets

        current_label = period_state["period"]
        override_tag  = (f" + {len(mem_overrides)} override(s)" if mem_overrides else "")
        outage_tag    = (f" + {len(outage_state)} outage(s)" if outage_state else "")
        run_label     = f"{current_label}{override_tag}{outage_tag}"
        state["last_run_label"] = run_label   

        if not comparison["has_baseline"]:
            comparison["baseline_net"]   = net
            comparison["has_baseline"]   = True
            comparison["baseline_label"] = run_label
            comparison["has_modified"]   = False
            comparison["modified_net"]   = None
            comparison["modified_label"] = ""
            print(f"[CMP] Auto-pinned baseline: {run_label}")
        else:
            comparison["modified_net"]   = net
            comparison["has_modified"]   = True
            comparison["modified_label"] = run_label
            print(f"[CMP] Stored as modified: {run_label}")

        state["status"]="Generating map..."
        build_map(net,bus_map,subgrid_nets)

        state["status"]="Generating SLD..."
        build_sld_html(net, bus_map, subgrid_nets)

        p = period_state["period"]
        outage_str = f" | Outages: {len(outage_state)}" if outage_state else ""
        state["status"] = f"Done | [{p}] | Buses: {len(net.bus)} | Lines: {len(net.line)}{outage_str}"
        state["done"]=True; state["running"]=False
    except Exception as _exc:
        tb = traceback.format_exc()
        hint = ""
        if "LoadflowNotConverged" in tb:
            hint = (
                "\n\n─── HINT ─────────────────────────────────────────────\n"
                "Power flow did not converge. The network is unrealisable\n"
                "with the current modifications. Common causes:\n"
                "  * Generator dispatch too large or too small\n"
                "  * Q limits conflict with dispatch (min/max_q_mvar)\n"
                "  * A bus is islanded from the slack\n"
                "\nClick Reset to Base Case to restore the default network."
            )
        elif "OPFNotConverged" in tb or "runopp" in tb.lower():
            hint = (
                "\n\n─── HINT ─────────────────────────────────────────────\n"
                "OPF did not converge. Common causes:\n"
                "  * Cost curve segments are infeasible\n"
                "  * p_mw outside [min_p_mw, max_p_mw]\n"
                "  * Total generation cannot meet total load\n"
                "\nClick Reset to Base Case to restore the default network."
            )
        state["error"] = tb + hint
        state["status"] = "Error -- see error window"
        state["running"] = False; state["done"] = False


# === MAP GENERATION HELPERS ===
def offset_polyline(coords, offset_m):
    import math
    if len(coords) < 2 or abs(offset_m) < 0.1:
        return coords
    result = []
    n = len(coords)
    for i in range(n):
        p1, p2 = (coords[i], coords[i+1]) if i < n-1 else (coords[i-1], coords[i])
        lat1, lng1 = p1[0], p1[1]
        lat2, lng2 = p2[0], p2[1]
        dlat = lat2 - lat1
        dlng = lng2 - lng1
        mlat = 111320.0
        mlng = 111320.0 * math.cos(math.radians((lat1+lat2)/2))
        dy = dlat * mlat
        dx = dlng * mlng
        length = math.sqrt(dx*dx + dy*dy)
        if length < 1e-9:
            result.append(coords[i])
            continue
        px = -dy / length
        py =  dx / length
        result.append([coords[i][0] + offset_m*px/mlat,
                       coords[i][1] + offset_m*py/mlng])
    return result


# === INTERACTIVE MAP GENERATION ===
def build_map(net,bus_map,subgrid_nets):
    global OSM_LINES, OSM_NAME_INDEX, OSM_RTREE
    OSM_LINES, OSM_NAME_INDEX, OSM_RTREE = _load_osm_vertices()

    geo_map={}
    for folder in SUBGRIDS:
        sg=subgrid_nets[folder][0]
        geo_csv=os.path.join(period_base_dir(),folder,"geo_coords.csv")
        if os.path.exists(geo_csv):
            gdf=pd.read_csv(geo_csv); gdf.columns=gdf.columns.str.strip()
            gdf = gdf.loc[:, ~gdf.columns.str.startswith("Unnamed")]
            bc=next((c for c in gdf.columns if c.lower() in ("bus","name")),None)
            if bc:
                gdf[bc]=gdf[bc].str.strip()
                for _,row in gdf.iterrows(): geo_map[row[bc]]=(float(row["lat"]),float(row["lng"]))
        else:
            b=SUBGRID_BOUNDS[folder]; np.random.seed(abs(hash(folder))%(2**31))
            ns=list(sg.bus["name"])
            lats=np.random.uniform(*b["lat"],len(ns)); lngs=np.random.uniform(*b["lng"],len(ns))
            for i,n in enumerate(ns): geo_map[n]=(lats[i],lngs[i])
    all_lats=[v[0] for v in geo_map.values()]; all_lngs=[v[1] for v in geo_map.values()]
    if not all_lats: return
    center=(np.mean(all_lats),np.mean(all_lngs))

    b2s={ci:f for (f,_),ci in bus_map.items()}
    eb=set(net.ext_grid["bus"].values); gb=set(net.gen["bus"].values)
    def btype(i): return "slack" if i in eb else ("pv" if i in gb else "pq")

    def vn_color(vn):
        if   vn >= 200: return "#ff0000"
        elif vn >= 120: return "#f1a542"
        elif vn >= 60:  return "#08f11c"
        else:           return "#aaaaaa"

    def loading_color(loading):
        if loading > 90:  return "#e74c3c"
        elif loading > 70: return "#e67e22"
        else:              return "#1D9E75"

    m=folium.Map(location=center,zoom_start=8,tiles="OpenStreetMap")
    fg={f:folium.FeatureGroup(name=f,show=True) for f in SUBGRIDS}
    fg_tie=folium.FeatureGroup(name="Tie Lines",show=True)
    fg_outage=folium.FeatureGroup(name="Outaged Lines",show=True)
    fg_gen_dispatch=folium.FeatureGroup(name="Generator Dispatch",show=False)
    fg_load_demand=folium.FeatureGroup(name="Load Demand",show=False)    
    tie_set={t[3] for t in TIE_LINES}

    from collections import defaultdict
    corridor_lines = defaultdict(list)
    for _, row in net.line.iterrows():
        fn = net.bus.loc[row["from_bus"], "name"].strip()
        tn = net.bus.loc[row["to_bus"],   "name"].strip()
        if fn not in geo_map or tn not in geo_map:
            continue
        corridor_lines[frozenset([fn, tn])].append(row)

    OFFSET_STEP = 0  

    def _make_line_corridor_popup(uid, rows_data, active_tab=0):
        n_tabs = len(rows_data)

        def _lcol(loading):
            return ("#e74c3c" if loading > 90 else
                    "#e67e22" if loading > 70 else "#1D9E75")

        tab_colors = [("#444444" if rd["is_outage"]
                       else _lcol(float(rd["res"]["loading_percent"])))
                      for rd in rows_data]

        cjs = "[" + ",".join(f"'{c}'" for c in tab_colors) + "]"
        tab_js = (
            "<script>"
            "function swl(uid,idx,n,colors){"
            "for(var k=0;k<n;k++){"
            "document.getElementById(uid+'p'+k).style.display=k==idx?'block':'none';"
            "var b=document.getElementById(uid+'b'+k);"
            "b.style.background=k==idx?colors[k]:'#eee';"
            "b.style.color=k==idx?'white':'#333';}"
            "}"
            f"window.addEventListener('load',function(){{swl('{uid}',{active_tab},{n_tabs},{cjs});}});"
            "</script>"
        )

        btn_html = ""
        for t, rd in enumerate(rows_data):
            lbl = f"Line&nbsp;{t+1}" if n_tabs > 1 else "Details"
            btn_html += (
                f"<button id='{uid}b{t}' onclick=\"swl('{uid}',{t},{n_tabs},{cjs})\" "
                f"style='flex:1;padding:5px 2px;font-size:10px;border:none;"
                f"cursor:pointer;border-radius:4px;background:#eee;color:#333;'>"
                f"{lbl}</button>"
            )

        panels_html = ""
        for t, rd in enumerate(rows_data):
            if rd["is_outage"]:
                badge      = "<span class='badge outage'>&#9889; OUT OF SERVICE</span>"
                rows_inner = ""
            else:
                res     = rd["res"]
                loading = float(res["loading_percent"])
                bcls    = "crit" if loading > 90 else "warn" if loading > 70 else "ok"
                badge   = f"<span class='badge {bcls}'>Loading: {loading:.1f}%</span>"
                rows_inner = (
                    f"<tr><td>P from</td><td>{float(res['p_from_mw']):.3f} MW</td></tr>"
                    f"<tr><td>P to</td><td>{float(res['p_to_mw']):.3f} MW</td></tr>"
                    f"<tr><td>Q from</td><td>{float(res['q_from_mvar']):.3f} Mvar</td></tr>"
                    f"<tr><td>Q to</td><td>{float(res['q_to_mvar']):.3f} Mvar</td></tr>"
                    f"<tr><td>I from</td><td>{float(res['i_from_ka']):.4f} kA</td></tr>"
                    f"<tr><td>Max I</td><td>{float(rd['max_i_ka']):.4f} kA</td></tr>"
                )

            circuit_label = (f"{rd['name']} #ckt{rd['circuit']}"
                             if n_tabs > 1 else rd["name"])
            panels_html += (
                f"<div id='{uid}p{t}' style='display:none;'>"
                f"<div class='lname'>{circuit_label}</div>"
                f"{badge}<table>"
                f"<tr><td>From</td><td>{_hl.escape(rd['fn'])}</td></tr>"
                f"<tr><td>To</td><td>{_hl.escape(rd['tn'])}</td></tr>"
                f"<tr><td>Subgrid</td><td>{_hl.escape(rd['folder'])}</td></tr>"
                f"<tr><td>Voltage</td><td>{rd['vn']:.0f} kV</td></tr>"
                f"<tr><td>Length</td><td>{rd['length_km']:.2f} km</td></tr>"
                f"{rows_inner}</table></div>"
            )

        plural    = "s" if n_tabs > 1 else ""
        full_html = (
            "<!DOCTYPE html><html><head><meta charset=utf-8><style>"
            "body{margin:0;padding:6px;font-family:sans-serif;font-size:11px}"
            "h3{margin:0 0 4px;font-size:12px;color:#222}"
            ".tabs{display:flex;gap:2px;margin-bottom:5px}"
            ".lname{font-weight:700;margin-bottom:4px;color:#333}"
            ".badge{display:inline-block;padding:2px 8px;border-radius:10px;"
            "font-size:10px;font-weight:bold;margin-bottom:5px;color:white}"
            ".ok{background:#1D9E75}.warn{background:#e67e22}"
            ".crit{background:#e74c3c}.outage{background:#444}"
            "table{width:100%;border-collapse:collapse}"
            "td{padding:2px 3px}"
            "td:first-child{color:#666;white-space:nowrap;width:44%}"
            "td:last-child{font-weight:600;color:#222}"
            "</style></head><body>"
            f"<h3>Corridor ({n_tabs} line{plural})</h3>"
            f"<div class='tabs'>{btn_html}</div>"
            f"{panels_html}{tab_js}</body></html>"
        )
        return full_html, tab_colors

    lc_counter = [0]
    for corr, rows_in_corr in corridor_lines.items():
        n_par = len(rows_in_corr)

        _rfn0 = net.bus.loc[rows_in_corr[0]["from_bus"], "name"].strip()
        _rtn0 = net.bus.loc[rows_in_corr[0]["to_bus"],   "name"].strip()
        _la0, _ln0 = geo_map[_rfn0]
        _la1, _ln1 = geo_map[_rtn0]
        shared_base_path = None
        for _row in rows_in_corr:
            for (_a0, _o0, _a1, _o1) in [(_la0, _ln0, _la1, _ln1),
                                           (_la1, _ln1, _la0, _ln0)]:
                _c = find_osm_coords(_row["name"], _a0, _o0, _a1, _o1)
                if _c is not None:
                    shared_base_path = _c
                    break
            if shared_base_path is not None:
                break
        if shared_base_path is None:
            shared_base_path = [[_la0, _ln0], [_la1, _ln1]]

        uid = f"lc{lc_counter[0]}"; lc_counter[0] += 1

        rows_data = []
        for i, row in enumerate(rows_in_corr):
            rfn = net.bus.loc[row["from_bus"], "name"].strip()
            rtn = net.bus.loc[row["to_bus"],   "name"].strip()
            is_outage = (row["name"] in outage_state
                         or not row.get("in_service", True))
            res = (None if is_outage
                   else net.res_line.loc[row.name]
                   if row.name in net.res_line.index else None)
            rows_data.append({
                "name":      row["name"],
                "circuit":   i + 1,
                "fn":        rfn,
                "tn":        rtn,
                "folder":    b2s.get(row["from_bus"], SUBGRIDS[0]),
                "vn":        float(net.bus.loc[row["from_bus"], "vn_kv"]),
                "length_km": float(row["length_km"]),
                "max_i_ka":  float(row["max_i_ka"]),
                "is_outage": is_outage,
                "res":       res,
                "row":       row,
            })

        for par_idx, rd in enumerate(rows_data):
            row       = rd["row"]
            is_outage = rd["is_outage"]
            is_tie    = row["name"] in tie_set
            vn        = rd["vn"]
            folder    = rd["folder"]

            offset_m = (par_idx - (n_par - 1) / 2.0) * OFFSET_STEP
            path     = offset_polyline(shared_base_path, offset_m)

            if is_outage:
                lc_load = "#222222"
                lc_volt = "#222222"
            else:
                loading = float(rd["res"]["loading_percent"])
                lc_load = loading_color(loading)
                lc_volt = vn_color(vn)

            w       = (5 if vn >= 200 else 4 if vn >= 100 else 2) + (1 if n_par > 1 else 0)
            w       = w + 4 if is_outage else w
            dash    = "10 6" if is_tie else ("8 5" if is_outage else None)
            opacity = 0.65 if is_outage else 0.9

            popup_html, _ = _make_line_corridor_popup(uid, rows_data,
                                                      active_tab=par_idx)
            b64p    = base64.b64encode(popup_html.encode("utf-8")).decode("ascii")
            popup_h = min(170 + n_par * 30 + (150 if not is_outage else 30), 360)
            iframe  = (f"<iframe src='data:text/html;base64,{b64p}' "
                       f"width='270' height='{popup_h}' style='border:none;'></iframe>")

            tip_txt = (
                f"<b>{_hl.escape(str(row['name']))}</b>"
                + (" &#9889; OUT OF SERVICE" if is_outage
                   else f"<br>Loading: {float(rd['res']['loading_percent']):.1f}%")
                + (f"<br><i>{n_par} parallel lines — click to switch</i>"
                   if n_par > 1 else "")
            )

            target_fg = fg_tie if is_tie else fg[folder]

            if is_outage:
                halo = folium.PolyLine(
                    path,
                    color="#ffaa00",
                    weight=w + 10,
                    dash_array=None,
                    opacity=0.35,
                )
                halo.options["className"] = "outage-halo"
                halo.add_to(target_fg)

            cls = ("outage-line line-el"
                   if is_outage
                   else f"line-el lc-load-{lc_load.lstrip('#')} lc-volt-{lc_volt.lstrip('#')}")
            pl  = folium.PolyLine(
                path,
                color=lc_load,
                weight=w,
                dash_array=dash,
                opacity=opacity,
                tooltip=folium.Tooltip(tip_txt, sticky=True),
                popup=folium.Popup(iframe, max_width=290),
            )
            pl.options["className"] = cls
            if is_outage:
                pl.add_to(fg_outage)
            else:
                pl.add_to(fg_tie if is_tie else fg[folder])

    for _, row in net.trafo.iterrows():
        hn  = net.bus.loc[row["hv_bus"], "name"]
        ln2 = net.bus.loc[row["lv_bus"], "name"]
        if hn not in geo_map or ln2 not in geo_map:
            continue
        folder = b2s.get(row["hv_bus"], SUBGRIDS[0])
        res    = net.res_trafo.loc[row.name]
        tip    = (f"<b>{row['name']}</b><br>"
                  f"HV: {hn} ({net.bus.loc[row['hv_bus'],'vn_kv']:.0f} kV)<br>"
                  f"LV: {ln2} ({net.bus.loc[row['lv_bus'],'vn_kv']:.0f} kV)<br>"
                  f"Loading: <b>{res['loading_percent']:.1f}%</b>")
        hla, hlo = geo_map[hn]
        lla, llo = geo_map[ln2]
        osm_t = find_osm_coords(row["name"], hla, hlo, lla, llo)
        tpath = osm_t if osm_t else [[hla, hlo], [lla, llo]]
        folium.PolyLine(tpath, color="#8e44ad", weight=3, dash_array="4 4",
                        opacity=0.8,
                        tooltip=folium.Tooltip(tip)).add_to(fg[folder])
        
    fg_hvdc = folium.FeatureGroup(name="HVDC Interconnections", show=True)

    HVDC_ROUTES = {
        "LVHVDC": {
            "name":     "Leyte-Luzon HVDC (LVHVDC)",
            "vis_bus":  "ORMOC 1",
            "ext_name": "NAGA (Luzon)",
            "ext_lat":  13.616446887161999,
            "ext_lng":  123.24066053743866,
            "waypoints": [
                [11.4273, 124.7223],
                [11.6963, 124.6729],
                [11.9061, 124.5245],
                [12.1102, 124.2993],
                [12.3250, 124.0686],
                [12.4913, 123.8928],
                [12.7593, 123.6790],
                [13.0431, 123.5197],
                [13.2303, 123.4758],
                [13.4121, 123.3988],
                [13.5349, 123.3329],
            ],
            "color":   "#AD0000",
            "voltage": "±350 kV DC",
            "length":  "~440 km",
        },
        "VMHVDC": {
            "name":     "Visayas-Mindanao HVDC (VMHVDC)",
            "vis_bus":  "DUMANJUG",
            "ext_name": "LALA (Mindanao)",
            "ext_lat":  7.911200846486274,
            "ext_lng":  123.81129955086477,
            "waypoints": [
                [9.50,   123.45],  
                [9.05,   123.52],  
                [8.55,   123.60],  
                [8.15,   123.68],  
            ],
            "color":   "#0D14A0",
            "voltage": "±350 kV DC",
            "length":  "~300 km",
        },
    }

    def _arrow_icon(color, rotate_deg):
        svg = (
            f'<svg xmlns="http://www.w3.org/2000/svg" width="24" height="24" '
            f'viewBox="0 0 24 24" style="transform:rotate({rotate_deg}deg);">'
            f'<polygon points="12,2 22,22 12,17 2,22" '
            f'fill="{color}" stroke="#0b0f14" stroke-width="1.5"/>'
            f'</svg>'
        )
        return folium.DivIcon(
            html=svg,
            icon_size=(24, 24),
            icon_anchor=(12, 12),
        )

    def _bearing(lat1, lng1, lat2, lng2):
        import math
        d_lng = math.radians(lng2 - lng1)
        lat1r  = math.radians(lat1)
        lat2r  = math.radians(lat2)
        x = math.sin(d_lng) * math.cos(lat2r)
        y = (math.cos(lat1r) * math.sin(lat2r)
             - math.sin(lat1r) * math.cos(lat2r) * math.cos(d_lng))
        return (math.degrees(math.atan2(x, y)) + 360) % 360

    def _arrow_marker(lat, lng, bearing, color, fg):
        folium.Marker(
            location=[lat, lng],
            icon=_arrow_icon(color, bearing),
            z_index_offset=1000,
        ).add_to(fg)

    for hvdc_key, route in HVDC_ROUTES.items():
        vis_bus = route["vis_bus"]
        if vis_bus not in geo_map:
            print(f"[HVDC map] Bus '{vis_bus}' not in geo_map, skipping.")
            continue

        vis_lat, vis_lng = geo_map[vis_bus]
        ext_lat, ext_lng = route["ext_lat"], route["ext_lng"]

        full_path = (
            [[vis_lat, vis_lng]]
            + route["waypoints"]
            + [[ext_lat, ext_lng]]
        )

        hvdc_p_mw = None
        for h in HVDC_CONNECTIONS:
            if h["name"] == hvdc_key:
                if h["is_slack"] and not net.res_ext_grid.empty:
                    idx = net.ext_grid[net.ext_grid["name"] == hvdc_key].index
                    if len(idx):
                        hvdc_p_mw = float(net.res_ext_grid.loc[idx[0], "p_mw"])
                elif not h["is_slack"] and not net.res_gen.empty:
                    idx = net.gen[net.gen["name"] == hvdc_key].index
                    if len(idx):
                        hvdc_p_mw = float(net.res_gen.loc[idx[0], "p_mw"])
                break

        direction_str = ""
        if hvdc_p_mw is not None:
            if hvdc_p_mw > 0.5:
                direction_str = (f"<br>Flow: <b>{hvdc_p_mw:.1f} MW"
                                 f" ← {route['ext_name']} (Import to Visayas)</b>")
            elif hvdc_p_mw < -0.5:
                direction_str = (f"<br>Flow: <b>{abs(hvdc_p_mw):.1f} MW"
                                 f" → {route['ext_name']} (Export from Visayas)</b>")
            else:
                direction_str = "<br>Flow: <b>~0 MW (standby)</b>"

        is_hvdc_slack = next((h["is_slack"] for h in HVDC_CONNECTIONS if h["name"] == hvdc_key), False)
        slack_badge   = "&nbsp;<b>[SLACK BUS]</b>" if is_hvdc_slack else ""
        tip = (
            f"<b>{route['name']}</b>{slack_badge}<br>"
            f"Visayas side: <b>{vis_bus}</b><br>"
            f"External side: <b>{route['ext_name']}</b><br>"
            f"Voltage: {route['voltage']}<br>"
            f"Length: {route['length']}<br>"
            f"Capacity: ±{max(abs(h['min_p_mw']), abs(h['max_p_mw'])):.0f} MW"
            f"{direction_str}"
        )

        folium.PolyLine(
            full_path,
            color=route["color"],
            weight=4,
            dash_array="12 6",
            opacity=0.9,
            tooltip=folium.Tooltip(tip),
        ).add_to(fg_hvdc)


        bearing_to_vis = _bearing(
            full_path[1][0], full_path[1][1],
            full_path[0][0], full_path[0][1],
        )
        _arrow_marker(vis_lat, vis_lng, bearing_to_vis, route["color"], fg_hvdc)

        bearing_to_ext = _bearing(
            full_path[-2][0], full_path[-2][1],
            full_path[-1][0], full_path[-1][1],
        )
        _arrow_marker(ext_lat, ext_lng, bearing_to_ext, route["color"], fg_hvdc)

        ext_tip = (
            f"<b>{route['ext_name']}</b><br>"
            f"HVDC external terminal"
            + (" &nbsp;<b>[SLACK BUS]</b>" if is_hvdc_slack else "") +
            f"<br>{route['voltage']}<br>"
            f"<i>Click for details</i>"
        )
        if hvdc_p_mw is not None:
            if hvdc_p_mw > 0.5:
                flow_label = f"{hvdc_p_mw:.1f} MW"
                flow_dir   = f"Export → {route['ext_name']}"
                flow_color = "#e74c3c"
            elif hvdc_p_mw < -0.5:
                flow_label = f"{abs(hvdc_p_mw):.1f} MW"
                flow_dir   = f"Import ← Visayas"
                flow_color = "#27ae60"
            else:
                flow_label = "~0 MW"
                flow_dir   = "Standby"
                flow_color = "#95a5a6"
        else:
            flow_label = "N/A"
            flow_dir   = "No results"
            flow_color = "#95a5a6"

        cap_mw = max(abs(next(h['min_p_mw'] for h in HVDC_CONNECTIONS if h['name'] == hvdc_key)),
                     abs(next(h['max_p_mw'] for h in HVDC_CONNECTIONS if h['name'] == hvdc_key)))

        if hvdc_p_mw is not None:
            if hvdc_p_mw > 0.5:
                flow_label = f"{hvdc_p_mw:.1f} MW"
                flow_dir   = f"← Import to Visayas from {route['ext_name']}"
                flow_color = "#27ae60"
            elif hvdc_p_mw < -0.5:
                flow_label = f"{abs(hvdc_p_mw):.1f} MW"
                flow_dir   = f"→ Export from Visayas to {route['ext_name']}"
                flow_color = "#e74c3c"
            else:
                flow_label = "~0 MW"
                flow_dir   = "Standby"
                flow_color = "#95a5a6"
        else:
            flow_label = "N/A"
            flow_dir   = "No results"
            flow_color = "#95a5a6"

        cap_mw = max(
            abs(next(h['min_p_mw'] for h in HVDC_CONNECTIONS if h['name'] == hvdc_key)),
            abs(next(h['max_p_mw'] for h in HVDC_CONNECTIONS if h['name'] == hvdc_key))
        )

        slack_row = (
            "<tr><td>Bus Type</td>"
            "<td><b style='color:#e74c3c;'>SLACK</b></td></tr>"
            if is_hvdc_slack else
            "<tr><td>Bus Type</td>"
            "<td><b style='color:#27ae60;'>PV (Controlled Gen)</b></td></tr>"
        )
        slack_header_badge = (
            "&nbsp;<span style='font-size:9px;color:#e74c3c;"
            "font-weight:700;vertical-align:middle;'>[SLACK]</span>"
            if is_hvdc_slack else ""
        )

        ext_popup_html = (
            "<!DOCTYPE html><html><head><meta charset=utf-8><style>"
            "body{margin:0;padding:5px;font-family:sans-serif;font-size:11px;}"
            "h3{margin:0 0 4px;font-size:12px;color:#222;}"
            ".bname{font-weight:600;margin-bottom:3px;color:#333;}"
            "table{width:100%;border-collapse:collapse;}"
            "td{padding:1px 3px;}"
            "td:first-child{color:#666;white-space:nowrap;}"
            "</style></head><body>"
            f"<h3>HVDC External Terminal{slack_header_badge}</h3>"
            f"<div class='bname'>{_hl.escape(route['ext_name'])}</div>"
            "<table>"
            f"{slack_row}"
            f"<tr><td>HVDC Link</td><td><b>{_hl.escape(route['name'])}</b></td></tr>"
            f"<tr><td>Visayas Bus</td><td><b>{_hl.escape(vis_bus)}</b></td></tr>"
            f"<tr><td>Voltage</td><td>{_hl.escape(route['voltage'])}</td></tr>"
            f"<tr><td>Length</td><td>{_hl.escape(route['length'])}</td></tr>"
            f"<tr><td>Capacity</td><td>&#177;{cap_mw:.0f} MW</td></tr>"
            f"<tr><td>Flow</td><td><b style='color:{flow_color};'>"
            f"{_hl.escape(flow_label)}</b></td></tr>"
            f"<tr><td>Direction</td><td style='color:{flow_color};'>"
            f"{_hl.escape(flow_dir)}</td></tr>"
            "</table>"
            "</body></html>"
        )

        b64_ext = base64.b64encode(ext_popup_html.encode("utf-8")).decode("ascii")
        ext_iframe = (f"<iframe src='data:text/html;base64,{b64_ext}' "
                      f"width='220' height='190' style='border:none;'></iframe>")

        folium.CircleMarker(
            location=[ext_lat, ext_lng],
            radius=14,
            color="rgba(0,0,0,0)",
            weight=0,
            fill=True,
            fill_color="rgba(0,0,0,0)",
            fill_opacity=0,
            tooltip=folium.Tooltip(ext_tip),
            popup=folium.Popup(ext_iframe, max_width=240),
        ).add_to(fg_hvdc)

        folium.Marker(
            location=[ext_lat, ext_lng],
            icon=folium.DivIcon(
                html=(
                    f'<div style="font-size:7px;font-weight:bold;color:white;'
                    f'text-align:center;line-height:1;margin-top:-5px;'
                    f'margin-left:-20px;pointer-events:none;width:40px;">'
                    f'{route["ext_name"].split("(")[0].strip()}</div>'
                ),
                icon_size=(40, 20),
                icon_anchor=(20, 10),
            ),
        ).add_to(fg_hvdc)

    bus_gen_mw   = defaultdict(float)   
    bus_load_mw  = defaultdict(float)   
    bus_lmp      = {}                    

    for i in net.gen.index:
        bi = net.gen.loc[i, "bus"]
        if not net.res_gen.empty and i in net.res_gen.index:
            bus_gen_mw[bi] += float(net.res_gen.loc[i, "p_mw"])

    for i in net.load.index:
        bi = net.load.loc[i, "bus"]
        if not net.res_load.empty and i in net.res_load.index:
            bus_load_mw[bi] += float(net.res_load.loc[i, "p_mw"])

    for bidx in net.bus.index:
        if not net.res_bus.empty and bidx in net.res_bus.index:
            res = net.res_bus.loc[bidx]
            if "lam_p" in res.index:
                bus_lmp[bidx] = float(res["lam_p"])

    max_gen_mw  = max(bus_gen_mw.values(),  default=1.0) or 1.0
    max_load_mw = max(bus_load_mw.values(), default=1.0) or 1.0

    def _dispatch_radius(mw, max_mw, min_r=4, max_r=22):
        if mw <= 0: return min_r
        return min_r + (max_r - min_r) * (mw / max_mw) ** 0.5

    cg = defaultdict(list)
    for bidx, brow in net.bus.iterrows():
        n = brow["name"]
        if n not in geo_map: continue
        la, ln3 = geo_map[n]
        cg[(round(la,6), round(ln3,6))].append((bidx, brow, n))

    mc = [0]
    for (la, ln3), bl in cg.items():
        bls      = sorted(bl, key=lambda x: x[1]["vn_kv"], reverse=True)
        top_idx, _, top_name = bls[0]
        n_tabs   = len(bls)
        uid      = "m" + str(mc[0]); mc[0] += 1
        tc       = [BUS_COLORS[btype(b[0])] for b in bls]
        folder   = b2s.get(top_idx, SUBGRIDS[0])

        loc_gen_mw  = sum(bus_gen_mw.get(b[0],  0) for b in bls)
        loc_load_mw = sum(bus_load_mw.get(b[0], 0) for b in bls)
        loc_lmp     = bus_lmp.get(top_idx, None)
        has_gen     = loc_gen_mw > 0.01
        has_load    = loc_load_mw > 0.01

        cjs = "[" + ",".join("'" + c + "'" for c in tc) + "]"
        btn = ""
        for t, (bidx, brow, bname) in enumerate(bls):
            vn     = int(brow["vn_kv"]); active = (t == 0)
            btn   += ("<button id='" + uid + "b" + str(t) + "' "
                      "onclick=\"sw('" + uid + "'," + str(t) + "," + str(n_tabs) + "," + cjs + ")\" "
                      "style='flex:1;padding:5px 2px;font-size:11px;border:none;cursor:pointer;"
                      "background:" + ("#1D9E75" if active else "#eee") + ";"
                      "color:" + ("white" if active else "#333") + ";border-radius:4px;'>"
                      + str(vn) + " kV</button>")

        panels = ""
        for t, (bidx, brow, bname) in enumerate(bls):
            res = net.res_bus.loc[bidx]; bt = btype(bidx); sg = b2s.get(bidx, "")
            d2  = "block" if t == 0 else "none"
            lr  = ""
            if "lam_p" in res.index:
                lr = "<tr><td>LMP</td><td><b>&#8369;" + "{:.2f}".format(float(res["lam_p"])) + "/MWh</b></td></tr>"
            rows = ("<tr><td>Subgrid</td><td><b>" + sg + "</b></td></tr>"
                    "<tr><td>Type</td><td><b>" + bt.upper() + "</b></td></tr>"
                    "<tr><td>Vn</td><td>" + str(int(brow["vn_kv"])) + "kV</td></tr>"
                    "<tr><td>Vm</td><td><b>" + "{:.4f}".format(float(res["vm_pu"])) + "pu</b></td></tr>"
                    "<tr><td>Va</td><td>" + "{:.2f}".format(float(res["va_degree"])) + "&deg;</td></tr>"
                    "<tr><td>P</td><td>" + "{:.3f}".format(float(res["p_mw"])) + "MW</td></tr>"
                    "<tr><td>Q</td><td>" + "{:.3f}".format(float(res["q_mvar"])) + "Mvar</td></tr>" + lr)
            panels += ("<div id='" + uid + "p" + str(t) + "' style='display:" + d2 + ";'>"
                       "<div class='bname'>" + _hl.escape(str(bname)) + "</div>"
                       "<table>" + rows + "</table></div>")

        tab_js = ("<script>function sw(uid,idx,n,colors){"
                  "for(var k=0;k<n;k++){document.getElementById(uid+'p'+k).style.display=k==idx?'block':'none';"
                  "var b=document.getElementById(uid+'b'+k);"
                  "b.style.background=k==idx?'#1D9E75':'#eee';b.style.color=k==idx?'white':'#333';}"
                  "window.parent.postMessage({type:'recolor',uid:uid,color:colors[idx]},'*');}"
                  "</script>")

        plural = "s" if n_tabs > 1 else ""
        fp = ("<!DOCTYPE html><html><head><meta charset=utf-8><style>"
              "body{margin:0;padding:5px;font-family:sans-serif;font-size:11px;}"
              "h3{margin:0 0 4px;font-size:12px;color:#222;}.tabs{display:flex;gap:2px;margin-bottom:5px;}"
              ".bname{font-weight:600;margin-bottom:3px;color:#333;}"
              "table{width:100%;border-collapse:collapse;}td{padding:1px 3px;}"
              "td:first-child{color:#666;white-space:nowrap;}"
              "</style></head><body><h3>Substation (" + str(n_tabs) + " voltage level" + plural + ")</h3>"
              "<div class='tabs'>" + btn + "</div>" + panels + tab_js + "</body></html>")

        lmp_str  = "&#8369;{:.2f}/MWh".format(loc_lmp) if loc_lmp is not None else "--"
        gen_rows = ""
        for i in net.gen.index:
            if net.gen.loc[i, "bus"] not in [b[0] for b in bls]: continue
            gn    = net.gen.loc[i, "name"]
            p_mw  = float(net.res_gen.loc[i, "p_mw"]) if i in net.res_gen.index else 0
            ptype = _get_plant_type(gn)
            gen_rows += (f"<tr><td>{_hl.escape(gn)}</td>"
                         f"<td style='color:#22c55e;font-weight:600'>{ptype}</td>"
                         f"<td style='text-align:right;font-weight:600'>{p_mw:.1f} MW</td></tr>")

        load_rows = ""
        for i in net.load.index:
            if net.load.loc[i, "bus"] not in [b[0] for b in bls]: continue
            ln   = net.load.loc[i, "name"]
            p_mw = float(net.res_load.loc[i, "p_mw"]) if i in net.res_load.index else 0
            load_rows += (f"<tr><td>{_hl.escape(ln)}</td>"
                          f"<td style='text-align:right;font-weight:600'>{p_mw:.1f} MW</td></tr>")

        dp_html = ("<!DOCTYPE html><html><head><meta charset=utf-8><style>"
                   "body{margin:0;padding:6px;font-family:sans-serif;font-size:11px;}"
                   "h3{margin:0 0 4px;font-size:12px;color:#222;}"
                   ".sec{font-weight:700;color:#555;margin:6px 0 2px;font-size:10px;text-transform:uppercase;}"
                   "table{width:100%;border-collapse:collapse;margin-bottom:4px;}"
                   "td{padding:1px 3px;}td:first-child{color:#666;}"
                   ".lmp{font-size:13px;font-weight:700;color:#3b82f6;margin-bottom:6px;}"
                   "</style></head><body>"
                   f"<h3>{_hl.escape(str(top_name))}</h3>"
                   f"<div class='lmp'>LMP: {lmp_str}</div>")
        if gen_rows:
            dp_html += ("<div class='sec'>&#9654; Generator Dispatch</div>"
                        "<table><tr><th style='text-align:left'>Unit</th>"
                        "<th style='text-align:left'>Type</th>"
                        "<th style='text-align:right'>Dispatch</th></tr>"
                        + gen_rows + "</table>")
        if load_rows:
            dp_html += ("<div class='sec'>&#9660; Load Demand</div>"
                        "<table><tr><th style='text-align:left'>Load</th>"
                        "<th style='text-align:right'>Demand</th></tr>"
                        + load_rows + "</table>")
        if not gen_rows and not load_rows:
            dp_html += "<div style='color:#999;font-size:10px'>No generation or load at this location.</div>"
        dp_html += "</body></html>"

        b64v    = base64.b64encode(fp.encode("utf-8")).decode("ascii")
        b64d    = base64.b64encode(dp_html.encode("utf-8")).decode("ascii")
        itag_v  = ("<iframe src='data:text/html;base64," + b64v + "' width='220' height='"
                   + str(min(140 + n_tabs * 16, 280)) + "' style='border:none;'></iframe>")
        itag_d  = ("<iframe src='data:text/html;base64," + b64d + "' width='240' height='"
                   + str(min(160 + (4 if gen_rows else 0) + (4 if load_rows else 0), 340))
                   + "' style='border:none;'></iframe>")

        vl = " / ".join(str(int(b[1]["vn_kv"])) + " kV" for b in bls)
        tt = _hl.escape(str(top_name)) + "<br>" + folder + " | " + vl + "<br><i>Click for details</i>"

        if has_gen and has_load:
            disp_color = "#a855f7"   
        elif has_gen:
            disp_color = "#22c55e"   
        else:
            disp_color = "#3b82f6"   

        disp_r_gen  = _dispatch_radius(loc_gen_mw,  max_gen_mw)
        disp_r_load = _dispatch_radius(loc_load_mw, max_load_mw)
        disp_r      = max(disp_r_gen, disp_r_load)

        cm = folium.CircleMarker(
            location=[la, ln3],
            radius=BUS_RADIUS[btype(top_idx)],
            color="white", weight=2, fill=True, fill_color=tc[0], fill_opacity=0.95,
            tooltip=folium.Tooltip(tt),
            popup=folium.Popup(itag_v, max_width=260))
        cm.options["className"] = uid
        cm.add_to(fg[folder])

        if has_gen or has_load:
            gen_str  = f"{loc_gen_mw:.1f} MW gen" if has_gen else ""
            load_str = f"{loc_load_mw:.1f} MW load" if has_load else ""
            parts_tt = [x for x in [gen_str, load_str] if x]
            disp_tt  = (_hl.escape(str(top_name)) + "<br>"
                        + " | ".join(parts_tt)
                        + (f"<br>LMP: &#8369;{loc_lmp:.2f}/MWh" if loc_lmp is not None else "")
                        + "<br><i>Click for dispatch details</i>")
            if has_gen:
                gen_tt = (_hl.escape(str(top_name)) + "<br>"
                          + f"{loc_gen_mw:.1f} MW dispatched"
                          + (f"<br>LMP: &#8369;{loc_lmp:.2f}/MWh" if loc_lmp is not None else "")
                          + "<br><i>Click for dispatch details</i>")
                cm_gen = folium.CircleMarker(
                    location=[la, ln3],
                    radius=disp_r_gen,
                    color="white", weight=1.5, fill=True,
                    fill_color="#22c55e", fill_opacity=0.85,
                    tooltip=folium.Tooltip(gen_tt),
                    popup=folium.Popup(itag_d, max_width=260))
                cm_gen.options["className"] = uid + "g"
                cm_gen.add_to(fg_gen_dispatch)

            if has_load:
                load_tt = (_hl.escape(str(top_name)) + "<br>"
                           + f"{loc_load_mw:.1f} MW demand"
                           + (f"<br>LMP: &#8369;{loc_lmp:.2f}/MWh" if loc_lmp is not None else "")
                           + "<br><i>Click for demand details</i>")
                cm_load = folium.CircleMarker(
                    location=[la, ln3],
                    radius=disp_r_load,
                    color="white", weight=1.5, fill=True,
                    fill_color="#3b82f6", fill_opacity=0.85,
                    tooltip=folium.Tooltip(load_tt),
                    popup=folium.Popup(itag_d, max_width=260))
                cm_load.options["className"] = uid + "l"
                cm_load.add_to(fg_load_demand)

        folium.Marker(
            location=[la, ln3],
            icon=folium.DivIcon(
                html="<div style='font-size:7px;font-weight:bold;color:white;"
                     "text-align:center;line-height:1;margin-top:-5px;margin-left:-4px;"
                     "pointer-events:none;'>"
                     + "/".join(str(b[0]) for b in bls) + "</div>",
                icon_size=(28, 20), icon_anchor=(14, 10)
            )).add_to(fg[folder])

    for f in SUBGRIDS: fg[f].add_to(m)
    fg_tie.add_to(m)
    fg_hvdc.add_to(m)
    fg_outage.add_to(m)
    fg_gen_dispatch.add_to(m)
    fg_load_demand.add_to(m)    
    folium.LayerControl(collapsed=False).add_to(m)
    
    outage_css = (
        "<style>"
        "@keyframes outage-pulse{"
        "0%  {stroke-opacity:1.0}"
        "50% {stroke-opacity:0.55}"
        "100%{stroke-opacity:1.0}"
        "}"
        ".outage-line{"
        "animation:outage-pulse 1.2s ease-in-out infinite;"
        "pointer-events:stroke!important;"
        "}"
        "@keyframes halo-pulse{"
        "0%  {stroke-opacity:0.40}"
        "50% {stroke-opacity:0.15}"
        "100%{stroke-opacity:0.40}"
        "}"
        ".outage-halo{"
        "animation:halo-pulse 1.2s ease-in-out infinite;"
        "pointer-events:none!important;"
        "}"
        "</style>"
    )
    recolor_js = (
        "<script>window.addEventListener('message',function(e){"
        "if(!e.data||e.data.type!=='recolor')return;"
        "var el=document.querySelector('path.'+e.data.uid);"
        "if(el){el.style.fill=e.data.color;el.style.fillOpacity='0.95';}"
        "});</script>"
    )
    hit_js = (
        "<style>.leaflet-marker-icon{pointer-events:none!important;}"
        ".leaflet-marker-shadow{pointer-events:none!important;}</style>"
        "<script>function expandHits(){"
        "document.querySelectorAll('.leaflet-overlay-pane path.leaflet-interactive')"
        ".forEach(function(p){var f=p.getAttribute('fill');"
        "if(f&&f!=='none'){p.style.strokeWidth='28px';p.style.strokeOpacity='0';"
        "p.style.paintOrder='stroke fill';}});}setTimeout(expandHits,1000);</script>"
    )
    toggle_panel = (
        '<div id="line-mode-panel" style="'
        'position:fixed;top:260px;right:10px;z-index:1000;'
        'background:white;padding:12px 16px;border-radius:8px;'
        'border:1px solid #ccc;font-size:12px;line-height:2;min-width:170px;">'
        '<b style="display:block;margin-bottom:6px;">Line color mode</b>'
        '<label style="display:flex;align-items:center;gap:8px;cursor:pointer;">'
        '<input type="radio" name="linemode" value="loading" checked '
        'onchange="setLineMode(\'loading\')">'
        '<span>Line loading level</span></label>'
        '<label style="display:flex;align-items:center;gap:8px;cursor:pointer;">'
        '<input type="radio" name="linemode" value="voltage" '
        'onchange="setLineMode(\'voltage\')">'
        '<span>Voltage level</span></label>'
        '<div id="legend-loading" style="margin-top:8px;border-top:1px solid #eee;padding-top:8px;">'
        '<b style="font-size:11px;">Loading</b><br>'
        '<span style="color:#1D9E75;font-size:16px;">&#9644;</span> &lt;70%<br>'
        '<span style="color:#e67e22;font-size:16px;">&#9644;</span> 70-90%<br>'
        '<span style="color:#e74c3c;font-size:16px;">&#9644;</span> &gt;90%'
        '</div>'
        '<div id="legend-voltage" style="margin-top:8px;border-top:1px solid #eee;padding-top:8px;display:none;">'
        '<b style="font-size:11px;">Voltage</b><br>'
        '<span style="color:#ff0000;font-size:16px;">&#9644;</span> 230 kV<br>'
        '<span style="color:#f1a542;font-size:16px;">&#9644;</span> 138 kV<br>'
        '<span style="color:#08f11c;font-size:16px;">&#9644;</span> 69 kV<br>'
        '<span style="color:#2c3e50;font-size:16px;">&#9146;</span> Tie line'
        '</div>'
        '</div>'
        '<script>'
        'var _lineMode="loading";'
        'function setLineMode(mode){'
        'document.getElementById("legend-loading").style.display=mode==="loading"?"block":"none";'
        'document.getElementById("legend-voltage").style.display=mode==="voltage"?"block":"none";'
        'document.querySelectorAll(".leaflet-overlay-pane path.line-el").forEach(function(p){'
        'var cls=p.getAttribute("class")||"";'
        'var rx=mode==="loading"?/lc-load-([0-9a-fA-F]{6})/:/ lc-volt-([0-9a-fA-F]{6})/;'
        'var mx=cls.match(rx);if(mx){p.style.stroke="#"+mx[1];}});}'
        '</script>'
    )
    left_legend = (
    '<div style="position:fixed;bottom:30px;left:30px;z-index:1000;'
    'background:white;padding:12px 16px;border-radius:8px;'
    'border:1px solid #ccc;font-size:12px;line-height:1.9;">'
    '<b>Bus type</b><br>'
    '<span style="color:red;">&#9679;</span> Slack<br>'
    '<span style="color:green;">&#9679;</span> PV (Generator)<br>'
    '<span style="color:blue;">&#9679;</span> PQ (Load)<br>'
    '<br><span style="color:#2c3e50;">&#9644;&#9644;</span> Tie line<br>'
    '<span style="color:#AD0000;">&#9644;&#9644;</span> LVHVDC (Luzon)<br>'
    '<span style="color:#0D14A0;">&#9644;&#9644;</span> VMHVDC (Mindanao)'
    '</div>'
    )
    reload_beacon = (
        "<script>"
        "(function(){"
        "var _ts=null;"
        "function _check(){"
        "fetch('/map_timestamp.txt?_='+Date.now())"
        ".then(function(r){return r.text();})"
        ".then(function(t){"
        "t=t.trim();"
        "if(_ts===null){_ts=t;}"
        "else if(t!==_ts){window.location.reload();}"
        "}).catch(function(){});"
        "}"
        "setInterval(_check,3000);"
        "})();"
        "</script>"
    )
    for jh in [recolor_js, hit_js, toggle_panel, left_legend, reload_beacon]:
        m.get_root().html.add_child(folium.Element(jh))
    auto_save = (not dpg.does_item_exist("auto_generate_map")
                 or dpg.get_value("auto_generate_map"))
    try:
        out     = os.path.join(state["base_dir"], MAP_FILE)
        out_tmp = out + ".tmp"
        m.save(out_tmp)   
        if auto_save:
            if os.path.exists(out):
                os.replace(out_tmp, out)
            else:
                os.rename(out_tmp, out)
            import time
            with open(os.path.join(state["base_dir"], "map_timestamp.txt"), "w") as _tf:
                _tf.write(str(time.time()))
            print(f"[map] Saved to {out}")
        else:
            if os.path.exists(out):
                os.replace(out_tmp, out)
            else:
                os.rename(out_tmp, out)
            print("[map] Built but auto-save skipped — file updated without timestamp beacon")
        state["map_ready"] = True
    except Exception as e:
        print(f"[map] ERROR saving map: {e}")
        print(traceback.format_exc())
def start_map_server():
    if state["map_server"]:
        try: state["map_server"].shutdown()
        except: pass
    class QH(__import__("http.server",fromlist=["SimpleHTTPRequestHandler"]).SimpleHTTPRequestHandler):
        def log_message(self,*a): pass
        def translate_path(self,path):
            return os.path.join(state["base_dir"],path.lstrip("/"))
    class Srv(socketserver.TCPServer):
        allow_reuse_address=True
    srv=Srv(("",MAP_PORT),QH); state["map_server"]=srv
    threading.Thread(target=srv.serve_forever,daemon=True).start()


# === UTILITY AND LOOKUP FUNCTIONS ===
def get_b2s():
    return {ci:f for (f,_),ci in state["bus_map"].items()}

def get_gen_folder():
    gf={}
    for f,(sg,_) in state["subgrid_nets"].items():
        for _,r in sg.gen.iterrows(): gf[r["name"]]=f
    return gf

def _get_gen_cost(gn):
    hvdc_names = {h["name"]: h for h in HVDC_CONNECTIONS if not h["is_slack"]}
    if gn in hvdc_names:
        return hvdc_names[gn]["cost"]
    sn = state.get("subgrid_nets")
    if not sn: return 0.0
    gf = get_gen_folder()
    folder = gf.get(gn)
    if not folder: return 0.0
    try:
        _, gc = sn[folder]
        row = gc[gc["name"] == gn].iloc[0]
        return float(row.get("c1", 0))
    except Exception:
        return 0.0

PLANT_TYPE_MAP = {
    "AMLA_G01": "Hydroelectric",
    "BBGPP_G01": "Unknown",
    "BDPP_U01": "Diesel", "BDPP_U02": "Diesel", "BDPP_U03": "Diesel", "BDPP_U04": "Diesel",
    "BIDPP_G01": "Diesel",
    "BILGPP": "Unknown",
    "BISCOM_G01": "Biomass",
    "CABI_G01": "Biomass",
    "CALASOL_G01": "Solar",
    "CARMENDPP_U01": "Diesel", "CARMENDPP_U02": "Diesel",
    "CARMENDPP_U03": "Diesel", "CARMENDPP_U04": "Diesel",
    "CARSOL_G01": "Solar",
    "CASA_G01": "Biomass", "CASA_G02": "Biomass",
    "CEDC_U01": "Coal", "CEDC_U02": "Coal", "CEDC_U03": "Coal",
    "CENPRI_U01": "Diesel", "CENPRI_U02": "Diesel", "CENPRI_U03": "Diesel",
    "CENPRI_U04": "Diesel", "CENPRI_U05": "Diesel",
    "CLBYBNK_G01": "Diesel",
    "COSMO_G01": "Solar",
    "CPPC_U01": "Diesel", "CPPC_U02": "Diesel", "CPPC_U03": "Diesel",
    "CPPC_U04": "Diesel", "CPPC_U05": "Diesel", "CPPC_U06": "Diesel",
    "CPPC_U07": "Diesel", "CPPC_U08": "Diesel", "CPPC_U09": "Diesel", "CPPC_U10": "Diesel",
    "DAGSOL_G01": "Solar",
    "EAUC_U01": "Diesel", "EAUC_U02": "Diesel", "EAUC_U03": "Diesel",
    "EAUC_U04": "Diesel", "EAUC_U05": "Diesel",
    "FFHC_G01": "Biomass",
    "HELIOS_G01": "Solar",
    "HPCO_G01": "Biomass", "HPCO_G02": "Biomass",
    "IASMOD_G01": "Diesel", "IASMOD_G02": "Diesel", "IASMOD_G03": "Diesel",
    "IASMOD_G04": "Diesel", "IASMOD_G05": "Diesel", "IASMOD_G06": "Diesel",
    "ISIDROSOL_G01": "Solar",
    "JANOPO_G01": "Hydroelectric",
    "KABAN_BAT": "BESS",
    "KSPC_G01": "Coal", "KSPC_G02": "Coal",
    "LEYTE_A": "Geothermal",
    "LGPP_G01": "Geothermal",
    "LOBOC_G01": "Hydroelectric", "LOBOC_G02": "Hydroelectric",
    "MANSOL_G01": "Solar",
    "MNTSOL_G01": "Solar",
    "NABASDPP_U01": "Diesel", "NABASDPP_U02": "Diesel",
    "NASULO_G01": "Geothermal",
    "NTNEGB_G01": "Biomass",
    "ORMOC_BAT": "BESS",
    "PAL1A_U01": "Geothermal", "PAL1A_U02": "Geothermal", "PAL1A_U03": "Geothermal",
    "PAL2A_U01": "Geothermal", "PAL2A_U02": "Geothermal",
    "PAL2A_U03": "Geothermal", "PAL2A_U04": "Geothermal",
    "PALM_G01": "Coal",
    "PDPP1_U02": "Diesel", "PDPP1_U03": "Diesel", "PDPP1_U05": "Diesel",
    "PDPP3_C": "Diesel", "PDPP3_E": "Diesel", "PDPP3_G": "Diesel", "PDPP3_H": "Diesel",
    "PEDC_U01": "Coal", "PEDC_U02": "Coal", "PEDC_U03": "Coal",
    "PHSOL_G01": "Solar",
    "PWIND_G01": "Wind", "PWIND_G02": "Wind",
    "SACASL_G01": "Solar", "SACASL_G02": "Solar",
    "SACSUN_G01": "Solar",
    "SCBE_G01": "Biomass",
    "SCBIOP_G01": "Biomass",
    "SEPSOL_G01": "Solar",
    "SEVILL_G01": "Hydroelectric",
    "SLWIND_G01": "Wind",
    "SONEG_BAT": "BESS",
    "STBPB1_U01": "Diesel", "STBPB1_U02": "Diesel",
    "STBPB1_U03": "Diesel", "STBPB1_U04": "Diesel",
    "STNEGB_G01": "Biomass",
    "SUWECO_G01": "Hydroelectric",
    "SYLSOL_G01": "Solar",
    "TAFTSOL_G01": "Solar",
    "TAFT_G01": "Solar",
    "THVI_U01": "Coal", "THVI_U02": "Coal",
    "TIMBA_G01": "Hydroelectric",
    "TOLEDO_BAT": "BESS",
    "TOLSOL_G01": "Solar",
    "TONGO_BAT": "Unknown",
    "TPC_G01": "Coal", "TPC_G02": "Coal",
    "TPLPB4_U01": "Diesel", "TPLPB4_U02": "Diesel",
    "TPLPB4_U03": "Diesel", "TPLPB4_U04": "Diesel",
    "TPVI_U01": "Diesel", "TPVI_U02": "Diesel", "TPVI_U03": "Diesel",
    "TPVI_U04": "Diesel", "TPVI_U05": "Diesel", "TPVI_U06": "Diesel",
    "UBAY_BAT": "BESS",
    "URC_G01": "Biomass",
    "SLYSOL_G01" : "Solar",
    "VISTASOL_G01" : "Solar",
    "UTH_G01": "Solar",
    "VITASOL_G01": "Solar",
    "VMC_G01": "Biomass", "VMC_G02": "Biomass",
    "LVHVDC": "HVDC",
    "VMHVDC": "HVDC",
}

def _get_plant_type(gn):
    import re
    sn = state.get("subgrid_nets")
    if sn:
        gf = get_gen_folder()
        folder = gf.get(gn)
        if not folder:
            clean_gn = re.sub(r"^\d+\s*", "", str(gn).strip())
            folder = gf.get(clean_gn)
        if folder:
            try:
                _, gc = sn[folder]
                row = gc[gc["name"] == gn]
                if row.empty:
                    clean_gn = re.sub(r"^\d+\s*", "", str(gn).strip())
                    row = gc[gc["name"].str.replace(r"^\d+\s*", "", regex=True) == clean_gn]
                if not row.empty and "plant_type" in gc.columns:
                    val = str(row.iloc[0]["plant_type"]).strip()
                    if val and val.lower() not in ("nan", "none", ""):
                        return val
            except Exception:
                pass
    clean = re.sub(r"^\d+\s*", "", str(gn).strip())
    return PLANT_TYPE_MAP.get(clean, PLANT_TYPE_MAP.get(gn, "--"))


# === GUI TABLE POPULATION ===
def fill_bus_table():
    net = state["net"]; b2s = get_b2s()
    eb = set(net.ext_grid["bus"].values); gb = set(net.gen["bus"].values)
    def btype(i): return "SLACK" if i in eb else ("PV" if i in gb else "PQ")
    sg_filter = filter_state["subgrid"]
    dpg.delete_item("bus_table", children_only=True)
    for col in ["Bus Name","Subgrid","Type","Vn (kV)","Vm (pu)","Va (deg)","P (MW)","Q (Mvar)","LMP (PHP/MWh)"]:
        dpg.add_table_column(label=col, parent="bus_table")
    for bidx, brow in net.bus.iterrows():
        sg = b2s.get(bidx, "")
        if sg_filter != "All" and sg != sg_filter: continue
        res = net.res_bus.loc[bidx]
        lmp = f"{res['lam_p']:.2f}" if "lam_p" in res.index else "--"
        with dpg.table_row(parent="bus_table"):
            for v in [brow["name"], sg, btype(bidx),
                      f"{brow['vn_kv']:.0f}", f"{res['vm_pu']:.4f}",
                      f"{res['va_degree']:.2f}", f"{res['p_mw']:.3f}",
                      f"{res['q_mvar']:.3f}", lmp]:
                dpg.add_text(v)

def fill_line_table():
    net=state["net"]; b2s=get_b2s()
    tie_set={t[3] for t in TIE_LINES}; sg_filter=filter_state["subgrid"]
    dpg.delete_item("line_table", children_only=True)
    for col in ["Line Name","Type","Subgrid","From Bus","To Bus","Vn (kV)","Loading (%)","P_from (MW)","Q_from (Mvar)","I_from (kA)"]:
        dpg.add_table_column(label=col, parent="line_table")
    for lidx, row in net.line.iterrows():
        fn  = net.bus.loc[row["from_bus"],"name"]
        tn  = net.bus.loc[row["to_bus"],  "name"]
        vn  = net.bus.loc[row["from_bus"],"vn_kv"]
        is_tie = row["name"] in tie_set
        sg  = b2s.get(row["from_bus"], "Tie") if not is_tie else "Tie"
        if sg_filter != "All" and sg != sg_filter: continue
        res = net.res_line.loc[lidx]
        lt  = "Tie" if is_tie else ("HV" if vn >= 110 else "MV")
        from_bus_name = net.bus.loc[row["from_bus"], "name"] if row["from_bus"] in net.bus.index else str(row["from_bus"])
        to_bus_name   = net.bus.loc[row["to_bus"],   "name"] if row["to_bus"]   in net.bus.index else str(row["to_bus"])
        with dpg.table_row(parent="line_table"):
            for v in [row["name"], lt, sg, from_bus_name, to_bus_name, f"{vn:.0f}",
                      f"{res['loading_percent']:.1f}",
                      f"{res['p_from_mw']:.3f}", f"{res['q_from_mvar']:.3f}",
                      f"{res['i_from_ka']:.4f}"]:
                dpg.add_text(v)

def fill_gen_table():
    net=state["net"]; b2s=get_b2s(); gf=get_gen_folder()
    sg_filter=filter_state["subgrid"]
    dpg.delete_item("gen_table", children_only=True)
    for col in ["Generator","Plant Type","Subgrid","Bus","Dispatch (MW)","Q (Mvar)","Cost (PHP/MWh)","LMP (PHP/MWh)"]:
        dpg.add_table_column(label=col, parent="gen_table")
    for i in net.gen.index:
        gn = net.gen.loc[i,"name"]
        if gn in {h["name"] for h in HVDC_CONNECTIONS if not h["is_slack"]}:
            sg = next(h["subgrid"] for h in HVDC_CONNECTIONS if h["name"] == gn)
        else:
            sg = gf.get(gn, "")
        if sg_filter != "All" and sg != sg_filter: continue
        res = net.res_gen.loc[i]; bi = net.gen.loc[i,"bus"]
        bus_name   = net.bus.loc[bi, "name"] if bi in net.bus.index else str(bi)
        cost       = _get_gen_cost(gn)
        plant_type = _get_plant_type(gn)
        lmp        = net.res_bus.loc[bi, "lam_p"] if "lam_p" in net.res_bus.columns and bi in net.res_bus.index else 0.0
        with dpg.table_row(parent="gen_table"):
            for v in [gn, plant_type, sg, bus_name, f"{res['p_mw']:.3f}", f"{res['q_mvar']:.3f}", f"{cost:.2f}", f"{lmp:.2f}"]:
                dpg.add_text(v)

def fill_load_table():
    net=state["net"]; b2s=get_b2s(); sg_filter=filter_state["subgrid"]
    dpg.delete_item("load_table", children_only=True)
    for col in ["Load","Subgrid","Bus","P (MW)","LMP (PHP/MWh)"]:
        dpg.add_table_column(label=col, parent="load_table")
    for i in net.load.index:
        bi = net.load.loc[i,"bus"]; sg = b2s.get(bi, "")
        if sg_filter != "All" and sg != sg_filter: continue
        res = net.res_load.loc[i]
        bus_name = net.bus.loc[bi, "name"] if bi in net.bus.index else str(bi)
        lmp = net.res_bus.loc[bi,"lam_p"] if "lam_p" in net.res_bus.columns else 0
        with dpg.table_row(parent="load_table"):
            for v in [net.load.loc[i,"name"], sg, bus_name, f"{res['p_mw']:.3f}", f"{lmp:.2f}"]:
                dpg.add_text(v)

def fill_all_tables():
    fill_bus_table(); fill_line_table(); fill_gen_table(); fill_load_table()


# === STATE RESET AND COMPARISON ===
def reset_overrides():
    mem_overrides.clear()
    outage_state.clear()
    if dpg.does_item_exist("outage_status"):
        dpg.set_value("outage_status", "No active outages.")
    _refresh_outage_list()
    comparison["modified_net"] = None
    comparison["has_modified"]  = False
    for tag in ["cmp_bus_table","cmp_line_table","cmp_gen_table","cmp_load_table"]:
        if dpg.does_item_exist(tag):
            dpg.delete_item(tag, children_only=True)
    if dpg.does_item_exist("upload_status"):
        dpg.set_value("upload_status", "Overrides cleared -- using default CSVs.")
    if dpg.does_item_exist("mod_status"):
        dpg.set_value("mod_status", "No active overrides.")
    if dpg.does_item_exist("active_overrides"):
        dpg.set_value("active_overrides", "None")
    _refresh_active_overrides_label()
    base_label = next((lbl for lbl, folder in _period_label_to_folder.items() if folder == "Base"), "Base")
    period_state["period"] = base_label
    if dpg.does_item_exist("period_combo"):
        dpg.set_value("period_combo", base_label)
    dpg.set_value("status_text", "Overrides and period cleared -- using base data.")
    _refresh_mod_status()

def reset_constraints():
    global user_constraints
    user_constraints = dict(CONSTRAINT_DEFAULTS)
    dpg.set_value("c_vm_min",    CONSTRAINT_DEFAULTS["vm_min_pu"])
    dpg.set_value("c_vm_max",    CONSTRAINT_DEFAULTS["vm_max_pu"])
    dpg.set_value("c_line_load", CONSTRAINT_DEFAULTS["max_line_loading"])
    dpg.set_value("c_trafo_load",CONSTRAINT_DEFAULTS["max_trafo_loading"])
    dpg.set_value("c_nr_iter",   int(CONSTRAINT_DEFAULTS["nr_max_iteration"]))
    dpg.set_value("c_opf_iter",  int(CONSTRAINT_DEFAULTS["opf_max_iteration"]))
    dpg.set_value("c_tie_r",     CONSTRAINT_DEFAULTS["tie_r_ohm_per_km"])
    dpg.set_value("c_tie_x",     CONSTRAINT_DEFAULTS["tie_x_ohm_per_km"])
    dpg.set_value("c_tie_c",     CONSTRAINT_DEFAULTS["tie_c_nf_per_km"])
    if dpg.does_item_exist("constraints_status"):
        dpg.set_value("constraints_status", "All constraints reset to defaults.")

def fill_comparison_tab():
    if not comparison["has_baseline"] or not comparison["has_modified"]:
        return
    bnet     = comparison["baseline_net"]
    mnet     = comparison["modified_net"]
    b2s      = {ci: f for (f,_), ci in state["bus_map"].items()}
    gf       = get_gen_folder()
    sg_filter = filter_state["subgrid"]

    def clr(val, ref, lower_is_better=False):
        try:
            v, r = float(val), float(ref)
            if abs(v - r) < 1e-4: return (200, 200, 200, 255)
            better = (v < r) if lower_is_better else (v > r)
            return (80, 200, 120, 255) if better else (220, 80, 80, 255)
        except Exception:
            return (180, 180, 180, 255)

    eb = set(bnet.ext_grid["bus"].values)
    gb = set(bnet.gen["bus"].values)
    def btype(i): return "SLACK" if i in eb else ("PV" if i in gb else "PQ")

    dpg.delete_item("cmp_bus_table", children_only=True)
    for col in ["Bus Name", "Subgrid", "Type", "Vn (kV)",
                "Vm base", "Vm mod", "Va base", "Va mod",
                "P base(MW)", "P mod(MW)", "Q base(Mvar)", "Q mod(Mvar)",
                "LMP base", "LMP mod"]:
        dpg.add_table_column(label=col, parent="cmp_bus_table")

    m_bus_lookup = {n: i for i, n in mnet.bus["name"].items()}
    for bidx_b, brow_b in bnet.bus.iterrows():
        sg = b2s.get(bidx_b, "")
        if sg_filter != "All" and sg != sg_filter:
            continue
        midx = m_bus_lookup.get(brow_b["name"])
        if midx is None or midx not in mnet.res_bus.index or bidx_b not in bnet.res_bus.index:
            continue
        br    = bnet.res_bus.loc[bidx_b]
        mr    = mnet.res_bus.loc[midx]
        b_lmp = float(br["lam_p"]) if "lam_p" in br.index else 0.0
        m_lmp = float(mr["lam_p"]) if "lam_p" in mr.index else 0.0
        vm_better = abs(float(mr["vm_pu"]) - 1.0) < abs(float(br["vm_pu"]) - 1.0)
        with dpg.table_row(parent="cmp_bus_table"):
            dpg.add_text(brow_b["name"])
            dpg.add_text(sg)
            dpg.add_text(btype(bidx_b))
            dpg.add_text(f"{float(brow_b['vn_kv']):.0f}")
            dpg.add_text(f"{float(br['vm_pu']):.4f}")
            dpg.add_text(f"{float(mr['vm_pu']):.4f}",
                         color=(80,200,120,255) if vm_better else (220,80,80,255))
            dpg.add_text(f"{float(br['va_degree']):.2f}")
            dpg.add_text(f"{float(mr['va_degree']):.2f}")
            dpg.add_text(f"{float(br['p_mw']):.3f}")
            dpg.add_text(f"{float(mr['p_mw']):.3f}",
                         color=clr(mr["p_mw"], br["p_mw"]))
            dpg.add_text(f"{float(br['q_mvar']):.3f}")
            dpg.add_text(f"{float(mr['q_mvar']):.3f}",
                         color=clr(mr["q_mvar"], br["q_mvar"]))
            dpg.add_text(f"{b_lmp:.2f}")
            dpg.add_text(f"{m_lmp:.2f}",
                         color=clr(m_lmp, b_lmp, lower_is_better=True))

    dpg.delete_item("cmp_line_table", children_only=True)
    for col in ["Line Name", "Type", "Subgrid", "From Bus", "To Bus", "Vn (kV)",
                "Loading base(%)", "Loading mod(%)",
                "P_from base(MW)", "P_from mod(MW)",
                "Q_from base(Mvar)", "Q_from mod(Mvar)"]:
        dpg.add_table_column(label=col, parent="cmp_line_table")

    tie_set      = {t[3] for t in TIE_LINES}
    m_line_lookup = {n: i for i, n in mnet.line["name"].items()}
    for lidx_b, lrow_b in bnet.line.iterrows():
        is_tie = lrow_b["name"] in tie_set
        sg     = b2s.get(lrow_b["from_bus"], "Tie") if not is_tie else "Tie"
        if sg_filter != "All" and sg != sg_filter:
            continue
        midx = m_line_lookup.get(lrow_b["name"])
        if midx is None:
            continue
        vn   = float(bnet.bus.loc[lrow_b["from_bus"], "vn_kv"])
        lt   = "Tie" if is_tie else ("HV" if vn >= 110 else "MV")
        fn   = bnet.bus.loc[lrow_b["from_bus"], "name"] if lrow_b["from_bus"] in bnet.bus.index else "--"
        tn   = bnet.bus.loc[lrow_b["to_bus"],   "name"] if lrow_b["to_bus"]   in bnet.bus.index else "--"

        b_load_str  = (f"{float(bnet.res_line.loc[lidx_b, 'loading_percent']):.1f}"
                       if lidx_b in bnet.res_line.index else "--")
        b_pfrom_str = (f"{float(bnet.res_line.loc[lidx_b, 'p_from_mw']):.3f}"
                       if lidx_b in bnet.res_line.index else "--")
        b_qfrom_str = (f"{float(bnet.res_line.loc[lidx_b, 'q_from_mvar']):.3f}"
                       if lidx_b in bnet.res_line.index else "--")

        is_outaged = lrow_b["name"] in outage_state
        if is_outaged:
            m_load_str  = "TRIPPED"
            m_load_col  = (220, 80, 80, 255)
            m_pfrom_str = "--"
            m_qfrom_str = "--"
            m_pfrom_col = (180, 180, 180, 255)
            m_qfrom_col = (180, 180, 180, 255)
        elif midx in mnet.res_line.index:
            m_load      = float(mnet.res_line.loc[midx, "loading_percent"])
            b_load      = float(bnet.res_line.loc[lidx_b, "loading_percent"]) \
                          if lidx_b in bnet.res_line.index else m_load
            m_load_str  = f"{m_load:.1f}"
            m_load_col  = clr(m_load, b_load, lower_is_better=True)
            m_pfrom     = float(mnet.res_line.loc[midx, "p_from_mw"])
            b_pfrom     = float(bnet.res_line.loc[lidx_b, "p_from_mw"]) \
                          if lidx_b in bnet.res_line.index else m_pfrom
            m_pfrom_str = f"{m_pfrom:.3f}"
            m_pfrom_col = clr(m_pfrom, b_pfrom, lower_is_better=True)
            m_qfrom     = float(mnet.res_line.loc[midx, "q_from_mvar"])
            b_qfrom     = float(bnet.res_line.loc[lidx_b, "q_from_mvar"]) \
                          if lidx_b in bnet.res_line.index else m_qfrom
            m_qfrom_str = f"{m_qfrom:.3f}"
            m_qfrom_col = clr(m_qfrom, b_qfrom, lower_is_better=True)
        else:
            m_load_str  = "--"; m_load_col  = (180,180,180,255)
            m_pfrom_str = "--"; m_pfrom_col = (180,180,180,255)
            m_qfrom_str = "--"; m_qfrom_col = (180,180,180,255)

        with dpg.table_row(parent="cmp_line_table"):
            dpg.add_text(lrow_b["name"])
            dpg.add_text(lt)
            dpg.add_text(sg)
            dpg.add_text(fn)
            dpg.add_text(tn)
            dpg.add_text(f"{vn:.0f}")
            dpg.add_text(b_load_str)
            dpg.add_text(m_load_str,  color=m_load_col)
            dpg.add_text(b_pfrom_str)
            dpg.add_text(m_pfrom_str, color=m_pfrom_col)
            dpg.add_text(b_qfrom_str)
            dpg.add_text(m_qfrom_str, color=m_qfrom_col)

    dpg.delete_item("cmp_gen_table", children_only=True)
    for col in ["Generator", "Plant Type", "Subgrid", "Bus",
                "P base(MW)", "P mod(MW)",
                "Q base(Mvar)", "Q mod(Mvar)",
                "LMP base(PHP)", "LMP mod(PHP)"]:
        dpg.add_table_column(label=col, parent="cmp_gen_table")

    m_gen_lookup = {n: i for i, n in mnet.gen["name"].items()}
    for gidx_b, grow_b in bnet.gen.iterrows():
        gn   = grow_b["name"]
        if gn in {h["name"] for h in HVDC_CONNECTIONS if not h["is_slack"]}:
            sg = next(h["subgrid"] for h in HVDC_CONNECTIONS if h["name"] == gn)
        else:
            sg = gf.get(gn, "")
        if sg_filter != "All" and sg != sg_filter:
            continue
        midx = m_gen_lookup.get(gn)
        if midx is None or midx not in mnet.res_gen.index:
            continue
        br    = bnet.res_gen.loc[gidx_b]
        mr    = mnet.res_gen.loc[midx]
        bi_b  = bnet.gen.loc[gidx_b, "bus"]
        bi_m  = mnet.gen.loc[midx, "bus"]
        bname = bnet.bus.loc[bi_b, "name"] if bi_b in bnet.bus.index else str(bi_b)
        b_lmp = (float(bnet.res_bus.loc[bi_b, "lam_p"])
                 if "lam_p" in bnet.res_bus.columns and bi_b in bnet.res_bus.index else 0.0)
        m_lmp = (float(mnet.res_bus.loc[bi_m, "lam_p"])
                 if "lam_p" in mnet.res_bus.columns and bi_m in mnet.res_bus.index else 0.0)
        with dpg.table_row(parent="cmp_gen_table"):
            dpg.add_text(gn)
            dpg.add_text(_get_plant_type(gn))
            dpg.add_text(sg)
            dpg.add_text(bname)
            dpg.add_text(f"{float(br['p_mw']):.3f}")
            dpg.add_text(f"{float(mr['p_mw']):.3f}",
                         color=clr(mr["p_mw"], br["p_mw"]))
            dpg.add_text(f"{float(br['q_mvar']):.3f}")
            dpg.add_text(f"{float(mr['q_mvar']):.3f}",
                         color=clr(mr["q_mvar"], br["q_mvar"]))
            dpg.add_text(f"{b_lmp:.2f}")
            dpg.add_text(f"{m_lmp:.2f}",
                         color=clr(m_lmp, b_lmp, lower_is_better=True))

    dpg.delete_item("cmp_load_table", children_only=True)
    for col in ["Load", "Subgrid", "Bus",
                "P base(MW)", "P mod(MW)",
                "LMP base(PHP)", "LMP mod(PHP)"]:
        dpg.add_table_column(label=col, parent="cmp_load_table")

    m_load_lookup = {n: i for i, n in mnet.load["name"].items()}
    for lidx_b, lrow_b in bnet.load.iterrows():
        ln  = lrow_b["name"]
        bi_b = bnet.load.loc[lidx_b, "bus"]
        sg   = b2s.get(bi_b, "")
        if sg_filter != "All" and sg != sg_filter:
            continue
        if lidx_b not in bnet.res_load.index:
            continue
        midx = m_load_lookup.get(ln)
        if midx is None or midx not in mnet.res_load.index:
            continue
        br    = bnet.res_load.loc[lidx_b]
        mr    = mnet.res_load.loc[midx]
        bi_m  = mnet.load.loc[midx, "bus"]
        bname = bnet.bus.loc[bi_b, "name"] if bi_b in bnet.bus.index else str(bi_b)
        b_lmp = (float(bnet.res_bus.loc[bi_b, "lam_p"])
                 if "lam_p" in bnet.res_bus.columns and bi_b in bnet.res_bus.index else 0.0)
        m_lmp = (float(mnet.res_bus.loc[bi_m, "lam_p"])
                 if "lam_p" in mnet.res_bus.columns and bi_m in mnet.res_bus.index else 0.0)
        with dpg.table_row(parent="cmp_load_table"):
            dpg.add_text(ln)
            dpg.add_text(sg)
            dpg.add_text(bname)
            dpg.add_text(f"{float(br['p_mw']):.3f}")
            dpg.add_text(f"{float(mr['p_mw']):.3f}")
            dpg.add_text(f"{b_lmp:.2f}")
            dpg.add_text(f"{m_lmp:.2f}",
                         color=clr(m_lmp, b_lmp, lower_is_better=True))

    _update_comparison_header()


# === EVENT HANDLERS AND UPLOAD LOGIC ===
def on_filter_change(sender, app_data):
    filter_state["subgrid"] = app_data
    if state["net"] is not None:
        fill_all_tables()
    if comparison["has_baseline"] and comparison["has_modified"]:
        fill_comparison_tab()

def refresh_period_list():
    periods = discover_periods()
    if dpg.does_item_exist("period_combo"):
        dpg.configure_item("period_combo", items=periods)

def _refresh_active_overrides_label():
    if not dpg.does_item_exist("active_overrides_label"):
        return
    if mem_overrides:
        lines_list = [f"  {sg}/{comp}" for (sg, comp) in sorted(mem_overrides)]
        dpg.set_value("active_overrides_label", "\n".join(lines_list))
    else:
        dpg.set_value("active_overrides_label", "  None")

def handle_upload(sender, app_data):
    selections = app_data.get("selections", {})
    if not selections:
        upload_state["status"] = "No file selected."
        _refresh_upload_status(); return
    src_path = list(selections.values())[0]
    sg = upload_state["subgrid"]; comp = upload_state["component"]
    required_cols = {
        "gens.csv":          {"bus", "p_mw", "name"},
        "loads.csv":         {"bus", "p_mw", "name"},
        "shunts.csv":        {"bus", "q_mvar", "name"},
        "gen_costs.csv":     {"name", "plant_type", "min_p_mw", "max_p_mw", "dep_capacity",
                              "min_q_mvar", "max_q_mvar", "p1", "p2", "c1"},
        "lines.csv":         {"from_bus", "to_bus", "length_km", "r_ohm_per_km",
                              "x_ohm_per_km", "c_nf_per_km", "max_i_ka", "name"},
        "transformers.csv":  {"hv_bus", "lv_bus", "std_type", "name"},
        "std_types.csv":     {"name", "sn_mva", "vn_hv_kv", "vn_lv_kv",
                              "vk_percent", "vkr_percent", "pfe_kw",
                              "i0_percent", "shift_degree"},
        "buses.csv":         {"name", "vn_kv"},
        "geo_coords.csv":    {"bus", "lat", "lng"},
        "ext_grid.csv":      {"bus", "vm_pu", "name"},
        "ext_grid_costs.csv":{"name", "min_p_mw", "max_p_mw", "cost_php_per_mwh"},
    }
    no_schema = {"ext_grid.csv", "ext_grid_costs.csv",
                 "hvdc.csv", "visayas_line_vertices.csv"}
    try:
        df = pd.read_csv(src_path); df.columns = df.columns.str.strip()
        if comp in required_cols and comp not in no_schema:
            missing = required_cols[comp] - set(df.columns)
            if missing:
                upload_state["status"] = (
                    f"ERROR: missing columns: {', '.join(sorted(missing))}\n"
                    f"Required: {', '.join(sorted(required_cols[comp]))}"
                )
                _refresh_upload_status(); return

        if sg == "VISAYAS":
            if comp == "hvdc.csv":
                out_path = os.path.join(period_base_dir(), comp)
            elif comp == "visayas_line_vertices.csv":
                out_path = os.path.join(state["base_dir"], comp)
            else:
                out_path = None
            if out_path:
                df.to_csv(out_path, index=False)
                upload_state["status"] = (
                    f"Saved {comp} to disk at:\n{out_path}\n"
                    f"({len(df)} rows). Run Simulation to apply."
                )
                _refresh_upload_status()
                return
        mem_overrides[(sg, comp)] = df
        active = [f"{s}/{c}" for (s,c) in mem_overrides]
        upload_state["status"] = (
            f"Loaded {sg}/{comp} into memory ({len(df)} rows). TEMPORARY.\n"
            f"Active overrides: {', '.join(active)}\n"
            f"Run Simulation to apply.")
        upload_state["preview_df"] = df
        _rebuild_preview_table(df)
        if dpg.does_item_exist("mod_status"):
            dpg.set_value("mod_status", f"Overrides: {', '.join(active)}" if active else "")
        if dpg.does_item_exist("active_overrides"):
            dpg.set_value("active_overrides",
                "\n".join(f"  {s} / {c}" for (s,c) in mem_overrides) or "None")
        _refresh_active_overrides_label()
    except Exception as e:
        upload_state["status"] = f"ERROR: {e}"
    _refresh_upload_status()
    _refresh_mod_status()

def _refresh_upload_status():
    if dpg.does_item_exist("upload_status"):
        dpg.set_value("upload_status", upload_state["status"])

def _rebuild_preview_table(df, remaining=0):
    if not dpg.does_item_exist("preview_table"):
        return
    df = df.loc[:, ~df.columns.str.startswith("Unnamed")]
    dpg.delete_item("preview_table", children_only=True)
    for col in df.columns:
        dpg.add_table_column(label=col, parent="preview_table",
                             width_fixed=True, init_width_or_weight=120)
    for _, row in df.iterrows():
        with dpg.table_row(parent="preview_table"):
            for v in row:
                dpg.add_text("" if pd.isna(v) else str(v))
    if remaining > 0:
        with dpg.table_row(parent="preview_table"):
            dpg.add_text(f"... {remaining} more rows not shown ...",
                         color=(120, 120, 120, 200))

def open_file_dialog():
    dpg.show_item("upload_file_dialog")

def on_upload_sg_change(sender, app_data):
    upload_state["subgrid"] = app_data
    if app_data == "VISAYAS":
        dpg.configure_item("upload_comp_combo", items=GLOBAL_UPLOADABLE,
                           default_value=GLOBAL_UPLOADABLE[0])
        upload_state["component"] = GLOBAL_UPLOADABLE[0]
    else:
        dpg.configure_item("upload_comp_combo", items=ALL_UPLOADABLE,
                           default_value=ALL_UPLOADABLE[0])
        upload_state["component"] = ALL_UPLOADABLE[0]
    _update_existing_info()

def on_upload_comp_change(sender, app_data):
    upload_state["component"] = app_data
    _update_existing_info()

def _update_existing_info():
    sg   = upload_state["subgrid"]
    comp = upload_state["component"]

    if sg == "VISAYAS":
        if comp == "hvdc.csv":
            path = os.path.join(period_base_dir(), comp)
        elif comp == "visayas_line_vertices.csv":
            path = output_path(comp)
            if not os.path.exists(path):
                path = resource_path(os.path.join("visualization", comp))
        else:
            path = None

        if path and os.path.exists(path):
            try:
                if comp == "visayas_line_vertices.csv":
                    df_count = pd.read_csv(path, usecols=[0])
                    total_rows = len(df_count)
                    df = pd.read_csv(path, nrows=20)
                    df = df.loc[:, ~df.columns.str.startswith("Unnamed")]
                    remaining = max(0, total_rows - 20)
                    info = (f"Global file: {comp}  ({total_rows} total rows)"
                            f"\nShowing first 20 rows.  Path: {path}")
                    upload_state["preview_df"] = df
                    _rebuild_preview_table(df, remaining=remaining)
                else:
                    df = pd.read_csv(path)
                    df = df.loc[:, ~df.columns.str.startswith("Unnamed")]
                    info = (f"Global file: {comp}  "
                            f"({len(df)} rows, {len(df.columns)} cols)"
                            f"\nPath: {path}")
                    upload_state["preview_df"] = df
                    _rebuild_preview_table(df)
            except Exception as e:
                info = f"Could not read {comp}: {e}"
                dpg.delete_item("preview_table", children_only=True)
        else:
            info = f"Global file: {comp} — not found (will be created on upload)"
            dpg.delete_item("preview_table", children_only=True)

        if dpg.does_item_exist("existing_info"):
            dpg.set_value("existing_info", info)
        if dpg.does_item_exist("upload_status"):
            dpg.set_value("upload_status", "")
        return

    if (sg, comp) in mem_overrides:
        df = mem_overrides[(sg, comp)]
        info = (f"[In-memory override] {sg}/{comp}  "
                f"({len(df)} rows, {len(df.columns)} cols) -> not yet saved to disk")
        upload_state["preview_df"] = df
        _rebuild_preview_table(df)
        if dpg.does_item_exist("existing_info"):
            dpg.set_value("existing_info", info)
        if dpg.does_item_exist("upload_status"):
            dpg.set_value("upload_status", "")
        return

    label      = period_state["period"]
    raw_folder = _period_label_to_folder.get(label, "Base")

    if raw_folder != "Base":
        path = os.path.join(resource_path(raw_folder), sg, comp)
    else:
        path = os.path.join(resource_path("visualization"), sg, comp)

    if os.path.exists(path):
        try:
            df = pd.read_csv(path)
            df = df.loc[:, ~df.columns.str.startswith("Unnamed")]
            info = (f"Current file: {sg}/{comp}  "
                    f"({len(df)} rows, {len(df.columns)} cols)"
                    f"\nPeriod: {label}  |  Path: {path}")
            upload_state["preview_df"] = df
            _rebuild_preview_table(df)
        except Exception as e:
            info = f"Could not read current file: {e}"
            dpg.delete_item("preview_table", children_only=True)
    else:
        info = f"No file found for {sg}/{comp} in period '{label}'"
        dpg.delete_item("preview_table", children_only=True)

    if dpg.does_item_exist("existing_info"):
        dpg.set_value("existing_info", info)
        if dpg.does_item_exist("upload_status"):
            dpg.set_value("upload_status", "")
        return

    path = os.path.join(period_base_dir(), sg, comp)
    if os.path.exists(path):
        try:
            df = pd.read_csv(path)
            df = df.loc[:, ~df.columns.str.startswith("Unnamed")]
            info = f"Current file: {sg}/{comp}  ({len(df)} rows, {len(df.columns)} cols)"
            upload_state["preview_df"] = df
            _rebuild_preview_table(df)
        except Exception as e:
            info = f"Could not read current file: {e}"
    else:
        info = f"No existing file at {sg}/{comp}"
        dpg.delete_item("preview_table", children_only=True)
    if dpg.does_item_exist("existing_info"):
        dpg.set_value("existing_info", info)


# === OUTAGE MANAGEMENT ===
def _refresh_outage_list():
    if not dpg.does_item_exist("outage_list_container"):
        return
    dpg.delete_item("outage_list_container", children_only=True)
    net = state["net"]
    if net is None:
        dpg.add_text("Run simulation first to load line list.",
                     parent="outage_list_container", color=(160,160,160))
        return
    search = dpg.get_value("outage_search").strip().lower() if dpg.does_item_exist("outage_search") else ""
    b2s = {ci: f for (f,_), ci in state["bus_map"].items()}
    sg_filter = dpg.get_value("outage_sg_filter") if dpg.does_item_exist("outage_sg_filter") else "All"

    shown = 0
    for _, row in net.line.iterrows():
        name = row["name"]
        sg   = b2s.get(row["from_bus"], "")
        if sg_filter != "All" and sg != sg_filter:
            continue
        if search and search not in name.lower():
            continue
        checked = name in outage_state
        tag = f"outage_chk_{name}"
        with dpg.group(horizontal=True, parent="outage_list_container"):
            dpg.add_checkbox(label="", tag=tag, default_value=checked,
                             callback=_on_outage_toggle, user_data=name)
            color = (220, 80, 80, 255) if checked else (200, 200, 200, 255)
            lbl_tag = f"outage_lbl_{name}"
            if dpg.does_item_exist(lbl_tag):
                dpg.delete_item(lbl_tag)
            dpg.add_text(f"[{sg}]  {name}", tag=lbl_tag, color=color)
        shown += 1

    if shown == 0:
        dpg.add_text("No lines match filter.", parent="outage_list_container",
                     color=(160,160,160))
    _update_outage_status()

def _on_outage_toggle(sender, app_data, user_data):
    line_name = user_data
    if app_data:
        outage_state.add(line_name)
    else:
        outage_state.discard(line_name)
    _update_outage_status()
    _recolor_outage_label(line_name, app_data)
    _refresh_mod_status()

def _recolor_outage_label(line_name, checked):
    tag = f"outage_lbl_{line_name}"
    if dpg.does_item_exist(tag):
        color = (220, 80, 80, 255) if checked else (200, 200, 200, 255)
        dpg.configure_item(tag, color=color)

def _refresh_mod_status():
    parts = []
    if mem_overrides:
        parts.append(f"{len(mem_overrides)} override(s)")
    if outage_state:
        parts.append(f"{len(outage_state)} outage(s)")
    label = "Active: " + ", ".join(parts) if parts else ""
    if dpg.does_item_exist("mod_status"):
        dpg.set_value("mod_status", label)

def _update_outage_status():
    if not dpg.does_item_exist("outage_status"):
        return
    if outage_state:
        names = ", ".join(sorted(outage_state))
        dpg.set_value("outage_status",
                      f"{len(outage_state)} line(s) tripped: {names}")
    else:
        dpg.set_value("outage_status", "No active outages.")

def _on_outage_search(sender, app_data):
    _refresh_outage_list()
    
def _on_tab_change(sender, app_data):
    if app_data == "tab_modify":
        _update_existing_info()


# === DEARPYGUI SETUP AND THEME ===
dpg.create_context()
dpg.create_viewport(title="Open Source Visayas Grid Model Simulator",
                    width=1400, height=860, resizable=True)
dpg.setup_dearpygui()

with dpg.theme() as global_theme:
    with dpg.theme_component(dpg.mvAll):
        dpg.add_theme_color(dpg.mvThemeCol_WindowBg,         (0, 0, 0, 255))
        dpg.add_theme_color(dpg.mvThemeCol_ChildBg,          (0, 0, 0, 255))
        dpg.add_theme_color(dpg.mvThemeCol_PopupBg,          (10, 10, 10, 255))
        dpg.add_theme_color(dpg.mvThemeCol_FrameBg,          (18, 18, 18, 255))
        dpg.add_theme_color(dpg.mvThemeCol_FrameBgHovered,   (35, 35, 35, 255))
        dpg.add_theme_color(dpg.mvThemeCol_FrameBgActive,    (50, 50, 50, 255))
        dpg.add_theme_color(dpg.mvThemeCol_TitleBg,          (0, 0, 0, 255))
        dpg.add_theme_color(dpg.mvThemeCol_TitleBgActive,    (0, 0, 0, 255))
        dpg.add_theme_color(dpg.mvThemeCol_MenuBarBg,        (0, 0, 0, 255))
        dpg.add_theme_color(dpg.mvThemeCol_ScrollbarBg,      (0, 0, 0, 255))
        dpg.add_theme_color(dpg.mvThemeCol_ScrollbarGrab,    (50, 80, 120, 255))
        dpg.add_theme_color(dpg.mvThemeCol_ScrollbarGrabHovered, (70, 110, 160, 255))
        dpg.add_theme_color(dpg.mvThemeCol_ScrollbarGrabActive,  (90, 140, 200, 255))

        dpg.add_theme_color(dpg.mvThemeCol_Button,           (30, 60, 100, 255))
        dpg.add_theme_color(dpg.mvThemeCol_ButtonHovered,    (50, 100, 160, 255))
        dpg.add_theme_color(dpg.mvThemeCol_ButtonActive,     (70, 130, 200, 255))

        dpg.add_theme_color(dpg.mvThemeCol_FrameBg,          (20, 40, 65, 255))
        dpg.add_theme_color(dpg.mvThemeCol_FrameBgHovered,   (35, 65, 100, 255))
        dpg.add_theme_color(dpg.mvThemeCol_FrameBgActive,    (50, 90, 140, 255))

        dpg.add_theme_color(dpg.mvThemeCol_CheckMark,        (100, 180, 255, 255))
        dpg.add_theme_color(dpg.mvThemeCol_SliderGrab,       (70, 130, 200, 255))
        dpg.add_theme_color(dpg.mvThemeCol_SliderGrabActive, (100, 160, 240, 255))

        dpg.add_theme_color(dpg.mvThemeCol_Tab,              (15, 35, 60, 255))
        dpg.add_theme_color(dpg.mvThemeCol_TabHovered,       (40, 90, 150, 255))
        dpg.add_theme_color(dpg.mvThemeCol_TabActive,        (50, 120, 200, 255))
        dpg.add_theme_color(dpg.mvThemeCol_TabUnfocused,     (10, 25, 45, 255))
        dpg.add_theme_color(dpg.mvThemeCol_TabUnfocusedActive, (35, 85, 140, 255))

        dpg.add_theme_color(dpg.mvThemeCol_Header,           (25, 55, 90, 255))
        dpg.add_theme_color(dpg.mvThemeCol_HeaderHovered,    (40, 85, 135, 255))
        dpg.add_theme_color(dpg.mvThemeCol_HeaderActive,     (55, 115, 180, 255))

        dpg.add_theme_color(dpg.mvThemeCol_TableRowBg,       (0, 0, 0, 255))
        dpg.add_theme_color(dpg.mvThemeCol_TableRowBgAlt,    (10, 20, 35, 255))
        dpg.add_theme_color(dpg.mvThemeCol_TableBorderLight, (30, 55, 85, 255))
        dpg.add_theme_color(dpg.mvThemeCol_TableBorderStrong,(45, 80, 120, 255))

        dpg.add_theme_color(dpg.mvThemeCol_ResizeGrip,       (40, 90, 150, 100))
        dpg.add_theme_color(dpg.mvThemeCol_ResizeGripHovered,(60, 120, 190, 180))
        dpg.add_theme_color(dpg.mvThemeCol_ResizeGripActive, (80, 150, 230, 255))

        dpg.add_theme_color(dpg.mvThemeCol_Text,             (220, 220, 220, 255))
        dpg.add_theme_color(dpg.mvThemeCol_TextDisabled,     (90, 90, 90, 255))
        dpg.add_theme_color(dpg.mvThemeCol_Separator,        (35, 65, 100, 255))
        dpg.add_theme_color(dpg.mvThemeCol_SeparatorHovered, (60, 110, 170, 255))
        dpg.add_theme_color(dpg.mvThemeCol_SeparatorActive,  (80, 140, 210, 255))

        dpg.add_theme_color(dpg.mvThemeCol_NavHighlight,     (60, 130, 210, 255))

        dpg.add_theme_color(dpg.mvThemeCol_Border,           (0, 0, 0, 255))
        dpg.add_theme_color(dpg.mvThemeCol_BorderShadow,     (0, 0, 0, 255))

        dpg.add_theme_style(dpg.mvStyleVar_FrameBorderSize,  1.5)
        dpg.add_theme_style(dpg.mvStyleVar_WindowBorderSize, 1.0)
        dpg.add_theme_style(dpg.mvStyleVar_TabBorderSize,    1.5)

        dpg.add_theme_style(dpg.mvStyleVar_FrameRounding,    3.0)
        dpg.add_theme_style(dpg.mvStyleVar_TabRounding,      3.0)
        dpg.add_theme_style(dpg.mvStyleVar_GrabRounding,     3.0)

        dpg.add_theme_color(dpg.mvThemeCol_Text,             (255, 255, 255, 255))

dpg.bind_theme(global_theme)
W, H = 1400, 860
# === START MENU AND INITIALIZATION ===
def _on_start_clicked():
    chosen_label = dpg.get_value("start_period_combo")
    period_state["period"] = chosen_label
    if dpg.does_item_exist("period_combo"):
        dpg.set_value("period_combo", chosen_label)
    dpg.configure_item("start_win",  show=False)
    dpg.configure_item("start_card", show=False)
    dpg.configure_item("main_win",   show=True)
    dpg.set_primary_window("main_win", True)
    _start_run()

_startup_periods = discover_periods()

SW, SH = 480, 320
cx = 420
cy = 160

with dpg.window(label="", tag="start_win", no_title_bar=True, no_move=True,
                no_resize=True, no_scrollbar=True, no_scroll_with_mouse=True,
                pos=(0, 0), width=W, height=H):
    with dpg.drawlist(width=W, height=H-219):
        dpg.draw_rectangle((0, 0), (W, H),
                           fill=(12, 17, 30, 255),
                           color=(12, 17, 30, 255))
        dpg.draw_rectangle((cx - 2, cy - 2), (cx + SW + 2, cy + SH + 2),
                           fill=(12, 20, 40, 255),
                           color=(50, 120, 200, 255),
                           thickness=2.0,
                           rounding=6.0)
with dpg.window(label="", tag="start_card", no_title_bar=True, no_move=True,
                no_resize=True, no_scrollbar=True, no_scroll_with_mouse=True,
                no_background=True, pos=(cx, cy), width=SW, height=SH):
    dpg.add_spacer(height=20)
    with dpg.group(horizontal=True):
        dpg.add_spacer(width=24)
        with dpg.group():
            dpg.add_text("Combined Visayas Grid", color=(100, 190, 255))
            dpg.add_text("Power Flow + OPF Simulator", color=(140, 140, 160))
    dpg.add_spacer(height=8)
    with dpg.group(horizontal=True):
        dpg.add_spacer(width=24)
        dpg.add_separator()
    dpg.add_spacer(height=16)
    with dpg.group(horizontal=True):
        dpg.add_spacer(width=24)
        dpg.add_text("Select Time Period:", color=(190, 190, 190))
    dpg.add_spacer(height=6)
    with dpg.group(horizontal=True):
        dpg.add_spacer(width=24)
        dpg.add_combo(_startup_periods, tag="start_period_combo",
                      default_value=_startup_periods[0] if _startup_periods else "Base",
                      width=SW - 52)
    dpg.add_spacer(height=4)
    with dpg.group(horizontal=True):
        dpg.add_spacer(width=24)
        dpg.add_text("Base = default data.  Other entries = timestamped folders.",
                     color=(100, 100, 120))
    dpg.add_spacer(height=20)
    with dpg.group(horizontal=True):
        dpg.add_spacer(width=24)
        dpg.add_button(label="  Start Simulation  ", tag="start_btn",
                       width=SW - 52, height=36, callback=_on_start_clicked)
    dpg.add_spacer(height=12)
    with dpg.group(horizontal=True):
        dpg.add_spacer(width=24)
        dpg.add_text("LEYTE-SAMAR  ·  CEBU  ·  NEGROS  ·  BOHOL  ·  PANAY",
                     color=(70, 95, 120))


# === MAIN WINDOW UI DEFINITION ===
with dpg.window(label="", tag="main_win", no_title_bar=True, no_move=True,
                no_resize=True, no_scrollbar=True, pos=(0, 0), width=W, height=H, show=False):

    with dpg.group(horizontal=True):
        dpg.add_text("Combined Visayas Grid", color=(255,255,255))
        dpg.add_spacer(width=12)
        dpg.add_text("Filter:", color=(160,160,160))
        dpg.add_combo(["All"] + SUBGRIDS, tag="sg_filter_combo", default_value="All",
                      width=140, callback=on_filter_change)
        dpg.add_spacer(width=8)
        dpg.add_text("Period:", color=(160,160,160))
        dpg.add_combo(discover_periods(), tag="period_combo", default_value="Base",
                      width=150, callback=lambda s,a: _on_period_change(a))
        
        dpg.add_spacer(width=12)

        dpg.add_button(label="Run Simulation", tag="run_btn", width=130, height=26,
                       callback=lambda: _start_run())
        
        dpg.add_button(label="Open Map", tag="map_btn", width=90, height=26,
                       enabled=False,
                       callback=lambda: webbrowser.open(f"http://localhost:{MAP_PORT}/{MAP_FILE}"))
        dpg.add_button(label="Open SLD", tag="sld_btn", width=90, height=26,
                       enabled=False,
                       callback=lambda: webbrowser.open(f"http://localhost:{MAP_PORT}/{SLD_FILE}"))
        dpg.add_button(label="Pin as Baseline", tag="pin_btn", width=120, height=26,
                       callback=lambda: _pin_as_baseline())
        dpg.add_button(label="Reset to Base Case", tag="reset_override_btn", width=150, height=26,
                       callback=lambda: reset_overrides())

    dpg.add_spacer(height=4)
    with dpg.group(horizontal=True):
        dpg.add_text("", tag="status_text", color=(160,220,160))
        dpg.add_spacer(width=16)
        dpg.add_text("", tag="mod_status", color=(220,180,60,255))
    dpg.add_separator()
    dpg.add_spacer(height=4)

    with dpg.child_window(width=-1, height=-1, border=False):

        with dpg.tab_bar(reorderable=True, tag="main_tab_bar",
                         callback=_on_tab_change):  
            with dpg.tab(label="Bus Voltages & LMP"):
                with dpg.child_window(height=-1, horizontal_scrollbar=True):

                    dpg.add_table(tag="bus_table", header_row=True, freeze_rows=1,
                        resizable=True, borders_innerH=True, borders_outerH=True,
                        borders_innerV=True, borders_outerV=True,
                        scrollY=True, scrollX=True, height=-1,
                        policy=dpg.mvTable_SizingFixedFit, no_pad_outerX=True)

            with dpg.tab(label="Line Loading"):
                with dpg.child_window(height=-1, horizontal_scrollbar=True):

                    dpg.add_table(tag="line_table", header_row=True, freeze_rows=1,
                        resizable=True, borders_innerH=True, borders_outerH=True,
                        borders_innerV=True, borders_outerV=True,
                        scrollY=True, scrollX=True, height=-1,
                        policy=dpg.mvTable_SizingFixedFit, no_pad_outerX=True)

            with dpg.tab(label="Generator Dispatch"):
                with dpg.child_window(height=-1, horizontal_scrollbar=True):

                    dpg.add_table(tag="gen_table", header_row=True, freeze_rows=1,
                        resizable=True, borders_innerH=True, borders_outerH=True,
                        borders_innerV=True, borders_outerV=True,
                        scrollY=True, scrollX=True, height=-1,
                        policy=dpg.mvTable_SizingFixedFit, no_pad_outerX=True)

            with dpg.tab(label="Load LMP"):
                with dpg.child_window(height=-1, horizontal_scrollbar=True):

                    dpg.add_table(tag="load_table", header_row=True, freeze_rows=1,
                        resizable=True, borders_innerH=True, borders_outerH=True,
                        borders_innerV=True, borders_outerV=True,
                        scrollY=True, scrollX=True, height=-1,
                        policy=dpg.mvTable_SizingFixedFit, no_pad_outerX=True)

            with dpg.tab(label="Modify Network", tag="tab_modify"):
                with dpg.child_window(height=-1, horizontal_scrollbar=True):

                    dpg.add_spacer(height=6)
                    with dpg.group(horizontal=True):
                        dpg.add_text("Subgrid:", color=(160,160,160))
                        dpg.add_combo(SUBGRIDS_WITH_GLOBAL, tag="upload_sg_combo",
                                        default_value=SUBGRIDS[0], width=150,
                                        callback=on_upload_sg_change)
                        dpg.add_spacer(width=16)
                        dpg.add_text("Component:", color=(160,160,160))
                        dpg.add_combo(ALL_UPLOADABLE, tag="upload_comp_combo",
                                      default_value=ALL_UPLOADABLE[0], width=160,
                                      callback=on_upload_comp_change)
                        dpg.add_spacer(width=16)
                        dpg.add_button(label="Upload CSV", width=110, height=28,
                                       callback=open_file_dialog)
                        dpg.add_spacer(width=16)
                        dpg.add_button(label="Reset to Default", width=130, height=28,
                                       callback=lambda: reset_overrides())
                    dpg.add_spacer(height=6)
                    dpg.add_text("Refer to User Manual for instructions on how to modify the system", color=(200,200,255))
                    dpg.add_separator(); dpg.add_spacer(height=3)
                    with dpg.group(horizontal=True):
                        dpg.add_text("Active overrides:", color=(200,200,255))
                        dpg.add_text("None", tag="active_overrides_label", color=(220,200,80,255))
                    dpg.add_separator(); dpg.add_spacer(height=2)
                    with dpg.group(horizontal=True):
                        dpg.add_text("Active overrides (session only):", color=(200,200,255))
                        dpg.add_text("None", tag="active_overrides", color=(220,200,80,255))
                    dpg.add_separator()
                    dpg.add_text("", tag="existing_info", color=(200,200,120))
                    dpg.add_spacer(height=2)
                    dpg.add_text("", tag="upload_status", color=(120,220,120))
                    dpg.add_spacer(height=4)
                    dpg.add_text("Default file preview:", color=(200,200,255))
                    dpg.add_spacer(height=4)
                    with dpg.child_window(height=-1, horizontal_scrollbar=True):

                        dpg.add_table(tag="preview_table", header_row=True,
                            resizable=True, borders_innerH=True, borders_outerH=True,
                            borders_innerV=True, borders_outerV=True,
                            scrollY=True, scrollX=True, height=-1,
                            policy=dpg.mvTable_SizingFixedFit, no_pad_outerX=True)

            with dpg.tab(label="Line Outages"):
                with dpg.child_window(height=-1, horizontal_scrollbar=True):

                    dpg.add_spacer(height=6)
                    dpg.add_text("Select lines to trip (set out-of-service). Run simulation to apply.",
                                 color=(200,200,255))
                    dpg.add_text("Outaged lines appear black on the map and grey on the SLD.",
                                 color=(160,160,160))
                    dpg.add_spacer(height=6)
                    with dpg.group(horizontal=True):
                        dpg.add_text("Search:", color=(160,160,160))
                        dpg.add_input_text(tag="outage_search", width=340,
                                           hint="type line name...",
                                           callback=_on_outage_search)
                        dpg.add_spacer(width=10)
                        dpg.add_text("Subgrid:", color=(160,160,160))
                        dpg.add_combo(["All"] + SUBGRIDS, tag="outage_sg_filter",
                                      default_value="All", width=140,
                                      callback=lambda s,a: _refresh_outage_list())
                        dpg.add_spacer(width=16)
                        dpg.add_button(label="Clear All Outages", width=140, height=26,
                                       callback=lambda: [outage_state.clear(),
                                                         _refresh_outage_list()])
                        dpg.add_spacer(width=16)
                        dpg.add_button(label="Refresh List", width=110, height=26,
                                       callback=lambda: _refresh_outage_list())
                    dpg.add_spacer(height=4)
                    dpg.add_text("", tag="outage_status", color=(220,180,80,255))
                    dpg.add_separator()
                    dpg.add_spacer(height=4)
                    with dpg.child_window(tag="outage_list_container",
                                          height=-1, horizontal_scrollbar=False,
                                          border=False):
                        dpg.add_text("Run simulation first to load line list.",
                                     color=(160,160,160))
            
            with dpg.tab(label="Comparison"):
                with dpg.child_window(height=-1, horizontal_scrollbar=True):

                    dpg.add_spacer(height=4)
                    dpg.add_text("Compare any two runs across periods, overrides, or outages.", color=(160,160,160))
                    dpg.add_text("1. Run any period/scenario  2. Click 'Pin as Baseline'  3. Run another  4. View diff here.", color=(160,160,160))
                    dpg.add_text("Green = improved vs baseline,  Red = worsened.", color=(160,160,160))
                    dpg.add_spacer(height=4)
                    dpg.add_text("Baseline:  None\nModified:  None",
                                 tag="cmp_header_text", color=(220,200,80,255))
                    dpg.add_spacer(height=6)
                    with dpg.tab_bar():
                        with dpg.tab(label="Bus Voltages & LMP"):
                            with dpg.child_window(height=-1, horizontal_scrollbar=True):
                                dpg.add_table(tag="cmp_bus_table", header_row=True, freeze_rows=1,
                                    resizable=True, borders_innerH=True, borders_outerH=True,
                                    borders_innerV=True, borders_outerV=True,
                                    scrollY=True, scrollX=True, height=-1,
                                    policy=dpg.mvTable_SizingFixedFit, no_pad_outerX=True)
                        with dpg.tab(label="Line Loading"):
                            with dpg.child_window(height=-1, horizontal_scrollbar=True):
                                dpg.add_table(tag="cmp_line_table", header_row=True, freeze_rows=1,
                                    resizable=True, borders_innerH=True, borders_outerH=True,
                                    borders_innerV=True, borders_outerV=True,
                                    scrollY=True, scrollX=True, height=-1,
                                    policy=dpg.mvTable_SizingFixedFit, no_pad_outerX=True)
                        with dpg.tab(label="Generator Dispatch"):
                            with dpg.child_window(height=-1, horizontal_scrollbar=True):
                                dpg.add_table(tag="cmp_gen_table", header_row=True, freeze_rows=1,
                                    resizable=True, borders_innerH=True, borders_outerH=True,
                                    borders_innerV=True, borders_outerV=True,
                                    scrollY=True, scrollX=True, height=-1,
                                    policy=dpg.mvTable_SizingFixedFit, no_pad_outerX=True)
                        with dpg.tab(label="Load LMP"):
                            with dpg.child_window(height=-1, horizontal_scrollbar=True):
                                dpg.add_table(tag="cmp_load_table", header_row=True, freeze_rows=1,
                                    resizable=True, borders_innerH=True, borders_outerH=True,
                                    borders_innerV=True, borders_outerV=True,
                                    scrollY=True, scrollX=True, height=-1,
                                    policy=dpg.mvTable_SizingFixedFit, no_pad_outerX=True)
                                
            with dpg.tab(label="Export"):
                with dpg.child_window(height=-1, horizontal_scrollbar=False):
                    dpg.add_spacer(height=8)
                    dpg.add_text("Export Results", color=(100,190,255))
                    dpg.add_text("Choose a destination folder when the file dialog opens.",
                                 color=(160,160,160))
                    dpg.add_spacer(height=10)
                    with dpg.group(horizontal=True):
                        dpg.add_text("Filename prefix:", color=(160,160,160))
                        dpg.add_input_text(tag="export_prefix", width=280,
                                           hint="e.g. peak_load_scenario",
                                           default_value="visayas")
                    dpg.add_spacer(height=10)
                    dpg.add_separator()
                    dpg.add_spacer(height=8)

                    dpg.add_text("Current Case -> 4 sheets: Bus, Lines, Generators, Loads",
                                 color=(200,200,255))
                    dpg.add_text("Exports all output tables into a single Excel file.",
                                 color=(160,160,160))
                    dpg.add_spacer(height=4)
                    dpg.add_button(label="Export Full Current Case",
                                   width=220, height=28,
                                   callback=lambda: export_current_case())
                    dpg.add_spacer(height=10)
                    dpg.add_separator()
                    dpg.add_spacer(height=8)

                    dpg.add_text("Comparison Report  ->  4 sheets",
                                 color=(200,200,255))
                    dpg.add_text("Exports baseline vs modified runs for all tables.",
                                 color=(160,160,160))
                    dpg.add_text("Requires a pinned baseline and a modified run.",
                                 color=(180,100,100))
                    dpg.add_spacer(height=4)
                    dpg.add_button(label="Export Comparison Report",
                                   width=220, height=28,
                                   callback=lambda: export_comparison())
                    dpg.add_spacer(height=10)
                    dpg.add_separator()
                    dpg.add_spacer(height=8)

                    dpg.add_text("Map & SLD Snapshots", color=(200,200,255))
                    dpg.add_text("Copies the latest map and SLD HTML files to a folder you choose.",
                                 color=(160,160,160))
                    dpg.add_spacer(height=4)
                    dpg.add_button(label="Export Map & SLD Snapshots",
                                   width=220, height=28,
                                   callback=lambda: export_visuals())
                    dpg.add_spacer(height=10)
                    dpg.add_separator()
                    dpg.add_spacer(height=8)
                    dpg.add_text("", tag="export_status", color=(120,220,120))
            with dpg.tab(label="Settings"):
                with dpg.child_window(height=-1, horizontal_scrollbar=False):
                    dpg.add_spacer(height=6)
                    dpg.add_text("Simulation Constraints", color=(100, 190, 255))
                    dpg.add_text("Changes apply on the next Run. Reset restores all defaults.",
                                color=(140, 140, 160))
                    dpg.add_separator()
                    dpg.add_spacer(height=6)

                    dpg.add_text("Voltage Limits", color=(200, 200, 255))
                    with dpg.group(horizontal=True):
                        dpg.add_text("Min Vm (pu):", color=(160, 160, 160))
                        dpg.add_spacer(width=8)
                        dpg.add_input_float(tag="c_vm_min", default_value=CONSTRAINT_DEFAULTS["vm_min_pu"],
                                            width=120, step=0.01, format="%.3f",
                                            callback=lambda s,a: user_constraints.update({"vm_min_pu": a}))
                    with dpg.group(horizontal=True):
                        dpg.add_text("Max Vm (pu):", color=(160, 160, 160))
                        dpg.add_spacer(width=8)
                        dpg.add_input_float(tag="c_vm_max", default_value=CONSTRAINT_DEFAULTS["vm_max_pu"],
                                            width=120, step=0.01, format="%.3f",
                                            callback=lambda s,a: user_constraints.update({"vm_max_pu": a}))

                    dpg.add_spacer(height=8)
                    dpg.add_separator()
                    dpg.add_text("Power Flow (Newton-Raphson)", color=(200, 200, 255))
                    with dpg.group(horizontal=True):
                        dpg.add_text("Max Iterations:", color=(160, 160, 160))
                        dpg.add_spacer(width=8)
                        dpg.add_input_int(tag="c_nr_iter",
                                        default_value=int(CONSTRAINT_DEFAULTS["nr_max_iteration"]),
                                        width=120, step=1,
                                        callback=lambda s,a: user_constraints.update({"nr_max_iteration": a}))

                    dpg.add_spacer(height=8)
                    dpg.add_separator()

                    dpg.add_text("Optimal Power Flow", color=(200, 200, 255))
                    with dpg.group(horizontal=True):
                        dpg.add_text("Max Iterations:", color=(160, 160, 160))
                        dpg.add_spacer(width=8)
                        dpg.add_input_int(tag="c_opf_iter",
                                        default_value=int(CONSTRAINT_DEFAULTS["opf_max_iteration"]),
                                        width=120, step=1,
                                        callback=lambda s,a: user_constraints.update({"opf_max_iteration": a}))

                    dpg.add_spacer(height=8)
                    dpg.add_separator()

                    dpg.add_text("Tie Line Parameters", color=(200, 200, 255))
                    with dpg.group(horizontal=True):
                        dpg.add_text("R (ohm/km):", color=(160, 160, 160))
                        dpg.add_spacer(width=8)
                        dpg.add_input_float(tag="c_tie_r", default_value=CONSTRAINT_DEFAULTS["tie_r_ohm_per_km"],
                                            width=120, step=0.001, format="%.4f",
                                            callback=lambda s,a: user_constraints.update({"tie_r_ohm_per_km": a}))
                    with dpg.group(horizontal=True):
                        dpg.add_text("X (ohm/km):", color=(160, 160, 160))
                        dpg.add_spacer(width=8)
                        dpg.add_input_float(tag="c_tie_x", default_value=CONSTRAINT_DEFAULTS["tie_x_ohm_per_km"],
                                            width=120, step=0.001, format="%.4f",
                                            callback=lambda s,a: user_constraints.update({"tie_x_ohm_per_km": a}))
                    with dpg.group(horizontal=True):
                        dpg.add_text("C (nF/km):", color=(160, 160, 160))
                        dpg.add_spacer(width=8)
                        dpg.add_input_float(tag="c_tie_c", default_value=CONSTRAINT_DEFAULTS["tie_c_nf_per_km"],
                                            width=120, step=1.0, format="%.1f",
                                            callback=lambda s,a: user_constraints.update({"tie_c_nf_per_km": a}))

                    dpg.add_spacer(height=12)
                    dpg.add_separator()


                    dpg.add_spacer(height=6)
                    dpg.add_button(label="Reset All to Defaults", width=180, height=28,
                                callback=reset_constraints)
                    dpg.add_spacer(height=4)
                    dpg.add_text("", tag="constraints_status", color=(220, 180, 80, 255))


# === DIALOGS AND POPUPS ===
with dpg.file_dialog(tag="upload_file_dialog", directory_selector=False,
                     show=False, callback=handle_upload, width=680, height=420, modal=True):
    dpg.add_file_extension(".csv", color=(100,220,100,255), custom_text="[CSV]")
    dpg.add_file_extension(".*")
    
def _close_error_win():
    state["error"] = None
    dpg.configure_item("error_win", show=False)

with dpg.window(label="Simulation Error", tag="error_win", modal=True,
                show=False, width=720, height=480, pos=(340,190)):
    dpg.add_input_text(tag="error_text", multiline=True, readonly=True,
                       width=700, height=410)
    dpg.add_button(label="Close", width=120, height=28, callback=_close_error_win)

with dpg.value_registry():
    dpg.add_int_value(tag="tables_filled", default_value=0)
    dpg.add_int_value(tag="comparison_filled", default_value=0) 

_update_existing_info()
refresh_period_list()


# === RUNTIME CALLBACKS AND EXPORT ===
def _on_period_change(period_name):
    period_state["period"] = period_name
    mem_overrides.clear()
    comparison["modified_net"]   = None
    comparison["has_modified"]   = False
    comparison["modified_label"] = ""
    state["net"]  = None
    state["done"] = False
    dpg.set_value("tables_filled",   0)
    dpg.set_value("comparison_filled", 0)
    if dpg.does_item_exist("mod_status"):
        dpg.set_value("mod_status", "")
    dpg.set_value("status_text", f"Period changed to '{period_name}' — click Run to load.")

    upload_state["subgrid"]    = SUBGRIDS[0]
    upload_state["component"]  = ALL_UPLOADABLE[0]
    if dpg.does_item_exist("upload_sg_combo"):
        dpg.set_value("upload_sg_combo",   SUBGRIDS[0])
    if dpg.does_item_exist("upload_comp_combo"):
        dpg.set_value("upload_comp_combo", ALL_UPLOADABLE[0])

    _update_existing_info()   
    _update_comparison_header()
    _refresh_mod_status()

def _update_comparison_header():
    if not dpg.does_item_exist("cmp_header_text"):
        return
    b = comparison["baseline_label"] or "None"
    m = comparison["modified_label"] or "None (run simulation again)"
    dpg.set_value("cmp_header_text",
                  f"Baseline:  {b}\nModified:  {m}")
    
def _pin_as_baseline():
    if state["net"] is None or state["running"]:
        dpg.set_value("status_text", "Run simulation first before pinning baseline.")
        return
    if not state.get("done", False):
        dpg.set_value("status_text", "Wait for simulation to finish before pinning baseline.")
        return
    run_label = state.get("last_run_label", period_state["period"])
    comparison["baseline_net"]   = state["net"]
    comparison["has_baseline"]   = True
    comparison["baseline_label"] = run_label
    comparison["has_modified"]   = False
    comparison["modified_net"]   = None
    comparison["modified_label"] = ""
    dpg.set_value("comparison_filled", 0)
    for tag in ["cmp_bus_table","cmp_line_table","cmp_gen_table","cmp_load_table"]:
        if dpg.does_item_exist(tag):
            dpg.delete_item(tag, children_only=True)
    _update_comparison_header()
    dpg.set_value("status_text", f"Baseline pinned: {run_label}")
    _refresh_mod_status()

def _export_filename(suffix):
    import re, datetime
    raw    = dpg.get_value("export_prefix").strip() if dpg.does_item_exist("export_prefix") else ""
    prefix = re.sub(r"[^\w\-]", "_", raw) if raw else "visayas"
    period = re.sub(r"[^\w\-]", period_state["period"], "")
    ts     = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    ov_tag = f"_ov{len(mem_overrides)}" if mem_overrides else ""
    ot_tag = f"_out{len(outage_state)}" if outage_state else ""
    return f"{prefix}_{period}{ov_tag}{ot_tag}_{ts}_{suffix}.xlsx"

def _set_export_status(msg, ok=True):
    color = (120, 220, 120) if ok else (220, 100, 100)
    if dpg.does_item_exist("export_status"):
        dpg.configure_item("export_status", color=color)
        dpg.set_value("export_status", msg)
    print(f"[EXPORT] {msg}")

def _write_info_sheet(wb, net):
    import datetime
    ws = wb.create_sheet("Run Info")
    ws.append(["Parameter", "Value"])
    ws.append(["Export Time",     datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws.append(["Period",          period_state["period"]])
    ws.append(["Active Outages",  ", ".join(sorted(outage_state)) or "None"])
    ws.append(["Active Overrides",", ".join(f"{s}/{c}" for s,c in mem_overrides) or "None"])
    ws.append(["Total Buses",     len(net.bus)])
    ws.append(["Total Lines",     len(net.line)])
    ws.append(["Total Generators",len(net.gen)])
    ws.append(["Total Loads",     len(net.load)])
    ws.append(["Baseline",        comparison.get("baseline_label","") or "None"])
    ws.append(["Modified",        comparison.get("modified_label","") or "None"])

def _do_export_current_case(sender, app_data):
    path = app_data.get("file_path_name", "")
    if not path: return
    if not path.endswith(".xlsx"): path += ".xlsx"
    try:
        import openpyxl
        net     = state["net"]
        b2s     = get_b2s()
        gf      = get_gen_folder()
        tie_set = {t[3] for t in TIE_LINES}
        eb      = set(net.ext_grid["bus"].values)
        gb      = set(net.gen["bus"].values)
        def btype(i): return "SLACK" if i in eb else ("PV" if i in gb else "PQ")

        wb = openpyxl.Workbook()

        ws = wb.active
        ws.title = "Bus Voltages & LMP"
        ws.append(["Bus Name","Subgrid","Type","Vn (kV)","Vm (pu)",
                   "Va (deg)","P (MW)","Q (Mvar)","LMP (PHP/MWh)"])
        for bidx, brow in net.bus.iterrows():
            sg  = b2s.get(bidx, "")
            res = net.res_bus.loc[bidx]
            lmp = round(float(res["lam_p"]), 4) if "lam_p" in res.index else ""
            ws.append([brow["name"], sg, btype(bidx), int(brow["vn_kv"]),
                       round(float(res["vm_pu"]), 6),
                       round(float(res["va_degree"]), 4),
                       round(float(res["p_mw"]), 4),
                       round(float(res["q_mvar"]), 4), lmp])

        ws2 = wb.create_sheet("Line Loading")
        ws2.append(["Line Name","Type","Subgrid","From Bus","To Bus","Vn (kV)",
                    "Loading (%)","P_from (MW)","Q_from (Mvar)",
                    "P_to (MW)","Q_to (Mvar)","I_from (kA)","Outaged"])
        for lidx, row in net.line.iterrows():
            fn     = net.bus.loc[row["from_bus"], "name"]
            tn     = net.bus.loc[row["to_bus"],   "name"]
            vn     = float(net.bus.loc[row["from_bus"], "vn_kv"])
            is_tie = row["name"] in tie_set
            sg     = b2s.get(row["from_bus"], "Tie") if not is_tie else "Tie"
            res    = net.res_line.loc[lidx]
            lt     = "Tie" if is_tie else ("HV" if vn >= 110 else "MV")
            ws2.append([row["name"], lt, sg, fn, tn, int(vn),
                        round(float(res["loading_percent"]), 3),
                        round(float(res["p_from_mw"]), 4),
                        round(float(res["q_from_mvar"]), 4),
                        round(float(res["p_to_mw"]), 4),
                        round(float(res["q_to_mvar"]), 4),
                        round(float(res["i_from_ka"]), 6),
                        "Yes" if row["name"] in outage_state else "No"])

        ws3 = wb.create_sheet("Generator Dispatch")
        ws3.append(["Generator","Plant Type","Subgrid","Bus",
                    "Dispatch (MW)","Q (Mvar)","Cost (PHP/MWh)","LMP (PHP/MWh)"])
        for i in net.gen.index:
            gn  = net.gen.loc[i, "name"]
            res = net.res_gen.loc[i]
            bi  = net.gen.loc[i, "bus"]
            if gn in {h["name"] for h in HVDC_CONNECTIONS if not h["is_slack"]}:
                sg = next(h["subgrid"] for h in HVDC_CONNECTIONS if h["name"] == gn)
            else:
                sg = gf.get(gn, "")
            bus_name = net.bus.loc[bi, "name"] if bi in net.bus.index else str(bi)
            lmp = float(net.res_bus.loc[bi, "lam_p"]) if "lam_p" in net.res_bus.columns and bi in net.res_bus.index else ""
            ws3.append([gn, _get_plant_type(gn), sg, bus_name,
                        round(float(res["p_mw"]), 4),
                        round(float(res["q_mvar"]), 4),
                        round(_get_gen_cost(gn), 4),
                        round(lmp, 4) if lmp != "" else ""])

        ws4 = wb.create_sheet("Load LMP")
        ws4.append(["Load","Subgrid","Bus","P (MW)","LMP (PHP/MWh)"])
        for i in net.load.index:
            bi       = net.load.loc[i, "bus"]
            sg       = b2s.get(bi, "")
            res      = net.res_load.loc[i]
            bus_name = net.bus.loc[bi, "name"] if bi in net.bus.index else str(bi)
            lmp      = float(net.res_bus.loc[bi, "lam_p"]) if "lam_p" in net.res_bus.columns else ""
            ws4.append([net.load.loc[i, "name"], sg, bus_name,
                        round(float(res["p_mw"]), 4),
                        round(lmp, 4) if lmp != "" else ""])

        _write_info_sheet(wb, net)
        wb.save(path)
        _set_export_status(f"Saved -> {os.path.basename(path)}")

    except Exception:
        _set_export_status(f"Export failed: {traceback.format_exc()}", ok=False)

def export_current_case():
    if state["net"] is None:
        _set_export_status("No results — run simulation first.", ok=False)
        return
    import re, datetime
    raw    = dpg.get_value("export_prefix").strip() if dpg.does_item_exist("export_prefix") else ""
    prefix = re.sub(r"[^\w\-]", "_", raw) if raw else "visayas"
    period = re.sub(r"[^\w\-]", "_", period_state["period"])
    ts     = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    ov_tag = f"_ov{len(mem_overrides)}" if mem_overrides else ""
    ot_tag = f"_out{len(outage_state)}" if outage_state else ""
    default = f"{prefix}_{period}{ov_tag}{ot_tag}_{ts}_results.xlsx"
    dpg.configure_item("export_current_dialog", default_filename=default)
    dpg.show_item("export_current_dialog")

def _do_export_comparison(sender, app_data):
    path = app_data.get("file_path_name", "")
    if not path: return
    if not path.endswith(".xlsx"): path += ".xlsx"
    try:
        import openpyxl
        bnet    = comparison["baseline_net"]
        mnet    = comparison["modified_net"]
        b2s     = {ci: f for (f,_), ci in state["bus_map"].items()}
        gf      = get_gen_folder()
        tie_set = {t[3] for t in TIE_LINES}
        eb      = set(bnet.ext_grid["bus"].values)
        gb      = set(bnet.gen["bus"].values)
        def btype(i): return "SLACK" if i in eb else ("PV" if i in gb else "PQ")

        wb = openpyxl.Workbook()

        ws = wb.active
        ws.title = "Bus Voltages & LMP"
        ws.append(["Bus Name","Subgrid","Type","Vn (kV)",
                   "Vm base","Vm mod","Δ Vm",
                   "Va base","Va mod","Δ Va",
                   "P base(MW)","P mod(MW)","Δ P",
                   "LMP base","LMP mod","Δ LMP"])
        m_bus = {n: i for i, n in mnet.bus["name"].items()}
        for bidx_b, brow_b in bnet.bus.iterrows():
            midx = m_bus.get(brow_b["name"])
            if midx is None or midx not in mnet.res_bus.index: continue
            br    = bnet.res_bus.loc[bidx_b]
            mr    = mnet.res_bus.loc[midx]
            b_lmp = float(br["lam_p"]) if "lam_p" in br.index else 0.0
            m_lmp = float(mr["lam_p"]) if "lam_p" in mr.index else 0.0
            ws.append([brow_b["name"], b2s.get(bidx_b,""), btype(bidx_b),
                       int(brow_b["vn_kv"]),
                       round(float(br["vm_pu"]),6), round(float(mr["vm_pu"]),6),
                       round(float(mr["vm_pu"])-float(br["vm_pu"]),6),
                       round(float(br["va_degree"]),4), round(float(mr["va_degree"]),4),
                       round(float(mr["va_degree"])-float(br["va_degree"]),4),
                       round(float(br["p_mw"]),4), round(float(mr["p_mw"]),4),
                       round(float(mr["p_mw"])-float(br["p_mw"]),4),
                       round(b_lmp,4), round(m_lmp,4), round(m_lmp-b_lmp,4)])

        ws2 = wb.create_sheet("Line Loading")
        ws2.append(["Line Name","Type","Subgrid","From Bus","To Bus","Vn (kV)",
                    "Loading base(%)","Loading mod(%)","Δ Loading","Status"])
        m_line = {n: i for i, n in mnet.line["name"].items()}
        for lidx_b, lrow_b in bnet.line.iterrows():
            midx   = m_line.get(lrow_b["name"])
            is_tie = lrow_b["name"] in tie_set
            fn     = bnet.bus.loc[lrow_b["from_bus"],"name"] if lrow_b["from_bus"] in bnet.bus.index else "--"
            tn     = bnet.bus.loc[lrow_b["to_bus"],  "name"] if lrow_b["to_bus"]   in bnet.bus.index else "--"
            vn     = float(bnet.bus.loc[lrow_b["from_bus"],"vn_kv"])
            lt     = "Tie" if is_tie else ("HV" if vn >= 110 else "MV")
            sg     = b2s.get(lrow_b["from_bus"],"Tie") if not is_tie else "Tie"
            b_load = round(float(bnet.res_line.loc[lidx_b,"loading_percent"]),3) if lidx_b in bnet.res_line.index else ""
            tripped = lrow_b["name"] in outage_state
            if tripped:
                m_load="TRIPPED"; delta=""; status="Tripped"
            elif midx is not None and midx in mnet.res_line.index:
                m_load = round(float(mnet.res_line.loc[midx,"loading_percent"]),3)
                delta  = round(m_load - b_load, 3) if b_load != "" else ""
                status = ""
            else:
                m_load=""; delta=""; status=""
            ws2.append([lrow_b["name"], lt, sg, fn, tn, int(vn),
                        b_load, m_load, delta, status])

        ws3 = wb.create_sheet("Generator Dispatch")
        ws3.append(["Generator","Plant Type","Subgrid",
                    "Dispatch base(MW)","Dispatch mod(MW)","Δ Dispatch",
                    "LMP base","LMP mod","Δ LMP"])
        m_gen = {n: i for i, n in mnet.gen["name"].items()}
        for gidx_b, grow_b in bnet.gen.iterrows():
            gn   = grow_b["name"]
            midx = m_gen.get(gn)
            if midx is None or midx not in mnet.res_gen.index: continue
            br   = bnet.res_gen.loc[gidx_b]
            mr   = mnet.res_gen.loc[midx]
            bi_b = bnet.gen.loc[gidx_b,"bus"]
            bi_m = mnet.gen.loc[midx,"bus"]
            b_lmp = float(bnet.res_bus.loc[bi_b,"lam_p"]) if "lam_p" in bnet.res_bus.columns and bi_b in bnet.res_bus.index else 0.0
            m_lmp = float(mnet.res_bus.loc[bi_m,"lam_p"]) if "lam_p" in mnet.res_bus.columns and bi_m in mnet.res_bus.index else 0.0
            if gn in {h["name"] for h in HVDC_CONNECTIONS if not h["is_slack"]}:
                sg = next(h["subgrid"] for h in HVDC_CONNECTIONS if h["name"]==gn)
            else:
                sg = gf.get(gn,"")
            ws3.append([gn, _get_plant_type(gn), sg,
                        round(float(br["p_mw"]),4), round(float(mr["p_mw"]),4),
                        round(float(mr["p_mw"])-float(br["p_mw"]),4),
                        round(b_lmp,4), round(m_lmp,4), round(m_lmp-b_lmp,4)])

        ws4 = wb.create_sheet("Load LMP")
        ws4.append(["Load","Subgrid","Bus",
                    "P base(MW)","P mod(MW)","Δ P",
                    "LMP base","LMP mod","Δ LMP"])
        m_load_idx = {n: i for i, n in mnet.load["name"].items()}
        for lidx_b, lrow_b in bnet.load.iterrows():
            ln   = lrow_b["name"]
            midx = m_load_idx.get(ln)
            if midx is None or midx not in mnet.res_load.index: continue
            bi_b = bnet.load.loc[lidx_b,"bus"]
            bi_m = mnet.load.loc[midx,"bus"]
            sg   = b2s.get(bi_b,"")
            bus_name = bnet.bus.loc[bi_b,"name"] if bi_b in bnet.bus.index else str(bi_b)
            br   = bnet.res_load.loc[lidx_b]
            mr   = mnet.res_load.loc[midx]
            b_lmp = float(bnet.res_bus.loc[bi_b,"lam_p"]) if "lam_p" in bnet.res_bus.columns and bi_b in bnet.res_bus.index else 0.0
            m_lmp = float(mnet.res_bus.loc[bi_m,"lam_p"]) if "lam_p" in mnet.res_bus.columns and bi_m in mnet.res_bus.index else 0.0
            ws4.append([ln, sg, bus_name,
                        round(float(br["p_mw"]),4), round(float(mr["p_mw"]),4),
                        round(float(mr["p_mw"])-float(br["p_mw"]),4),
                        round(b_lmp,4), round(m_lmp,4), round(m_lmp-b_lmp,4)])

        _write_info_sheet(wb, bnet)
        wb.save(path)
        _set_export_status(f"Comparison saved → {os.path.basename(path)}")

    except Exception:
        _set_export_status(f"Export failed: {traceback.format_exc()}", ok=False)

def export_comparison():
    if not comparison["has_baseline"] or not comparison["has_modified"]:
        _set_export_status("Pin a baseline and run a modified scenario first.", ok=False)
        return
    import re, datetime
    raw    = dpg.get_value("export_prefix").strip() if dpg.does_item_exist("export_prefix") else ""
    prefix = re.sub(r"[^\w\-]", "_", raw) if raw else "visayas"
    period = re.sub(r"[^\w\-]", "_", period_state["period"])
    ts     = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    default = f"{prefix}_{period}_{ts}_comparison.xlsx"
    dpg.configure_item("export_comparison_dialog", default_filename=default)
    dpg.show_item("export_comparison_dialog")

def _do_export_visuals(sender, app_data):
    import shutil
    folder = app_data.get("file_path_name", "")
    if not folder: return
    import re, datetime
    raw    = dpg.get_value("export_prefix").strip() if dpg.does_item_exist("export_prefix") else ""
    prefix = re.sub(r"[^\w\-]", "_", raw) if raw else "visayas"
    period = re.sub(r"[^\w\-]", "_", period_state["period"])
    ts     = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    copied = []
    for src_name, label in [(MAP_FILE, "map"), (SLD_FILE, "sld")]:
        src = os.path.join(state["base_dir"], src_name)
        if os.path.exists(src):
            ext = os.path.splitext(src_name)[1]
            dst = os.path.join(folder, f"{prefix}_{period}_{ts}_{label}{ext}")
            shutil.copy2(src, dst)
            copied.append(os.path.basename(dst))
    if copied:
        _set_export_status(f"Copied → {', '.join(copied)}")
    else:
        _set_export_status("No map/SLD found — run simulation first.", ok=False)

def export_visuals():
    if not os.path.exists(os.path.join(state["base_dir"], MAP_FILE)) and \
       not os.path.exists(os.path.join(state["base_dir"], SLD_FILE)):
        _set_export_status("No map/SLD found — run simulation first.", ok=False)
        return
    dpg.show_item("export_visuals_dialog")

def export_results(sender, app_data):
    if state["net"] is None:
        dpg.set_value("status_text", "No results to export — run simulation first.")
        return

    path = app_data.get("file_path_name", "")
    if not path:
        return
    if not path.endswith(".xlsx"):
        path += ".xlsx"

    try:
        import openpyxl
        import datetime

        net       = state["net"]
        b2s       = get_b2s()
        gf        = get_gen_folder()
        sg_filter = filter_state["subgrid"]
        tie_set   = {t[3] for t in TIE_LINES}

        eb = set(net.ext_grid["bus"].values)
        gb = set(net.gen["bus"].values)
        def btype(i): return "SLACK" if i in eb else ("PV" if i in gb else "PQ")

        wb = openpyxl.Workbook()

        def make_sheet(wb, title, headers, rows):
            ws = wb.create_sheet(title) if title != wb.active.title else wb.active
            ws.title = title
            ws.append(headers)
            for row in rows:
                ws.append(row)
            return ws

        bus_rows = []
        for bidx, brow in net.bus.iterrows():
            sg = b2s.get(bidx, "")
            if sg_filter != "All" and sg != sg_filter: continue
            res = net.res_bus.loc[bidx]
            lmp = round(float(res["lam_p"]), 2) if "lam_p" in res.index else None
            bus_rows.append([
                brow["name"], sg, btype(bidx), int(brow["vn_kv"]),
                round(float(res["vm_pu"]), 4),
                round(float(res["va_degree"]), 2),
                round(float(res["p_mw"]), 3),
                round(float(res["q_mvar"]), 3),
                lmp,
            ])
        make_sheet(wb, "Bus Voltages & LMP",
                   ["Bus Name","Subgrid","Type","Vn (kV)","Vm (pu)",
                    "Va (deg)","P (MW)","Q (Mvar)","LMP (PHP/MWh)"],
                   bus_rows)

        line_rows = []
        for lidx, row in net.line.iterrows():
            fn     = net.bus.loc[row["from_bus"], "name"]
            tn     = net.bus.loc[row["to_bus"],   "name"]
            vn     = float(net.bus.loc[row["from_bus"], "vn_kv"])
            is_tie = row["name"] in tie_set
            sg     = b2s.get(row["from_bus"], "Tie") if not is_tie else "Tie"
            if sg_filter != "All" and sg != sg_filter: continue
            res = net.res_line.loc[lidx]
            lt  = "Tie" if is_tie else ("HV" if vn >= 110 else "MV")
            line_rows.append([
                row["name"], lt, sg, fn, tn, int(vn),
                round(float(res["loading_percent"]), 1),
                round(float(res["p_from_mw"]), 3),
                round(float(res["q_from_mvar"]), 3),
                round(float(res["i_from_ka"]), 4),
            ])
        make_sheet(wb, "Line Loading",
                   ["Line Name","Type","Subgrid","From Bus","To Bus","Vn (kV)",
                    "Loading (%)","P_from (MW)","Q_from (Mvar)","I_from (kA)"],
                   line_rows)

        gen_rows = []
        for i in net.gen.index:
            gn = net.gen.loc[i, "name"]
            if gn in {h["name"] for h in HVDC_CONNECTIONS if not h["is_slack"]}:
                sg = next(h["subgrid"] for h in HVDC_CONNECTIONS if h["name"] == gn)
            else:
                sg = gf.get(gn, "")
            if sg_filter != "All" and sg != sg_filter: continue
            res      = net.res_gen.loc[i]
            bi       = net.gen.loc[i, "bus"]
            bus_name = net.bus.loc[bi, "name"] if bi in net.bus.index else str(bi)
            cost     = _get_gen_cost(gn)
            ptype    = _get_plant_type(gn)
            lmp      = round(float(net.res_bus.loc[bi, "lam_p"]), 2) \
                       if "lam_p" in net.res_bus.columns and bi in net.res_bus.index else None
            gen_rows.append([
                gn, ptype, sg, bus_name,
                round(float(res["p_mw"]), 3),
                round(float(res["q_mvar"]), 3),
                round(cost, 2), lmp,
            ])
        make_sheet(wb, "Generator Dispatch",
                   ["Generator","Plant Type","Subgrid","Bus",
                    "Dispatch (MW)","Q (Mvar)","Cost (PHP/MWh)","LMP (PHP/MWh)"],
                   gen_rows)

        load_rows = []
        for i in net.load.index:
            bi  = net.load.loc[i, "bus"]
            sg  = b2s.get(bi, "")
            if sg_filter != "All" and sg != sg_filter: continue
            res      = net.res_load.loc[i]
            bus_name = net.bus.loc[bi, "name"] if bi in net.bus.index else str(bi)
            lmp      = round(float(net.res_bus.loc[bi, "lam_p"]), 2) \
                       if "lam_p" in net.res_bus.columns else None
            load_rows.append([
                net.load.loc[i, "name"], sg, bus_name,
                round(float(res["p_mw"]), 3), lmp,
            ])
        make_sheet(wb, "Load LMP",
                   ["Load","Subgrid","Bus","P (MW)","LMP (PHP/MWh)"],
                   load_rows)

        if comparison["has_baseline"] and comparison["has_modified"]:
            bnet = comparison["baseline_net"]
            mnet = comparison["modified_net"]
            b2sb = {ci: f for (f,_), ci in state["bus_map"].items()}
            m_bus = {n: i for i, n in mnet.bus["name"].items()}

            cmp_rows = []
            for bidx_b, brow_b in bnet.bus.iterrows():
                sg = b2sb.get(bidx_b, "")
                if sg_filter != "All" and sg != sg_filter: continue
                midx = m_bus.get(brow_b["name"])
                if midx is None or midx not in mnet.res_bus.index: continue
                br    = bnet.res_bus.loc[bidx_b]
                mr    = mnet.res_bus.loc[midx]
                b_lmp = round(float(br["lam_p"]), 2) if "lam_p" in br.index else None
                m_lmp = round(float(mr["lam_p"]), 2) if "lam_p" in mr.index else None
                cmp_rows.append([
                    brow_b["name"], sg,
                    round(float(br["vm_pu"]), 4), round(float(mr["vm_pu"]), 4),
                    round(float(br["va_degree"]), 2), round(float(mr["va_degree"]), 2),
                    round(float(br["p_mw"]), 3), round(float(mr["p_mw"]), 3),
                    b_lmp, m_lmp,
                ])
            make_sheet(wb, "Comparison - Bus",
                       ["Bus Name","Subgrid",
                        "Vm base","Vm mod","Va base","Va mod",
                        "P base(MW)","P mod(MW)",
                        "LMP base(PHP)","LMP mod(PHP)"],
                       cmp_rows)

        info_rows = [
            ["Export Time",     datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["Period",          period_state["period"]],
            ["Subgrid Filter",  sg_filter],
            ["Active Outages",  ", ".join(sorted(outage_state)) or "None"],
            ["Active Overrides",", ".join(f"{s}/{c}" for s,c in mem_overrides) or "None"],
            ["Total Buses",     len(net.bus)],
            ["Total Lines",     len(net.line)],
            ["Total Gens",      len(net.gen)],
            ["Total Loads",     len(net.load)],
            ["Baseline",        comparison.get("baseline_label","") or "None"],
            ["Modified",        comparison.get("modified_label","") or "None"],
        ]
        make_sheet(wb, "Run Info", ["Parameter","Value"], info_rows)

        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        wb.save(path)
        dpg.set_value("status_text", f"Exported → {path}")
        print(f"[EXPORT] Saved → {path}")

    except Exception as e:
        dpg.set_value("status_text", f"Export failed: {e}")
        print(f"[EXPORT] ERROR: {traceback.format_exc()}")
        
def _start_run():
    if state["running"]: return
    state.update({"running": True, "done": False, "error": None,
                  "map_ready": False, "net": None})
    dpg.set_value("tables_filled", 0)
    dpg.set_value("comparison_filled", 0)
    dpg.configure_item("run_btn", enabled=False, label="Running...")
    dpg.configure_item("map_btn", enabled=False)
    if dpg.does_item_exist("sld_btn"):
        dpg.configure_item("sld_btn", enabled=False)
    threading.Thread(target=run_simulation, daemon=True).start()

_tab_bar_theme_applied   = False
_start_card_theme_applied = False

def update_ui():
    global _tab_bar_theme_applied, _start_card_theme_applied

    if not _start_card_theme_applied and dpg.does_item_exist("start_card"):
        with dpg.theme() as _sct:
            with dpg.theme_component(dpg.mvChildWindow):
                dpg.add_theme_color(dpg.mvThemeCol_ChildBg,  (12, 20, 38, 255))
                dpg.add_theme_color(dpg.mvThemeCol_Border,   (50, 120, 200, 255))
                dpg.add_theme_style(dpg.mvStyleVar_ChildBorderSize, 2.0)
                dpg.add_theme_style(dpg.mvStyleVar_ChildRounding,   6.0)
        dpg.bind_item_theme("start_card", _sct)
        _start_card_theme_applied = True
    dpg.set_value("status_text", state["status"])
    if state["error"] and not dpg.is_item_shown("error_win"):
        dpg.configure_item("error_win", show=True)
        dpg.set_value("error_text", state["error"])

    if state["done"] and not state["running"]:
        dpg.configure_item("run_btn", enabled=True, label="Run Simulation")
        dpg.configure_item("map_btn", enabled=True)
        if dpg.does_item_exist("sld_btn"):
            dpg.configure_item("sld_btn", enabled=True)

        # Fill standard results tables
        if state["net"] is not None and dpg.get_value("tables_filled") == 0:
            fill_all_tables()
            _refresh_outage_list()
            dpg.set_value("tables_filled", 1)

        # Trigger Comparison Tab Fill
        if state["done"] and not state["running"]:

            if (comparison["has_baseline"] and comparison["has_modified"]
                    and dpg.get_value("comparison_filled") == 0):
                try:
                    fill_comparison_tab()
                except Exception as _cmp_exc:
                    import traceback
                _update_comparison_header()   # ← add this at the very end

                dpg.set_value("comparison_filled", 1)
    if not state["done"] and not state["running"]:
        dpg.configure_item("run_btn", enabled=True, label="Run Simulation")
        
with dpg.file_dialog(tag="export_file_dialog", directory_selector=False,
                     show=False, callback=export_results,
                     width=680, height=420, modal=True,
                     default_filename="visayas_results.xlsx"):
    dpg.add_file_extension(".xlsx", color=(100,220,100,255), custom_text="[Excel]")
    dpg.add_file_extension(".*")
with dpg.file_dialog(tag="export_current_dialog", directory_selector=False,
                     show=False, callback=_do_export_current_case,
                     width=700, height=440, modal=True):
    dpg.add_file_extension(".xlsx", color=(100,220,100,255), custom_text="[Excel]")

with dpg.file_dialog(tag="export_comparison_dialog", directory_selector=False,
                     show=False, callback=_do_export_comparison,
                     width=700, height=440, modal=True):
    dpg.add_file_extension(".xlsx", color=(100,220,100,255), custom_text="[Excel]")

with dpg.file_dialog(tag="export_visuals_dialog", directory_selector=True,
                     show=False, callback=_do_export_visuals,
                     width=700, height=440, modal=True):
    pass

start_map_server()
dpg.set_primary_window("start_win", True)
dpg.show_viewport()

while dpg.is_dearpygui_running():
    update_ui()
    dpg.render_dearpygui_frame()

if state["map_server"]:
    state["map_server"].shutdown()
dpg.destroy_context()

#newxSDsdads